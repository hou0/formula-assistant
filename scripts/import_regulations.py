"""
法规数据库导入脚本
将三份 Excel 法规数据导入 SQLite 数据库。

数据来源：
1. 化妆品安全技术规范（2015年版）.xlsx    — 7张表（禁限用/准用组分）
2. 已上市产品原料使用信息（2025）.xlsx     — 已上市产品原料使用量
3. 国际化妆品安全评估数据索引...xlsx       — 国际安全评估数据

使用说明：
   数据库路径自动沿用项目的 userdata/safety.db
   安全运行，重复运行不会产生重复数据（先清空后导入）
"""

import sys
import os

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
sys.stdout.reconfigure(encoding='utf-8')

import database as db  # noqa: E402

# ── 文件路径 ──
DIR_LAWS = r'D:\工作文档\法律法规文件\法律法规文件'
FILE_TECH_SPEC = r'C:\Users\houping\Desktop\化妆品安全技术规范（2015年版.xlsx'
FILE_USAGE_MARKET = os.path.join(DIR_LAWS, '已上市产品原料使用信息 （2025）.xlsx')
FILE_USAGE_INTERNATIONAL = os.path.join(DIR_LAWS, '20250410《国际化妆品安全评估数据索引》收录的部分原料使用信息.xlsx')
FILE_EXPOSURE_SED = r'C:\Users\houping\Desktop\日均使用量和驻留因子查询、SED值、MOS值计算器（Wendy慧）(1).xlsx'

# ── 使用说明文本（安全评估规则） ──
USAGE_RULES_MARKET = """一、本原料使用信息是对我国注册备案有效化妆品中已使用、未收录在《化妆品安全技术规范》《国际权威机构化妆品安全评估索引》中原料使用信息的客观收录。未组织对所列原料的安全性进行系统评价，化妆品注册人、备案人在使用相关原料信息时，应当符合国家有关法律法规、强制性国家标准、技术规范的相关要求，开展化妆品安全评估并承担产品质量安全责任。
二、本原料使用信息收录的原料使用量，可为化妆品安全评估提供参考，化妆品注册人、备案人应当结合产品使用方法和作用部位，对数据的适用性进行评估，正确使用原料使用量。
三、相同作用部位的同一原料，若只有驻留类产品的原料使用量，淋洗类产品可参照驻留类使用。
四、相同使用方法的同一原料，可按照全身、躯干部位、面部（含颈部）、手足、头部、头发、口唇、眼部、指（趾）甲的顺序，后面作用部位可参照前面作用部位的原料使用量，但产品作用部位为眼部且参考其他部位使用量时，需另外评估眼刺激性。其中，口唇、眼部不可参照手足、头部、头发的原料使用量；体毛仅可参照全身或躯干部位的原料使用量；作用部位同时为头部和头发，可参照头部的原料使用量；作用部位同时为面部（含颈部）、眼部和/或口唇，可参照面部（含颈部）的原料使用量，作用部位包括眼部时，需另外评估眼刺激性；对于其他同时用于多个作用部位产品的原料使用量，选择使用相同使用方法的上一级作用部位的使用量。"""

USAGE_RULES_INTERNATIONAL = USAGE_RULES_MARKET  # 内容相同


def safe_str(val):
    """处理 Excel 单元格值，None → 空字符串"""
    if val is None:
        return ''
    return str(val).strip()


def find_data_start(ws, keyword='序号'):
    """找到数据标题行的行号（从1开始）"""
    for r in range(1, min(15, ws.max_row + 1)):
        v = str(ws.cell(row=r, column=1).value or '')
        if keyword in v:
            return r
    return 1


def import_banned_chemical(conn, ws):
    """导入禁用组分 表1"""
    cursor = conn.cursor()
    cursor.execute('DELETE FROM regulation_banned_chemical')
    data_start = find_data_start(ws)
    data_start += 1  # 跳过标题行
    count = 0
    for r in range(data_start, ws.max_row + 1):
        seq = safe_str(ws.cell(row=r, column=1).value)
        name_zh = safe_str(ws.cell(row=r, column=2).value)
        if not seq or not name_zh or not seq.isdigit():
            continue
        name_en = safe_str(ws.cell(row=r, column=3).value)
        cursor.execute(
            'INSERT INTO regulation_banned_chemical (seq, name_zh, name_en) VALUES (?, ?, ?)',
            (seq, name_zh, name_en)
        )
        count += 1
    conn.commit()
    return count


def import_banned_botanical(conn, ws):
    """导入禁用植(动)物组分 表2"""
    cursor = conn.cursor()
    cursor.execute('DELETE FROM regulation_banned_botanical')
    data_start = find_data_start(ws)
    data_start += 1
    count = 0
    for r in range(data_start, ws.max_row + 1):
        seq = safe_str(ws.cell(row=r, column=1).value)
        name_zh = safe_str(ws.cell(row=r, column=2).value)
        if not seq or not seq.isdigit():
            continue
        latin_name = safe_str(ws.cell(row=r, column=3).value)
        # 收集剩余列的附加说明（如果有）
        notes_parts = []
        for c in range(4, ws.max_column + 1):
            v = safe_str(ws.cell(row=r, column=c).value)
            if v:
                notes_parts.append(v)
        notes = '; '.join(notes_parts) if notes_parts else ''
        cursor.execute(
            'INSERT INTO regulation_banned_botanical (seq, name_zh, latin_name, notes) VALUES (?, ?, ?, ?)',
            (seq, name_zh, latin_name, notes)
        )
        count += 1
    conn.commit()
    return count


def import_restricted(conn, ws):
    """导入限用组分 表3"""
    cursor = conn.cursor()
    cursor.execute('DELETE FROM regulation_restricted')

    # 找到真实标题行（包含 INCI 关键词）
    header_row = None
    for r in range(1, 10):
        vals = [str(ws.cell(row=r, column=c).value or '') for c in range(1, ws.max_column + 1)]
        if any('INCI' in v for v in vals):
            header_row = r
            break

    if header_row is None:
        print('  [ERROR] 表3找不到标题行')
        return 0

    data_start = header_row + 1
    count = 0
    current_seq = ''
    for r in range(data_start, ws.max_row + 1):
        # 序号可能为空（续行），携带上一个序号
        seq_raw = safe_str(ws.cell(row=r, column=1).value)
        if seq_raw:
            current_seq = seq_raw
        name_zh = safe_str(ws.cell(row=r, column=2).value)
        name_en = safe_str(ws.cell(row=r, column=3).value)
        inci_name = safe_str(ws.cell(row=r, column=4).value)
        scope_of_use = safe_str(ws.cell(row=r, column=5).value)
        max_conc = safe_str(ws.cell(row=r, column=6).value)
        restrictions = safe_str(ws.cell(row=r, column=7).value)
        label_req = safe_str(ws.cell(row=r, column=8).value)

        # 跳过全空行
        if not name_zh and not name_en and not scope_of_use and not max_conc:
            # 但可能是续行（只有 scope/max_conc 等），不跳过
            if not scope_of_use and not max_conc and not restrictions:
                continue

        cursor.execute(
            'INSERT INTO regulation_restricted '
            '(seq, name_zh, name_en, inci_name, scope_of_use, max_concentration, restrictions, label_requirements) '
            'VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
            (current_seq, name_zh, name_en, inci_name, scope_of_use, max_conc, restrictions, label_req)
        )
        count += 1
    conn.commit()
    return count


def import_allowed_preservative(conn, ws):
    """导入准用防腐剂 表4"""
    cursor = conn.cursor()
    cursor.execute('DELETE FROM regulation_allowed_preservative')
    header_row = None
    for r in range(1, 10):
        vals = [str(ws.cell(row=r, column=c).value or '') for c in range(1, ws.max_column + 1)]
        if any('INCI' in v for v in vals):
            header_row = r
            break
    if header_row is None:
        print('  [ERROR] 表4找不到标题行')
        return 0

    data_start = header_row + 1
    count = 0
    current_seq = ''
    for r in range(data_start, ws.max_row + 1):
        seq_raw = safe_str(ws.cell(row=r, column=1).value)
        if seq_raw:
            current_seq = seq_raw
        name_zh = safe_str(ws.cell(row=r, column=2).value)
        name_en = safe_str(ws.cell(row=r, column=3).value)
        inci_name = safe_str(ws.cell(row=r, column=4).value)
        max_conc = safe_str(ws.cell(row=r, column=5).value)
        scope_cond = safe_str(ws.cell(row=r, column=6).value)
        label_req = safe_str(ws.cell(row=r, column=7).value)

        if not name_zh and not name_en and not max_conc and not scope_cond:
            continue

        cursor.execute(
            'INSERT INTO regulation_allowed_preservative '
            '(seq, name_zh, name_en, inci_name, max_concentration, scope_and_conditions, label_requirements) '
            'VALUES (?, ?, ?, ?, ?, ?, ?)',
            (current_seq, name_zh, name_en, inci_name, max_conc, scope_cond, label_req)
        )
        count += 1
    conn.commit()
    return count


def import_allowed_sunscreen(conn, ws):
    """导入准用防晒剂 表5"""
    cursor = conn.cursor()
    cursor.execute('DELETE FROM regulation_allowed_sunscreen')
    header_row = None
    for r in range(1, 10):
        vals = [str(ws.cell(row=r, column=c).value or '') for c in range(1, ws.max_column + 1)]
        if any('INCI' in v for v in vals):
            header_row = r
            break
    if header_row is None:
        print('  [ERROR] 表5找不到标题行')
        return 0

    data_start = header_row + 1
    count = 0
    current_seq = ''
    for r in range(data_start, ws.max_row + 1):
        seq_raw = safe_str(ws.cell(row=r, column=1).value)
        if seq_raw:
            current_seq = seq_raw
        name_zh = safe_str(ws.cell(row=r, column=2).value)
        name_en = safe_str(ws.cell(row=r, column=3).value)
        inci_name = safe_str(ws.cell(row=r, column=4).value)
        max_conc = safe_str(ws.cell(row=r, column=5).value)
        restrictions = safe_str(ws.cell(row=r, column=6).value)
        label_req = safe_str(ws.cell(row=r, column=7).value)

        if not name_zh and not name_en and not max_conc:
            continue

        cursor.execute(
            'INSERT INTO regulation_allowed_sunscreen '
            '(seq, name_zh, name_en, inci_name, max_concentration, restrictions, label_requirements) '
            'VALUES (?, ?, ?, ?, ?, ?, ?)',
            (current_seq, name_zh, name_en, inci_name, max_conc, restrictions, label_req)
        )
        count += 1
    conn.commit()
    return count


def import_allowed_colorant(conn, ws):
    """导入准用着色剂 表6"""
    cursor = conn.cursor()
    cursor.execute('DELETE FROM regulation_allowed_colorant')
    # 表6的标题行比较特殊，row 3 是主标题，row 4-5 是列说明
    data_start = 6  # 数据从第6行开始
    count = 0
    current_seq = ''
    for r in range(data_start, ws.max_row + 1):
        seq_raw = safe_str(ws.cell(row=r, column=1).value)
        if seq_raw and seq_raw.isdigit():
            current_seq = seq_raw

        name_zh = safe_str(ws.cell(row=r, column=5).value)  # CI中文名
        color_index = safe_str(ws.cell(row=r, column=2).value)
        ci_generic_name = safe_str(ws.cell(row=r, column=3).value)
        color = safe_str(ws.cell(row=r, column=4).value)

        # 使用范围：4类，标记为 '+' 或 '×'
        s1 = 1 if safe_str(ws.cell(row=r, column=6).value) in ('+', '×', '1') else 0
        s2 = 1 if safe_str(ws.cell(row=r, column=7).value) in ('+', '×', '1') else 0
        s3 = 1 if safe_str(ws.cell(row=r, column=8).value) in ('+', '×', '1') else 0
        s4 = 1 if safe_str(ws.cell(row=r, column=9).value) in ('+', '×', '1') else 0

        restrictions = safe_str(ws.cell(row=r, column=10).value)

        if not current_seq and not name_zh:
            # 尝试看是不是注释行
            if '注' in str(ws.cell(row=r, column=1).value or ''):
                break  # 到底部注释了，结束
            continue

        cursor.execute(
            'INSERT INTO regulation_allowed_colorant '
            '(seq, name_zh, color_index, ci_generic_name, color, '
            'scope_1, scope_2, scope_3, scope_4, restrictions) '
            'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (current_seq, name_zh, color_index, ci_generic_name, color,
             s1, s2, s3, s4, restrictions)
        )
        count += 1
    conn.commit()
    return count


def import_allowed_hair_dye(conn, ws):
    """导入准用染发剂 表7"""
    cursor = conn.cursor()
    cursor.execute('DELETE FROM regulation_allowed_hair_dye')
    header_row = None
    for r in range(1, 10):
        vals = [str(ws.cell(row=r, column=c).value or '') for c in range(1, ws.max_column + 1)]
        if any('INCI' in v for v in vals):
            header_row = r
            break
    if header_row is None:
        print('  [ERROR] 表7找不到标题行')
        return 0

    data_start = header_row + 1
    count = 0
    current_seq = ''
    for r in range(data_start, ws.max_row + 1):
        seq_raw = safe_str(ws.cell(row=r, column=1).value)
        if seq_raw:
            current_seq = seq_raw
        name_zh = safe_str(ws.cell(row=r, column=2).value)
        inci_name = safe_str(ws.cell(row=r, column=3).value)
        max_ox = safe_str(ws.cell(row=r, column=4).value)
        max_nonox = safe_str(ws.cell(row=r, column=5).value)
        restrictions = safe_str(ws.cell(row=r, column=6).value)
        label_req = safe_str(ws.cell(row=r, column=7).value)

        if not current_seq and not name_zh:
            if '注' in str(ws.cell(row=r, column=1).value or ''):
                break
            continue

        cursor.execute(
            'INSERT INTO regulation_allowed_hair_dye '
            '(seq, name_zh, inci_name, max_conc_oxidative, max_conc_non_oxidative, '
            'restrictions, label_requirements) '
            'VALUES (?, ?, ?, ?, ?, ?, ?)',
            (current_seq, name_zh, inci_name, max_ox, max_nonox, restrictions, label_req)
        )
        count += 1
    conn.commit()
    return count


def import_usage_data(conn, ws, table_name):
    """导入已上市/国际原料使用信息（共用逻辑）"""
    cursor = conn.cursor()
    if table_name == 'safety_usage_market':
        cursor.execute('DELETE FROM safety_usage_market')
    else:
        cursor.execute('DELETE FROM safety_usage_international')

    header_row = None
    for r in range(1, 10):
        v = str(ws.cell(row=r, column=1).value or '')
        if '序号' in v or '序' in v:
            header_row = r
            break
    if header_row is None:
        print(f'  [ERROR] {table_name}找不到标题行')
        return 0

    data_start = header_row + 1
    count = 0
    current_name_zh = ''
    current_inci = ''
    current_catalog = ''
    for r in range(data_start, ws.max_row + 1):
        seq_raw = safe_str(ws.cell(row=r, column=1).value)
        catalog_seq = safe_str(ws.cell(row=r, column=2).value)
        name_zh = safe_str(ws.cell(row=r, column=3).value)
        inci_name = safe_str(ws.cell(row=r, column=4).value)
        used_part = safe_str(ws.cell(row=r, column=5).value)
        method = safe_str(ws.cell(row=r, column=6).value)
        max_pct = safe_str(ws.cell(row=r, column=7).value)
        remarks = safe_str(ws.cell(row=r, column=8).value)

        # 有序号的是新的原料条目
        if seq_raw:
            current_catalog = catalog_seq
            current_name_zh = name_zh
            current_inci = inci_name
        else:
            # 续行，继承原料名称
            if not name_zh:
                name_zh = current_name_zh
            if not inci_name:
                inci_name = current_inci

        if not current_name_zh and not name_zh and not used_part and not method:
            continue

        cursor.execute(
            f'INSERT INTO {table_name} '
            '(catalog_seq, name_zh, inci_name, used_part, method, max_percent, remarks) '
            'VALUES (?, ?, ?, ?, ?, ?, ?)',
            (current_catalog, name_zh, inci_name, used_part, method, max_pct, remarks)
        )
        count += 1
    conn.commit()
    return count


def import_exposure_daily_usage(conn, ws):
    """导入暴露量数据（日均使用量 & 驻留因子）
    Sheet 2 结构: 序号 | 数据来源 | 使用部位 | 产品类型 | 日均使用量(g/day) | 驻留因子 | 参考说明
    数据来源: 欧盟SCCS(1-26行), 日本JCIA(27-67行), 荷兰RIVM(68-97行), U.S.EPA(98行)
    """
    cursor = conn.cursor()
    cursor.execute('DELETE FROM exposure_daily_usage')
    count = 0
    # 数据从第2行开始 (第1行是列标题)
    for r in range(2, ws.max_row + 1):
        seq = safe_str(ws.cell(row=r, column=1).value)
        source = safe_str(ws.cell(row=r, column=2).value)
        site = safe_str(ws.cell(row=r, column=3).value)
        product = safe_str(ws.cell(row=r, column=4).value)
        amount = ws.cell(row=r, column=5).value
        if amount is not None:
            try:
                amount = float(amount)
            except (ValueError, TypeError):
                amount = None
        retention = ws.cell(row=r, column=6).value
        if retention is not None:
            try:
                retention = float(retention)
            except (ValueError, TypeError):
                retention = None
        reference = safe_str(ws.cell(row=r, column=7).value)

        if not source and not site and not product:
            continue

        cursor.execute(
            'INSERT INTO exposure_daily_usage '
            '(source, seq, application_site, product_category, daily_amount_g, retention_factor, reference) '
            'VALUES (?, ?, ?, ?, ?, ?, ?)',
            (source, seq, site, product, amount, retention, reference)
        )
        count += 1
    conn.commit()
    return count


def import_exposure_body_weight(conn, ws):
    """导入年龄段体重默认值
    位置: Sheet 2 右侧区域 (cols 12-14, rows 1-6)
    来源: EFSA 2012a
    """
    cursor = conn.cursor()
    cursor.execute('DELETE FROM exposure_body_weight')
    count = 0
    for r in range(1, 7):
        age = safe_str(ws.cell(row=r, column=12).value)
        weight = ws.cell(row=r, column=13).value
        if weight is not None:
            try:
                weight = float(weight)
            except (ValueError, TypeError):
                continue
        else:
            continue
        notes = safe_str(ws.cell(row=r, column=14).value)

        cursor.execute(
            'INSERT INTO exposure_body_weight (age_group, default_weight_kg, source, notes) '
            'VALUES (?, ?, ?, ?)',
            (age, weight, 'EFSA 2012a', notes)
        )
        count += 1
    conn.commit()
    return count


def main():
    print('=' * 60)
    print('法规数据库导入工具')
    print('=' * 60)

    # 初始化数据库（确保所有表存在）
    db.init_db()
    conn = db.get_db()
    print(f'\n数据库路径: {db.DB_PATH}')
    print()

    try:
        # ── 壹：化妆品安全技术规范（2015年版）──
        print('[1/4] 正在导入化妆品安全技术规范（2015年版）...')
        wb = openpyxl.load_workbook(FILE_TECH_SPEC, data_only=True)
        sheets = wb.sheetnames

        # 表1 禁用组分
        n = import_banned_chemical(conn, wb[sheets[0]])
        print(f'  ✓ 禁用组分 表1: {n} 条')

        # 表2 禁用植(动)物组分
        n = import_banned_botanical(conn, wb[sheets[1]])
        print(f'  ✓ 禁用植(动)物组分 表2: {n} 条')

        # 表3 限用组分
        n = import_restricted(conn, wb[sheets[2]])
        print(f'  ✓ 限用组分 表3: {n} 条')

        # 表4 准用防腐剂
        n = import_allowed_preservative(conn, wb[sheets[3]])
        print(f'  ✓ 准用防腐剂 表4: {n} 条')

        # 表5 准用防晒剂
        n = import_allowed_sunscreen(conn, wb[sheets[4]])
        print(f'  ✓ 准用防晒剂 表5: {n} 条')

        # 表6 准用着色剂
        n = import_allowed_colorant(conn, wb[sheets[5]])
        print(f'  ✓ 准用着色剂 表6: {n} 条')

        # 表7 准用染发剂
        n = import_allowed_hair_dye(conn, wb[sheets[6]])
        print(f'  ✓ 准用染发剂 表7: {n} 条')

        wb.close()

        # ── 贰：已上市产品原料使用信息（2025）──
        print('\n[2/4] 正在导入已上市产品原料使用信息（2025）...')
        if os.path.exists(FILE_USAGE_MARKET):
            wb = openpyxl.load_workbook(FILE_USAGE_MARKET, data_only=True)
            ws = wb[wb.sheetnames[0]]
            n = import_usage_data(conn, ws, 'safety_usage_market')
            print(f'  ✓ 已上市产品原料使用信息: {n} 条')
            print('  📋 使用说明（安全评估规则）已保存到脚本常量 USAGE_RULES_MARKET')
            wb.close()
        else:
            print(f'  ⚠ 文件未找到: {FILE_USAGE_MARKET}')

        # ── 叁：国际化妆品安全评估数据索引 ──
        print('\n[3/4] 正在导入国际化妆品安全评估数据索引...')
        if os.path.exists(FILE_USAGE_INTERNATIONAL):
            wb = openpyxl.load_workbook(FILE_USAGE_INTERNATIONAL, data_only=True)
            ws = wb[wb.sheetnames[0]]
            n = import_usage_data(conn, ws, 'safety_usage_international')
            print(f'  ✓ 国际安全评估索引: {n} 条')
            wb.close()
        else:
            print(f'  ⚠ 文件未找到: {FILE_USAGE_INTERNATIONAL}')

        # ── 肆：暴露量数据（日均使用量 & 驻留因子）──
        print('\n[4/4] 正在导入暴露量数据（日均使用量 & 驻留因子）...')
        if os.path.exists(FILE_EXPOSURE_SED):
            wb = openpyxl.load_workbook(FILE_EXPOSURE_SED, data_only=True)
            sheets = wb.sheetnames
            # Sheet 2: 日均使用量和驻留因子
            ws = wb[sheets[1]]
            n = import_exposure_daily_usage(conn, ws)
            print(f'  ✓ 日均使用量 & 驻留因子: {n} 条记录')
            n_w = import_exposure_body_weight(conn, ws)
            print(f'  ✓ 年龄段体重默认值: {n_w} 个年龄段')
            wb.close()
        else:
            print(f'  ⚠ 文件未找到: {FILE_EXPOSURE_SED}')

        # ── 汇总 ──
        print()
        print('=' * 60)
        print('导入完成！数据库汇总')
        print('=' * 60)
        tables = [
            'regulation_banned_chemical', 'regulation_banned_botanical',
            'regulation_restricted', 'regulation_allowed_preservative',
            'regulation_allowed_sunscreen', 'regulation_allowed_colorant',
            'regulation_allowed_hair_dye',
            'safety_usage_market', 'safety_usage_international',
            'exposure_daily_usage', 'exposure_body_weight',
        ]
        total = 0
        for t in tables:
            row = conn.execute(f'SELECT COUNT(*) FROM {t}').fetchone()
            cnt = row[0]
            print(f'  {t}: {cnt} 条')
            total += cnt
        print('  ─────────────────────────')
        print(f'  总计: {total} 条')
        print('=' * 60)

    except Exception as e:
        print(f'\n[ERROR] {e}')
        import traceback
        traceback.print_exc()
    finally:
        conn.close()


if __name__ == '__main__':
    main()
