"""构建统一安全评估数据库 Excel (安全评估数据库.xlsx)

从现有数据源整合:
  目录         ← userdata/已使用化妆品原料目录I.xlsx + 目录II.xlsx
  已上市使用量   ← safety.db safety_usage_market
  国际索引       ← safety.db safety_usage_international
  别名           ← (空结构)
  毒理学数据     ← toxicology_seed.json
  暴露量参数     ← safety.db exposure_daily_usage (97条)
  体重参数       ← safety.db exposure_body_weight
  CosIng        ← cosing_data.json
  风险物质规则    ← risk_control_rules.json
  风险物质识别    ← risk_substances.json
"""
import json
import sqlite3
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE, 'data')
OUTPUT = os.path.join(DATA_DIR, '安全评估数据库.xlsx')
DB_PATH = os.path.join(BASE, 'userdata', 'safety.db')
CAT_I_XLSX = os.path.join(BASE, 'userdata', '已使用化妆品原料目录I.xlsx')
CAT_II_XLSX = os.path.join(BASE, 'userdata', '已使用化妆品原料目录II.xlsx')
DATA_XLSX = os.path.join(BASE, 'userdata', 'data.xlsx')
TOX_JSON = os.path.join(BASE, 'userdata', 'toxicology_seed.json')
COSING_JSON = os.path.join(BASE, 'userdata', 'cosing_data.json')
RISK_JSON = os.path.join(BASE, 'userdata', 'risk_control_rules.json')
RISK_SUB_JSON = os.path.join(BASE, 'userdata', 'risk_substances.json')

# ── Styles ──────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ALT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_sheet(ws, headers, rows, col_widths=None, col_formats=None):
    """Write headers + data rows + apply formatting."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for ri, row in enumerate(rows, 2):
        alt = ri % 2 == 0
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val if val is not None else '')
            cell.border = THIN_BORDER
            cell.alignment = LEFT_WRAP
            if alt:
                cell.fill = ALT_FILL

    ws.freeze_panes = 'A2'
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    if col_formats:
        for ci, fmt in enumerate(col_formats, 1):
            for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1,
                                    min_col=ci, max_col=ci):
                for cell in row:
                    cell.number_format = fmt


# ══════════════════════════════════════════════════════════
# 1. Read Catalog I
# ══════════════════════════════════════════════════════════
def read_catalog_i():
    wb = load_workbook(CAT_I_XLSX)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(row)
    return rows  # (seq, cn_name, inci_name, remark)


# ══════════════════════════════════════════════════════════
# 2. Read Catalog II
# ══════════════════════════════════════════════════════════
def read_catalog_ii():
    wb = load_workbook(CAT_II_XLSX)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        rows.append(row)  # seq,cn_name,inci_name,en_name,cas,formula,structure,mol_weight,use_purpose,use_range,safe_amount,remark
    return rows


# ══════════════════════════════════════════════════════════
# 3. Read usage data from DB
# ══════════════════════════════════════════════════════════
def read_db_table(table):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(f'SELECT * FROM [{table}] ORDER BY id')
    rows = cur.fetchall()
    conn.close()
    return rows  # includes id column


# ══════════════════════════════════════════════════════════
# 4. Read toxicology from JSON
# ══════════════════════════════════════════════════════════
def read_toxicology():
    with open(TOX_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data.get('entries', data)


# ══════════════════════════════════════════════════════════
# 5. Read CosIng
# ══════════════════════════════════════════════════════════
def read_cosing():
    with open(COSING_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    rows = []
    for inci_name, info in data.items():
        rows.append((inci_name, info.get('cas', ''), info.get('ec', '')))
    rows.sort(key=lambda x: x[0])
    return rows


# ══════════════════════════════════════════════════════════
# 6. Read exposure from DB (97 rows)
def read_exposure():
    rows = read_db_table('exposure_daily_usage')
    # DB columns: id, source, seq, application_site, product_category,
    #             daily_amount_g, retention_factor, reference
    return [(r[4], r[3], r[1], r[5], r[6], r[7]) for r in rows]


def read_body_weight():
    rows = read_db_table('exposure_body_weight')
    # DB columns: id, age_group, default_weight_kg, source, notes
    return [(r[1], r[2], r[3], r[4]) for r in rows]


def read_body_weight():
    rows = read_db_table('exposure_body_weight')
    # DB columns: id, age_group, default_weight_kg, source, notes
    return [(r[1], r[2], r[3], r[4]) for r in rows]


# ══════════════════════════════════════════════════════════
# 7. Read risk control rules
# ══════════════════════════════════════════════════════════
def read_risk_rules():
    with open(RISK_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    rows = []
    for cat in data.get('categories', []):
        label = cat.get('label', '')
        ref = cat.get('ref', '')
        measures = cat.get('measures', [])
        for m in measures:
            if isinstance(m, dict):
                rows.append((label, m.get('measure', ''), m.get('detail', ''), ref))
            else:
                rows.append((label, str(m), '', ref))
    return rows


def read_authority_opinions():
    with open(TOX_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    rows = []
    for e in data.get('entries', []):
        ref = e.get('ref', '')
        if 'CIR' not in ref and 'SCCS' not in ref:
            continue
        authority = ref.split('.')[0].split(' ')[0] if '.' in ref else ''
        if authority not in ('CIR', 'SCCS'):
            authority = 'SCCS' if 'SCCS' in ref else 'CIR'
        noael = e.get('noael_mg_kg_day', '')
        study_type = e.get('study_type', '')
        opinion = f'经{study_type}毒性试验，该原料的NOAEL为{noael}mg/kg bw/day，采用100倍安全系数，其在化妆品中安全使用浓度需根据MoS计算确定。'
        rows.append((
            e.get('name_zh', ''),
            e.get('name_inci', ''),
            e.get('cas_no', ''),
            authority,
            opinion,
            ref,
        ))
    return rows


def read_risk_substances():
    with open(RISK_SUB_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    rows = []
    for r in data.get('general_risk_substances', []):
        rows.append((
            '通用',
            r.get('name', ''),
            '；'.join(r.get('trigger_keywords', [])),
            '否',
            r.get('limit', ''),
            r.get('reference', ''),
            r.get('description', ''),
        ))
    for r in data.get('child_risk_substances', []):
        rows.append((
            '儿童产品',
            r.get('name', ''),
            '；'.join(r.get('trigger_keywords', [])),
            '否',
            r.get('limit', ''),
            r.get('reference', ''),
            r.get('description', ''),
        ))
    return rows


# ══════════════════════════════════════════════════════════
# BUILD
# ══════════════════════════════════════════════════════════
def main():
    os.makedirs(DATA_DIR, exist_ok=True)

    print("正在读取数据源...")
    cat_i = read_catalog_i()
    print(f"  目录I: {len(cat_i)} 条")
    cat_ii = read_catalog_ii()
    print(f"  目录II: {len(cat_ii)} 条")
    usage_market = read_db_table('safety_usage_market')
    print(f"  已上市使用量: {len(usage_market)} 条")
    usage_intl = read_db_table('safety_usage_international')
    print(f"  国际索引: {len(usage_intl)} 条")
    tox_data = read_toxicology()
    print(f"  毒理学: {len(tox_data)} 条")
    cosing = read_cosing()
    print(f"  CosIng: {len(cosing)} 条")
    exposure = read_exposure()
    print(f"  暴露量: {len(exposure)} 条")
    risk_data = read_risk_rules()
    print(f"  风险规则: {len(risk_data)} 条")
    risk_sub = read_risk_substances()
    print(f"  风险物质: {len(risk_sub)} 条")
    opinions = read_authority_opinions()
    print(f"  权威评估意见: {len(opinions)} 条")
    bwt = read_body_weight()
    print(f"  体重参数: {len(bwt)} 条")

    wb = Workbook()

    # ── 1. 目录 ──
    ws = wb.active
    ws.title = '目录'
    headers = ['序号', '目录类别', '标准中文名称', 'INCI名称/英文名称', '英文名称',
               'CAS号', '原料分子式', '相对分子质量', '使用目的', '使用范围',
               '安全使用量(%)', '备注']
    col_widths = [8, 10, 30, 35, 35, 18, 20, 12, 12, 30, 15, 20]

    # Build merged: add cat_i with category=I, merge cat_ii with extra fields
    cat_ii_dict = {}
    for r in cat_ii:
        seq = str(r[0]).strip() if r[0] is not None else ''
        cat_ii_dict[seq] = r  # seq,cn,inci,en,cas,formula,struct,mw,use_purpose,use_range,safe_amt,remark

    merged = []
    for r in cat_i:
        seq = str(r[0]).strip() if r[0] is not None else ''
        cn = r[1] or ''
        inci = r[2] or ''
        remark = r[3] or ''
        en = ''
        cas = ''
        formula = ''
        mw = ''
        purpose = ''
        use_range = ''
        safe_amt = ''

        if seq in cat_ii_dict:
            ii = cat_ii_dict[seq]
            en = ii[3] or ''
            cas = ii[4] or ''
            formula = ii[5] or ''
            mw = ii[7] or ''
            purpose = ii[8] or ''
            use_range = ii[9] or ''
            safe_amt = ii[10] or ''
            remark = remark or ii[11] or ''

        merged.append((seq, 'I', cn, inci, en, cas, formula, mw,
                       purpose, use_range, safe_amt, remark))

    # add pure cat_ii entries not in cat_i (shouldn't happen but safe)
    for seq, ii in cat_ii_dict.items():
        if not any(r[0] and str(r[0]).strip() == seq for r in merged):
            merged.append((seq, 'II', ii[1] or '', ii[2] or '', ii[3] or '',
                          ii[4] or '', ii[5] or '', ii[7] or '',
                          ii[8] or '', ii[9] or '', ii[10] or '', ii[11] or ''))

    merged.sort(key=lambda x: x[0])
    style_sheet(ws, headers, merged, col_widths)
    print("  目录 写入完成")

    # ── 2. 已上市使用量 ──
    ws2 = wb.create_sheet('已上市使用量')
    headers2 = ['序号', '目录序号', '原料名称', 'INCI名称', '作用部位',
                '使用方法', '最大使用量(%)', '备注']
    rows2 = [(r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7])
             for r in usage_market]
    style_sheet(ws2, headers2, rows2, [8, 12, 30, 35, 20, 15, 15, 20])
    print(f"  已上市使用量 写入完成 ({len(rows2)} 条)")

    # ── 3. 国际索引 ──
    ws3 = wb.create_sheet('国际索引')
    style_sheet(ws3, headers2,  # same structure
                [(r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7])
                 for r in usage_intl],
                [8, 12, 30, 35, 20, 15, 15, 20])
    print(f"  国际索引 写入完成 ({len(usage_intl)} 条)")

    # ── 4. 别名 ──
    ws4 = wb.create_sheet('别名')
    alias_headers = ['目录序号', '曾用标准中文名称', '曾用INCI名称', '备注']
    style_sheet(ws4, alias_headers, [], [12, 30, 35, 20])
    print("  别名 (空结构) 写入完成")

    # ── 5. 毒理学数据 ──
    ws5 = wb.create_sheet('毒理学数据')
    tox_headers = ['原料名称', 'INCI名称', 'CAS号', 'NOAEL(mg/kg/day)',
                   '研究类型', '物种', '不确定系数', '暴露途径', '参考文献']
    # Flatten entries
    tox_rows = []
    if isinstance(tox_data, list) and all(isinstance(x, dict) for x in tox_data):
        for e in tox_data:
            tox_rows.append((
                e.get('name_zh', ''), e.get('name_inci', ''),
                e.get('cas_no', ''), e.get('noael_mg_kg_day', ''),
                e.get('study_type', ''), e.get('species', ''),
                e.get('uncertainty_factor', ''), e.get('route', ''),
                e.get('ref', '')
            ))
    style_sheet(ws5, tox_headers, tox_rows,
                [25, 30, 18, 18, 15, 10, 14, 12, 40])
    print(f"  毒理学数据 写入完成 ({len(tox_rows)} 条)")

    # ── 6. 暴露量参数 ──
    ws6 = wb.create_sheet('暴露量参数')
    exp_headers = ['产品类别', '使用部位', '数据来源', '日均使用量(g)',
                   '驻留因子', '备注']
    style_sheet(ws6, exp_headers, exposure,
                [18, 15, 16, 16, 10, 30])
    print(f"  暴露量参数 写入完成 ({len(exposure)} 条)")

    # ── 7. 体重参数 ──
    ws7 = wb.create_sheet('体重参数')
    bw_headers = ['年龄段', '默认体重(kg)', '数据来源', '备注']
    style_sheet(ws7, bw_headers, bwt,
                [16, 14, 16, 30])
    print(f"  体重参数 写入完成 ({len(bwt)} 条)")

    # ── 9. CosIng ──
    ws9 = wb.create_sheet('CosIng')
    cos_headers = ['INCI名称', 'CAS号', 'EC号']
    style_sheet(ws9, cos_headers, cosing,
                [45, 20, 20])
    print(f"  CosIng 写入完成 ({len(cosing)} 条)")

    # ── 10. 风险物质规则 ──
    ws10 = wb.create_sheet('风险物质规则')
    risk_headers = ['风险类别', '控制措施', '措施详情', '参考依据']
    style_sheet(ws10, risk_headers, risk_data,
                [20, 30, 50, 40])
    print(f"  风险物质规则 写入完成 ({len(risk_data)} 条)")

    # ── 11. 风险物质识别 ──
    ws11 = wb.create_sheet('风险物质识别')
    risk_sub_headers = ['风险类型', '物质名称', '触发关键词', '所有产品触发',
                        '限值', '参考来源', '评估说明']
    style_sheet(ws11, risk_sub_headers, risk_sub,
                [12, 18, 50, 12, 30, 35, 60])
    print(f"  风险物质识别 写入完成 ({len(risk_sub)} 条)")

    # ── 12. 权威评估意见 ──
    ws12 = wb.create_sheet('权威评估意见')
    opinion_headers = ['原料名称', 'INCI名', 'CAS号', '评估机构', '评估意见', '参考编号']
    style_sheet(ws12, opinion_headers, opinions,
                [18, 35, 18, 10, 60, 50])
    print(f"  权威评估意见 写入完成 ({len(opinions)} 条)")

    # ── Save with retry + fallback ──
    import tempfile
    fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', dir=DATA_DIR)
    os.close(fd)
    try:
        wb.save(tmp_path)
    except Exception:
        os.remove(tmp_path)
        raise

    # Try to replace old file; fallback to new file if locked
    import shutil
    try:
        if os.path.exists(OUTPUT):
            os.remove(OUTPUT)
        os.rename(tmp_path, OUTPUT)
    except PermissionError:
        base, ext = os.path.splitext(OUTPUT)
        backup = base + '_new' + ext
        if os.path.exists(backup):
            os.remove(backup)
        os.rename(tmp_path, backup)
        print(f"\n! 原文件被锁定，已保存至: {backup}")
        print(f"  请关闭占用程序后手动替换: {OUTPUT}")
    else:
        print(f"\n统一数据库已保存至: {OUTPUT}")

    # Summary
    total = (len(cat_i) + len(usage_market) + len(usage_intl)
             + len(tox_rows) + len(exposure) + len(bwt) + len(cosing)
             + len(risk_data) + len(risk_sub) + len(opinions))
    print(f"  共写入 {total} 条数据，12 个工作表")


if __name__ == '__main__':
    main()
