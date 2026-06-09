import os
import sys
import json
import io
import re
import traceback

from datetime import datetime

# pandas 和 openpyxl 延迟导入，加快启动速度
# import pandas as pd
# from openpyxl.styles import Font, PatternFill, Alignment
# from openpyxl import Workbook

from PySide6.QtWidgets import (
    QMainWindow, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QPushButton, QLabel, QLineEdit,
    QDoubleSpinBox, QHeaderView, QMessageBox, QFileDialog,
    QPlainTextEdit, QTextBrowser, QProgressBar, QFrame, QAbstractItemView,
    QDialog, QFormLayout, QDialogButtonBox, QCompleter, QCheckBox,
    QComboBox, QInputDialog, QGroupBox,
    QGridLayout, QAbstractSpinBox
)
from PySide6.QtCore import Qt, QStringListModel, Signal, QThread
from PySide6.QtGui import QColor

import database as db
import safety_engine as se
from catalog_mgr import CatalogManager
import report_generator as rg
import project_manager as pm
import review_workflow as rw


# ── helpers ──

def _parse_comp_str(s):
    parts = re.split(r'[，、]|, ', s)
    comps = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        m = re.match(r'(.+?)[\(\（](\d+\.?\d*)%?[\)\）]', p)
        if m:
            comps.append({'name': m.group(1).strip(), 'percent': float(m.group(2))})
        else:
            comps.append({'name': p, 'percent': 100})
    return comps


def _comp_str_to_display(composition):
    parts = []
    for c in composition:
        name = c.get('name', '')
        pct = c.get('percent', 0)
        parts.append(f"{name}({pct}%)")
    return '; '.join(parts)


def _build_composition_from_natural(raw_name, comp_str, catalog):
    comps = []
    comp_incis = []
    if comp_str:
        for part in re.split(r'[，、]|, ', comp_str):
            part = part.strip()
            if not part:
                continue
            m = re.match(r'(.+?)[\(\（](\d+\.?\d*)%?[\)\）]', part)
            if m:
                cn = m.group(1).strip()
                pct = float(m.group(2))
                inci = catalog.match_inci(cn)
                comps.append({'name': cn, 'inci': inci, 'percent': pct})
                if inci and inci not in comp_incis:
                    comp_incis.append(inci)
            else:
                inci = catalog.match_inci(part)
                comps.append({'name': part, 'inci': inci, 'percent': 100})
                if inci and inci not in comp_incis:
                    comp_incis.append(inci)
    if not comps:
        comps = [{'name': raw_name, 'inci': '', 'percent': 100}]
    return comps, comp_incis


# INCI名称到中文名称的映射表（修正常见错误映射）
INCI_TO_CHINESE = {
    # 蜡类
    'BEESWAX': '蜂蜡',
    'CERA ALBA': '白蜂蜡',
    'CERA FLAVA': '黄蜂蜡',
    # 甘油酯类
    'GLYCERIN': '甘油',
    'GLYCERYL STEARATE': '硬脂酸甘油酯',
    'GLYCERYL OLEATE': '油酸甘油酯',
    'C10-18 TRIGLYCERIDES': 'C10-18甘油三酯',
    'HYDROGENATED C12-18 TRIGLYCERIDES': '氢化C12-18甘油三酯',
    'CAPRYLIC/CAPRIC TRIGLYCERIDE': '辛酸/癸酸甘油三酯',
    # PEG类
    'PEG-40 STEARATE': 'PEG-40硬脂酸酯',
    'PEG-120 STEARATE': 'PEG-120硬脂酸酯',
    'PEG-60 HYDROGENATED CASTOR OIL': 'PEG-60氢化蓖麻油',
    'PEG-40 HYDROGENATED CASTOR OIL': 'PEG-40氢化蓖麻油',
    # 脂肪酸类
    'STEARIC ACID': '硬脂酸',
    'OLEIC ACID': '油酸',
    'PALMITIC ACID': '棕榈酸',
    'MYRISTIC ACID': '肉豆蔻酸',
    # 脂肪醇类
    'CETYL ALCOHOL': '鲸蜡醇',
    'STEARYL ALCOHOL': '硬脂醇',
    'CETEARYL ALCOHOL': '鲸蜡硬脂醇',
    'LAURYL ALCOHOL': '月桂醇',
    # 表面活性剂
    'SODIUM LAURETH SULFATE': '月桂醇聚醚硫酸酯钠',
    'SODIUM LAURYL SULFATE': '月桂醇硫酸酯钠',
    'COCAMPHOSPHATE': '椰油基磷酸酯',
    'C9-15 PARETH-8': 'C9-15链烷醇聚醚-8',
    # 防腐剂
    'METHYLPARABEN': '羟苯甲酯',
    'ETHYLPARABEN': '羟苯乙酯',
    'PROPYLPARABEN': '羟苯丙酯',
    'BUTYLPARABEN': '羟苯丁酯',
    'PHENOXYETHANOL': '苯氧乙醇',
    'DMDM HYDANTOIN': 'DMDM乙内酰脲',
    'IMIDAZOLIDINYL UREA': '咪唑烷基脲',
    'PHENETHYL ALCOHOL': '苯乙醇',
    'ETHYLHEXYLGLYCERIN': '乙基己基甘油',
    'CAPRYLOHYDROXAMIC ACID': '辛酰羟肟酸',
    'GLYCERYL CAPRYLATE': '甘油辛酸酯',
    # 染发剂
    '1,3-BIS-(2,4-DIAMINOPHENOXY)PROPANE': '1,3-双(2,4-二氨基苯氧基)丙烷',
    'N,N-BIS(2-HYDROXYETHYL)-P-PHENYLENEDIAMINE SULFATE': 'N,N-双(2-羟乙基)对苯二胺硫酸盐',
    'P-PHENYLENEDIAMINE': '对苯二胺',
    'RESORCINOL': '间苯二酚',
    # 多元醇
    'BUTYLENE GLYCOL': '丁二醇',
    '1,3-PROPANEDIOL': '1,3-丙二醇',
    '1,2-HEXANEDIOL': '1,2-己二醇',
    'PENTYLENE GLYCOL': '戊二醇',
    # 糖类
    'XYLITOL': '木糖醇',
    'ANHYDROXYLITOL': '脱水木糖醇',
    'INOSITOL': '肌醇',
    # 植物提取物
    'CENTELLA ASIATICA LEAF EXTRACT': '积雪草叶提取物',
    'TREMELLA FUCIFORMIS EXTRACT': '银耳提取物',
    'BELLIS PERENNIS FLOWER EXTRACT': '雏菊花提取物',
    'BUTYROSPERMUM PARKII BUTTER': '牛油果树果脂',
    'EUGLENA GRACILIS POLYSACCHARIDE': '细小裸藻多糖',
    # 透明质酸
    'SODIUM HYALURONATE': '透明质酸钠',
    # 其他功效成分
    'PANTENOL': '泛醇',
    'TOCOPHEROL': '生育酚',
    'TOCOPHERYL ACETATE': '生育酚乙酸酯',
    'BISABOLOL': '红没药醇',
    'DIPOTASSIUM GLYCYRRHIZATE': '甘草酸二钾',
    'SQUALANE': '角鲨烷',
    'HYDROXYPROPYL TETRAHYDROPYRANTRIOL': '羟丙基四氢吡喃三醇',
    'ECTOINE': '四氢甲基嘧啶羧酸',
    'ERGOTHIONEINE': '麦角硫因',
    'SOLUBLE COLLAGEN': '可溶性胶原',
    'HYDROLYZED COLLAGEN': '水解胶原',
    'HYDROLYZED RHODOPHYCEAE EXTRACT': '水解红藻提取物',
    'SODIUM CITRATE': '柠檬酸钠',
    # 乳化剂
    'CETEARYL OLIVATE': '鲸蜡硬脂醇橄榄油酸酯',
    'SORBITAN OLIVATE': '山梨坦橄榄油酸酯',
    'AMMONIUM ACRYLOYLDIMETHYLTAURATE/VP COPOLYMER': '丙烯酰二甲基牛磺酸铵/VP共聚物',
    # 增稠剂
    'XANTHAN GUM': '黄原胶',
    # 螯合剂
    'DISODIUM EDTA': 'EDTA二钠',
    # 无机盐
    'SODIUM CHLORIDE': '氯化钠',
    'POTASSIUM CHLORIDE': '氯化钾',
    'POTASSIUM PHOSPHATE': '磷酸钾',
    'DIPOTASSIUM PHOSPHATE': '磷酸氢二钾',
    # 其他
    '1-METHYLHYDANTOIN-2-IMIDE': '1-甲基乙内酰脲-2-酰亚胺',
    'WATER': '水',
    'ALCOHOL': '乙醇',
    'PROPYLENE GLYCOL': '丙二醇',
    'ETHYLENE GLYCOL': '乙二醇',
    'SEA WATER': '海水',
}


def _generate_label_ingredients(formula, catalog):
    """
    根据配方生成化妆品标签成分表（符合《化妆品标签管理办法》要求）
    
    法规要求：
    1. 成分按含量降序排列（≥1%）
    2. <1%成分可任意排列在≥1%成分之后
    3. 使用INCI名称，可同时标注中文名称
    4. 香精可标注为"香精"无需列出具体成分
    5. 着色剂需标注CI索引号
    
    Args:
        formula: 配方数据列表
        catalog: CatalogManager实例
    
    Returns:
        标签成分表文本
    """
    # 收集所有成分及其实际含量
    ingredients = []
    
    for item in formula:
        composition = json.loads(item['composition'])
        added_percent = item['added_percent']
        
        for comp in composition:
            # 计算实际含量
            actual_percent = added_percent * comp['percent'] / 100.0
            
            # 获取INCI名称（优先从成分中获取，其次从原料中获取）
            inci_name = comp.get('inci', '').strip() or item.get('name_inci', '').strip()
            
            # 获取中文名称（优先从成分中获取，其次从原料中获取）
            chinese_name = comp.get('name', '').strip()
            
            # 如果INCI名称在映射表中，修正中文名称
            if inci_name:
                normalized_inci = inci_name.upper().replace(' ', '')
                for inci_key, chinese_val in INCI_TO_CHINESE.items():
                    # 精确匹配
                    if normalized_inci == inci_key.replace(' ', ''):
                        # 使用映射表中的标准中文名称
                        chinese_name = chinese_val
                        break
                    # 模糊匹配（处理常见变体）
                    elif normalized_inci in inci_key.replace(' ', '') or inci_key.replace(' ', '') in normalized_inci:
                        # 对于植物提取物，匹配后缀
                        if 'EXTRACT' in inci_key and 'EXTRACT' in inci_name:
                            chinese_name = chinese_val
                            break
            
            # 特殊成分处理
            is_fragrance = False
            is_colorant = False
            ci_number = ''
            
            # 检查是否为香精
            if '香精' in chinese_name or 'fragrance' in inci_name.lower() or 'parfum' in inci_name.lower():
                is_fragrance = True
                inci_name = 'Fragrance'
                chinese_name = '香精'
            # 检查是否为着色剂（含CI号）
            elif 'CI ' in chinese_name or 'CI' in inci_name:
                is_colorant = True
                # 提取CI号
                ci_match = re.search(r'CI\s*\d+(\s*-\s*\d+)*', chinese_name + ' ' + inci_name, re.IGNORECASE)
                if ci_match:
                    ci_number = ci_match.group().replace(' ', '')
            
            # 清理INCI名称中的多余空格
            if inci_name:
                inci_name = ' '.join(inci_name.split())
            
            ingredients.append({
                'inci': inci_name,
                'chinese': chinese_name,
                'percent': actual_percent,
                'is_fragrance': is_fragrance,
                'is_colorant': is_colorant,
                'ci_number': ci_number
            })
    
    # 按含量排序（≥1%在前，<1%在后）
    ingredients.sort(key=lambda x: (-x['percent'] >= 1, -x['percent']))
    
    # 构建成分表文本
    lines = []
    lines.append("成分表：")
    
    # 植物/动物提取物关键词
    plant_keywords = ['EXTRACT', 'POLYSACCHARIDE', 'BUTTER', 'LEAF', 'FLOWER', 'FRUIT', 'WATER']
    
    for idx, ing in enumerate(ingredients):
        # 着色剂处理
        if ing['is_colorant'] and ing['ci_number']:
            line = ing['ci_number']
            if ing['chinese'] and ing['chinese'] != ing['ci_number']:
                line += f" ({ing['chinese']})"
        # 香精处理
        elif ing['is_fragrance']:
            line = '香精'
        else:
            # 检查中文名称是否已经包含拉丁学名（有括号）
            has_latin = '（' in ing.get('chinese', '') or '(' in ing.get('chinese', '')
            
            # 如果是植物提取物且中文名称没有拉丁学名，则添加
            if ing['inci'] and any(keyword in ing['inci'].upper() for keyword in plant_keywords) and not has_latin:
                if ing['chinese']:
                    line = f"{ing['chinese']}（{ing['inci']}）"
                else:
                    line = ing['inci']
            # 普通成分 - 只显示中文名称
            else:
                line = ing['chinese'] or ing['inci']
        
        lines.append(line)
    
    # 添加分隔符
    result = ', '.join(lines[1:])  # 跳过标题行
    
    # 添加合规说明
    footer = f"\n\n【合规说明】\n"
    footer += f"- 成分按含量降序排列（≥1%成分在前）\n"
    footer += f"- 香精统一标注为\"Fragrance (香精)\"\n"
    footer += f"- 着色剂标注CI索引号\n"
    footer += f"- 共 {len(ingredients)} 种成分"
    
    return result


def _export_formula_to_excel(formula, catalog):
    # 局部导入openpyxl，加快启动速度
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl import Workbook
    
    rows = []
    seq = 1
    for item in formula:
        composition = json.loads(item['composition'])
        added = item['added_percent']
        use_purpose = item['purpose_override'] if item['purpose_override'] else \
            (item['default_purpose'] if item['default_purpose'] else '')
        first = composition[0]
        rows.append({
            '原料序号': seq,
            '内部代码': item.get('internal_code', '') or '',
            '标准中文名称': first['name'],
            'INCI名称': first.get('inci', '') or item['name_inci'] or '',
            'CAS号': item.get('cas_number', '') or '',
            '原料含量(%)': added,
            '原料中成分含量(%)': first['percent'],
            '实际成分含量(%)': round(added * first['percent'] / 100, 4),
            '主要使用目的': use_purpose,
            '原料报送码': item['supplier_code'] if item['supplier_code'] else '',
            '备注': item['remarks'] if item['remarks'] else '',
            '是否新原料': '是' if item['is_new_material'] else '',
            '注册号/备案号': item['registration_number'] if item['registration_number'] else '',
            '_merge_seq': True,
            '_seq_value': seq
        })
        for comp in composition[1:]:
            rows.append({
                '原料序号': None,
                '内部代码': '',
                '标准中文名称': comp['name'],
                'INCI名称': comp.get('inci', ''),
                'CAS号': '',
                '原料含量(%)': None,
                '原料中成分含量(%)': comp['percent'],
                '实际成分含量(%)': round(added * comp['percent'] / 100, 4),
                '主要使用目的': '',
                '原料报送码': '',
                '备注': '',
                '是否新原料': 0,
                '注册号/备案号': '',
                '_merge_seq': False,
                '_seq_value': seq
            })
        seq += 1

    cols = ['原料序号', '内部代码', '标准中文名称', 'INCI名称', 'CAS号', '原料含量(%)', '原料中成分含量(%)',
            '实际成分含量(%)', '主要使用目的', '原料报送码', '备注', '是否新原料', '注册号/备案号']

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = '配方表'

    # Title row
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = '配方表'
    title_cell.font = Font(name='宋体', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Header row (row 2)
    headers = ['原料序号', '内部代码', '标准中文名称', 'INCI名称', 'CAS号', '原料含量(%)', '原料中成分含量(%)',
               '实际成分含量(%)', '主要使用目的', '原料报送码', '备注', '是否新原料', '注册号/备案号']
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(name='宋体', color="FFFFFF", bold=True)
    hdr_align = Alignment(horizontal="center", vertical="center")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align

    # Data rows starting from row 3
    for row_idx, row_data in enumerate(rows, 3):
        for col_idx, col_name in enumerate(cols, 1):
            value = row_data[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='宋体')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Apply number format
            if col_name == '原料含量(%)' and value is not None:
                cell.number_format = '0.00'
            elif col_name == '原料中成分含量(%)' and value is not None:
                cell.number_format = '0.00'
            elif col_name == '实际成分含量(%)' and value is not None:
                cell.number_format = '0.0000'

    # Merge cells for composite materials
    # Columns to merge: 原料序号(A), 内部代码(B), CAS号(E), 原料含量(%)(F), 主要使用目的(I), 原料报送码(J), 备注(K), 是否新原料(L), 注册号/备案号(M)
    merge_cols = [1, 2, 5, 6, 9, 10, 11, 12, 13]  # 1-indexed column numbers
    merge_start = 3
    current_seq = rows[0]['_seq_value'] if rows else None
    for i, row_data in enumerate(rows):
        if row_data['_seq_value'] != current_seq:
            end_row = i + 2  # +2 because rows is 0-indexed, Excel data starts at row 3
            if merge_start < end_row:
                for col in merge_cols:
                    ws.merge_cells(start_row=merge_start, start_column=col, end_row=end_row, end_column=col)
            merge_start = end_row + 1
            current_seq = row_data['_seq_value']
    # Merge the last group
    if merge_start <= len(rows) + 2:
        end_row = len(rows) + 2
        for col in merge_cols:
            ws.merge_cells(start_row=merge_start, start_column=col, end_row=end_row, end_column=col)

    # Set column widths
    column_widths = {
        'A': 10,  # 原料序号
        'B': 12,  # 内部代码
        'C': 25,  # 标准中文名称
        'D': 30,  # INCI名称
        'E': 15,  # CAS号
        'F': 12,  # 原料含量(%)
        'G': 15,  # 原料中成分含量(%)
        'H': 15,  # 实际成分含量(%)
        'I': 15,  # 主要使用目的
        'J': 20,  # 原料报送码
        'K': 20,  # 备注
        'L': 10,  # 是否新原料
        'M': 15   # 注册号/备案号
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output)
    output.seek(0)
    return output


# ── Worker for crawling ──

class CrawlWorker(QThread):
    log_updated = Signal(str)
    progress_updated = Signal(int, int, int)
    finished_signal = Signal(bool, int)
    cat_type = 'I'

    def run(self):
        try:
            sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            if self.cat_type == 'I':
                import crawlers.crawl_catalog_i as cm
            else:
                import crawlers.crawl_catalog_ii as cm

            from catalog_mgr import USERDATA_DIR as output_dir

            def cb(page, total, count):
                self.progress_updated.emit(page, total, count)
                self.log_updated.emit(f"第 {page}/{total} 页，累计 {count} 条")

            self.log_updated.emit("正在启动浏览器...")
            if self.cat_type == 'I':
                success = cm.crawl_with_edge(output_dir=output_dir, progress_callback=cb)
            else:
                success = cm.crawl_catalog_ii(output_dir=output_dir, progress_callback=cb)

            self.finished_signal.emit(success, 0)
        except Exception as e:
            self.log_updated.emit(f"爬取失败: {e}\n{traceback.format_exc()}")
            self.finished_signal.emit(False, 0)


class CasCrawlWorker(QThread):
    log_updated = Signal(str)
    progress_updated = Signal(int, int, int)
    finished_signal = Signal(bool, int)

    def __init__(self, mode=0):
        super().__init__()
        self.mode = mode  # 0=combined, 1=cosing only, 2=pubchem only

    def run(self):
        try:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            import crawl_cosing as cc
            import pandas as pd
            from catalog_mgr import USERDATA_DIR as userdata_dir

            catalog_path = os.path.join(userdata_dir, "ingredient_catalog.json")
            if not os.path.exists(catalog_path):
                self.log_updated.emit("目录 I 文件不存在，请先同步目录I")
                self.finished_signal.emit(False, 0)
                return

            catalog = pd.read_json(catalog_path)
            names = catalog['inci_name'].dropna().tolist()

            def cb(page, total, count):
                self.progress_updated.emit(page, total, count)
                self.log_updated.emit(f"第 {page}/{total} 个，已获取 {count} 条")

            self.log_updated.emit(f"共有 {len(names)} 个 INCI 名称待查询（仅查询未获取的）")

            if self.mode == 1:
                self.log_updated.emit("模式: 仅 COSING")
                cc.download_cosing_only(names, progress_callback=cb, incremental=True)
                result = cc.load_cosing_data()
            elif self.mode == 2:
                self.log_updated.emit("模式: 仅 PubChem")
                result = cc.download_pubchem_fallback(names, progress_callback=cb, incremental=True)
            else:
                self.log_updated.emit("模式: COSING + PubChem")
                cc.download_cosing_only(names, progress_callback=cb, incremental=True)
                result = cc.download_pubchem_fallback(names, progress_callback=cb, incremental=True)
            self.finished_signal.emit(True, len(result))
        except Exception as e:
            self.log_updated.emit(f"CAS号爬取失败: {e}\n{traceback.format_exc()}")
            self.finished_signal.emit(False, 0)


# ── Material Edit Dialog ──

class MaterialEditDialog(QDialog):
    def __init__(self, catalog, material=None, parent=None):
        super().__init__(parent)
        self.catalog = catalog
        self.material = material
        self.setWindowTitle("编辑原料" if material else "添加原料")
        self.setMinimumWidth(500)
        self._build_ui()
        if material:
            self._load_data()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.edit_name = QLineEdit()
        self.edit_name.setPlaceholderText("原料名称（必填）")
        form.addRow("原料名称:", self.edit_name)

        self.edit_code = QLineEdit()
        form.addRow("内部代码:", self.edit_code)

        self.edit_purpose = QLineEdit()
        form.addRow("使用目的:", self.edit_purpose)

        self.edit_supplier = QLineEdit()
        form.addRow("原料报送码:", self.edit_supplier)

        self.edit_supplier_name = QLineEdit()
        form.addRow("生产商:", self.edit_supplier_name)

        self.edit_cas = QLineEdit()
        form.addRow("CAS号:", self.edit_cas)

        self.edit_remarks = QLineEdit()
        form.addRow("备注:", self.edit_remarks)

        self.edit_inci = QLineEdit()
        self.edit_inci.setPlaceholderText("留空自动匹配")
        self.edit_inci.textChanged.connect(self._on_inci_changed)
        form.addRow("INCI名:", self.edit_inci)

        self.edit_composition = QPlainTextEdit()
        self.edit_composition.setPlaceholderText("例如：石蜡(40%)、白蜂蜡(17.5%)")
        self.edit_composition.setMaximumHeight(80)
        form.addRow("成分组成:", self.edit_composition)

        layout.addLayout(form)
        layout.addStretch()

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _on_inci_changed(self, text):
        """INCI名称变化时自动匹配CAS号"""
        if self.edit_cas.text().strip():
            return  # 已有CAS号不覆盖
        cas = self.catalog.match_cas(text)
        if cas:
            self.edit_cas.setText(cas)

    def _load_data(self):
        m = self.material
        self.edit_name.setText(m['name_zh'])
        self.edit_code.setText(m.get('internal_code', ''))
        self.edit_purpose.setText(m.get('default_purpose', ''))
        self.edit_supplier.setText(m.get('supplier_code', ''))
        self.edit_supplier_name.setText(m.get('supplier_name', ''))
        self.edit_cas.setText(m.get('cas_number', ''))
        self.edit_remarks.setText(m.get('remarks', ''))
        self.edit_inci.setText(m.get('name_inci', ''))
        comps = json.loads(m['composition']) if m['composition'] else []
        parts = []
        for c in comps:
            parts.append(f"{c['name']}({c['percent']}%)")
        self.edit_composition.setText('、'.join(parts))

    def get_data(self):
        name = self.edit_name.text().strip()
        internal_code = self.edit_code.text().strip()
        purpose = self.edit_purpose.text().strip()
        supplier_code = self.edit_supplier.text().strip()
        supplier_name = self.edit_supplier_name.text().strip()
        cas_number = self.edit_cas.text().strip()
        remarks = self.edit_remarks.text().strip()
        name_inci = self.edit_inci.text().strip()
        comp_str = self.edit_composition.toPlainText().strip()

        comps, comp_incis = _build_composition_from_natural(name, comp_str, self.catalog)

        if not name_inci:
            name_inci = self.catalog.match_inci(name)
        if not name_inci and comp_incis:
            name_inci = '、'.join(comp_incis)

        return {
            'name_zh': name,
            'name_inci': name_inci,
            'composition': json.dumps(comps, ensure_ascii=False),
            'default_purpose': purpose,
            'supplier_code': supplier_code,
            'supplier_name': supplier_name,
            'internal_code': internal_code,
            'cas_number': cas_number,
            'remarks': remarks,
            'trade_name': ''
        }


# ── Formula Design Tab ──

class FormulaDesignTab(QWidget):
    def __init__(self, catalog):
        super().__init__()
        self.catalog = catalog
        self._materials_cache = []
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # Top bar: search + add
        top = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索原料（名称/代码）...")
        self.completer = QCompleter()
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.completer.setFilterMode(Qt.MatchContains)
        self.search_input.setCompleter(self.completer)
        self.search_input.textChanged.connect(self._on_search_changed)
        top.addWidget(self.search_input, 1)

        self.spin_percent = QDoubleSpinBox()
        self.spin_percent.setRange(0, 100)
        self.spin_percent.setDecimals(2)
        self.spin_percent.setValue(1.0)
        self.spin_percent.setSuffix("%")
        self.spin_percent.setFixedWidth(100)
        top.addWidget(QLabel("添加量:"))
        top.addWidget(self.spin_percent)

        self.btn_add = QPushButton("添加")
        self.btn_add.clicked.connect(self._on_add)
        top.addWidget(self.btn_add)
        layout.addLayout(top)

        # Batch input bar
        batch_bar = QHBoxLayout()
        batch_bar.addWidget(QLabel("批量导入内部代码:"))
        self.batch_input = QLineEdit()
        self.batch_input.setPlaceholderText("用逗号、顿号或空格分隔")
        batch_bar.addWidget(self.batch_input, 1)
        self.btn_batch_add = QPushButton("批量添加")
        self.btn_batch_add.clicked.connect(self._on_batch_add)
        batch_bar.addWidget(self.btn_batch_add)
        layout.addLayout(batch_bar)

        # Formula table
        self.table = QTableWidget(0, 13)
        self.table.setHorizontalHeaderLabels([
            "原料序号", "内部代码", "标准中文名称", "INCI名称", "原料含量(%)",
            "原料中成分含量(%)", "实际成分含量(%)", "主要使用目的",
            "原料报送码", "备注", "是否新原料", "注册号/备案号", "操作"
        ])
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(7, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(9, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(10, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(11, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(12, QHeaderView.ResizeToContents)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked)
        self.table.verticalHeader().setVisible(False)
        self.table.setWordWrap(True)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.cellChanged.connect(self._on_cell_changed)
        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #d0d7de;
                gridline-color: #d0d7de;
            }
            QTableWidget::item {
                border: 1px solid #e5e7eb;
            }
            QTableWidget::item:selected {
                background-color: #dbeafe;
            }
            QHeaderView::section {
                background-color: #f3f4f6;
                border: 1px solid #d0d7de;
                padding: 4px;
            }
        """)
        layout.addWidget(self.table)

        # Bottom buttons
        bottom = QHBoxLayout()
        self.btn_clear = QPushButton("清空配方")
        self.btn_clear.clicked.connect(self._on_clear)
        bottom.addWidget(self.btn_clear)

        # v2 S8: 配方总含量实时提示
        self._total_percent_label = QLabel('总含量：—')
        self._total_percent_label.setStyleSheet(
            "padding: 4px 10px; border-radius: 4px; font-weight: bold; font-size: 12px;"
        )
        bottom.addWidget(self._total_percent_label)
        bottom.addStretch()
        
        bottom.addStretch()
        
        self.btn_export = QPushButton("一键导出备案配方表")
        self.btn_export.clicked.connect(self._on_export)
        bottom.addWidget(self.btn_export)
        
        self.btn_export_label = QPushButton("生成标签成分表")
        self.btn_export_label.clicked.connect(self._on_export_label)
        bottom.addWidget(self.btn_export_label)
        
        layout.addLayout(bottom)

    def refresh(self):
        self._materials_cache = db.get_all_raw_materials()
        self._reload_table()
        self._update_total_percent()

    def _reload_table(self):
        formula = db.get_current_formula()
        self.table.blockSignals(True)
        # Build expanded rows (similar to Excel export)
        expanded_rows = []
        for idx, item in enumerate(formula):
            comps = json.loads(item['composition'])
            use_purpose = item['purpose_override'] or item['default_purpose'] or ''
            # First component row
            expanded_rows.append({
                'row_type': 'main',
                'formula_id': item['id'],
                'raw_material_id': item['raw_material_id'],
                '原料序号': idx + 1,
                '内部代码': item.get('internal_code', '') or '',
                '标准中文名称': comps[0]['name'],
                'INCI名称': comps[0].get('inci', '') or item['name_inci'] or '',
                '原料含量(%)': item['added_percent'],
                '原料中成分含量(%)': comps[0]['percent'],
                '实际成分含量(%)': round(item['added_percent'] * comps[0]['percent'] / 100, 4),
                '主要使用目的': use_purpose,
                '原料报送码': item['supplier_code'] or '',
                '备注': item['remarks'] or '',
                '是否新原料': item['is_new_material'] or 0,
                '注册号/备案号': item['registration_number'] or ''
            })
            # Subsequent component rows
            for comp in comps[1:]:
                expanded_rows.append({
                    'row_type': 'comp',
                    'formula_id': item['id'],
                    'raw_material_id': item['raw_material_id'],
                    '原料序号': '',
                    '内部代码': '',
                    '标准中文名称': comp['name'],
                    'INCI名称': comp.get('inci', ''),
                    '原料含量(%)': '',
                    '原料中成分含量(%)': comp['percent'],
                    '实际成分含量(%)': round(item['added_percent'] * comp['percent'] / 100, 4),
                    '主要使用目的': '',
                    '原料报送码': '',
                    '备注': '',
                    '是否新原料': 0,
                    '注册号/备案号': ''
                })

        self.table.setRowCount(len(expanded_rows))
        cols = ['原料序号', '内部代码', '标准中文名称', 'INCI名称', '原料含量(%)', '原料中成分含量(%)',
                '实际成分含量(%)', '主要使用目的', '原料报送码', '备注', '是否新原料', '注册号/备案号']
        merge_cols = [0, 1, 4, 7, 8, 9, 10, 11, 12]

        # Track merge groups: group_id -> {'start': row_idx, 'count': n}
        merge_groups = {}
        group_id = None
        for i, row_data in enumerate(expanded_rows):
            if row_data['row_type'] == 'main':
                group_id = row_data['原料序号']
                merge_groups[group_id] = {'start': i, 'count': 1}
            else:
                if group_id is not None:
                    merge_groups[group_id]['count'] += 1

        for i, row_data in enumerate(expanded_rows):
            for col_idx, col_name in enumerate(cols):
                value = row_data[col_name]
                if col_name in ['原料含量(%)', '原料中成分含量(%)', '实际成分含量(%)']:
                    if value != '':
                        item = QTableWidgetItem(f"{value:.4f}" if col_name == '实际成分含量(%)' else f"{value:.2f}")
                    else:
                        item = QTableWidgetItem('')
                elif col_name == '是否新原料':
                    item = QTableWidgetItem('是' if value else '')
                else:
                    item = QTableWidgetItem(str(value) if value is not None else '')
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(i, col_idx, item)

            # Add delete button only on main rows
            if row_data['row_type'] == 'main':
                btn_w = QWidget()
                btn_l = QHBoxLayout(btn_w)
                btn_l.setContentsMargins(0, 0, 0, 0)
                btn_del = QPushButton("删除")
                btn_del.setFixedWidth(50)
                fid = row_data['formula_id']
                btn_del.clicked.connect(lambda checked, fid=fid: self._on_remove(fid))
                btn_l.addWidget(btn_del)
                self.table.setCellWidget(i, 12, btn_w)

        # Apply cell merging for multi-component materials
        for gid, info in merge_groups.items():
            if info['count'] > 1:
                start = info['start']
                span = info['count']
                for col in merge_cols:
                    self.table.setSpan(start, col, span, 1)

        self._expanded_rows = expanded_rows
        self._merge_groups = merge_groups
        self.table.blockSignals(False)

    def _on_cell_changed(self, row, col):
        if row >= len(self._expanded_rows):
            return
        row_data = self._expanded_rows[row]
        if row_data['row_type'] != 'main':
            return

        formula_id = row_data['formula_id']
        item = self.table.item(row, col)
        if not item:
            return
        new_text = item.text().strip()

        if col == 4:
            # 原料含量(%)
            try:
                new_pct = float(new_text)
            except ValueError:
                self._reload_table()
                return
            db.update_formula_percent(formula_id, new_pct)
            seq = row_data['原料序号']
            group = self._merge_groups.get(seq)
            if not group:
                return
            for i in range(group['start'], group['start'] + group['count']):
                rd = self._expanded_rows[i]
                comp_pct = rd['原料中成分含量(%)']
                if comp_pct != '':
                    actual = round(new_pct * float(comp_pct) / 100, 4)
                    actual_item = self.table.item(i, 6)
                    if actual_item:
                        actual_item.setText(f"{actual:.4f}")

        elif col == 7:
            # 主要使用目的 → purpose_override
            # Read current values from other editable cells
            reg_item = self.table.item(row, 11)
            is_new_item = self.table.item(row, 10)
            reg = reg_item.text().strip() if reg_item else row_data['注册号/备案号']
            is_new = is_new_item.text().strip() if is_new_item else ('是' if row_data['是否新原料'] else '')
            db.update_formula_material_info(formula_id, 1 if is_new == '是' else 0, reg, new_text)

        elif col == 9:
            # 备注 → remarks
            db.update_formula_remarks(formula_id, new_text)

        elif col == 10:
            # 是否新原料 → is_new_material
            reg_item = self.table.item(row, 11)
            purpose_item = self.table.item(row, 7)
            reg = reg_item.text().strip() if reg_item else row_data['注册号/备案号']
            purpose = purpose_item.text().strip() if purpose_item else row_data['主要使用目的']
            is_new = 1 if new_text == '是' else 0
            db.update_formula_material_info(formula_id, is_new, reg, purpose)

        elif col == 11:
            # 注册号/备案号 → registration_number
            purpose_item = self.table.item(row, 7)
            is_new_item = self.table.item(row, 10)
            purpose = purpose_item.text().strip() if purpose_item else row_data['主要使用目的']
            is_new = is_new_item.text().strip() if is_new_item else ('是' if row_data['是否新原料'] else '')
            db.update_formula_material_info(formula_id, 1 if is_new == '是' else 0, new_text, purpose)

        # v2 S8: 任意单元格变化后重算总含量
        self._update_total_percent()

    def _update_total_percent(self):
        """重算配方总含量并更新标签颜色。
        - 总含量 = 100% ± 0.01% → 绿色 ✓
        - 偏离 0.01-1% → 黄色 ⚠
        - 偏离 > 1% 或为 0% → 红色 ❌
        """
        if not hasattr(self, '_total_percent_label'):
            return
        formula = db.get_current_formula()
        if not formula:
            self._total_percent_label.setText('总含量：—')
            self._total_percent_label.setStyleSheet(
                "padding: 4px 10px; border-radius: 4px; font-weight: bold; font-size: 12px;"
                "background-color: #f3f4f6; color: #6b7280;"
            )
            return

        total = 0.0
        for item in formula:
            try:
                pct = float(item.get('added_percent') or 0)
                total += pct
            except (ValueError, TypeError):
                pass

        diff = abs(total - 100.0)
        if total == 0:
            label = '总含量：0.00%'
            bg, fg, icon = '#fee2e2', '#dc2626', '❌'
        elif diff <= 0.01:
            label = f'总含量：{total:.2f}%  {icon if False else "✓ 合格"}'
            bg, fg, icon = '#dcfce7', '#16a34a', '✓'
        elif diff <= 1.0:
            label = f'总含量：{total:.2f}%  ⚠ 偏离 {diff:.2f}%'
            bg, fg = '#fef3c7', '#ca8a04'
        else:
            label = f'总含量：{total:.2f}%  ❌ 偏离 {diff:.2f}%'
            bg, fg = '#fee2e2', '#dc2626'

        self._total_percent_label.setText(label)
        self._total_percent_label.setStyleSheet(
            f"padding: 4px 10px; border-radius: 4px; font-weight: bold; font-size: 12px;"
            f"background-color: {bg}; color: {fg};"
        )

    def _on_search_changed(self, text):
        if len(text) < 1:
            self.completer.setModel(QStringListModel([]))
            return
        results = []
        for m in self._materials_cache:
            if text.lower() in m['name_zh'].lower() or text.lower() in (m.get('internal_code') or '').lower():
                code = m.get('internal_code', '') or ''
                tag = f" [{code}]" if code else ""
                results.append(f"{m['id']}:{m['name_zh']}{tag}")
                if len(results) >= 10:
                    break
        self.completer.setModel(QStringListModel(results))
        if results:
            self.completer.complete()

    def _on_add(self):
        text = self.search_input.text().strip()
        if not text:
            QMessageBox.warning(self, "提示", "请先搜索并选择原料")
            return
        parts = text.split(':', 1)
        if parts[0].isdigit():
            mat_id = int(parts[0])
        else:
            for m in self._materials_cache:
                if m['name_zh'] == text or (m.get('internal_code') or '') == text:
                    mat_id = m['id']
                    break
            else:
                QMessageBox.warning(self, "提示", f"未找到原料: {text}")
                return
        mat = db.get_raw_material_by_id(mat_id)
        if mat:
            is_new = 0 if self.catalog.in_catalog_i(mat['name_zh']) else 1
        else:
            is_new = 1
        percent = self.spin_percent.value()
        db.add_to_formula(mat_id, percent, is_new)
        self._reload_table()
        self._update_total_percent()

    def _on_batch_add(self):
        text = self.batch_input.text().strip()
        if not text:
            QMessageBox.warning(self, "提示", "请输入内部代码")
            return

        import re
        codes = [c.strip() for c in re.split(r'[,，、\s]+', text) if c.strip()]
        if not codes:
            QMessageBox.warning(self, "提示", "未识别到有效的内部代码")
            return

        percent = self.spin_percent.value()
        added = []
        not_found = []
        for code in codes:
            for m in self._materials_cache:
                if m.get('internal_code', '') == code:
                    is_new = 0 if self.catalog.in_catalog_i(m['name_zh']) else 1
                    db.add_to_formula(m['id'], percent, is_new)
                    added.append(f"{m['name_zh']}({code})")
                    break
            else:
                not_found.append(code)

        msg = f"成功添加 {len(added)} 条"
        if not_found:
            msg += f"\n未找到: {', '.join(not_found)}"
        QMessageBox.information(self, "批量添加结果", msg)
        self.batch_input.clear()
        self._reload_table()
        self._update_total_percent()

    def _on_remove(self, fid):
        db.remove_from_formula(fid)
        self._reload_table()
        self._update_total_percent()

    def _on_clear(self):
        if QMessageBox.question(self, "确认", "确定清空全部配方？") == QMessageBox.Yes:
            db.clear_formula()
            self._reload_table()
            self._update_total_percent()

    def _on_export(self):
        formula = db.get_current_formula()
        if not formula:
            QMessageBox.warning(self, "提示", "当前配方为空，请先添加原料")
            return
        try:
            out = _export_formula_to_excel(formula, self.catalog)
            filename = f"备案配方表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            path, _ = QFileDialog.getSaveFileName(self, "保存备案配方表", filename, "Excel文件 (*.xlsx)")
            if path:
                with open(path, 'wb') as f:
                    f.write(out.getvalue())
                QMessageBox.information(self, "成功", f"导出成功: {os.path.basename(path)}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e) + "\n" + traceback.format_exc())

    def _on_export_label(self):
        formula = db.get_current_formula()
        if not formula:
            QMessageBox.warning(self, "提示", "当前配方为空，请先添加原料")
            return
        try:
            label_text = _generate_label_ingredients(formula, self.catalog)
            # 打开预览对话框
            dlg = LabelPreviewDialog(label_text, self)
            dlg.exec()
        except Exception as e:
            QMessageBox.critical(self, "生成失败", str(e) + "\n" + traceback.format_exc())


# ── Three Tables Preview Dialog ──

class ThreeTablesPreviewDialog(QDialog):
    """\u9884\u89c8 3 \u5927\u62a5\u544a\u8868 + \u7b2c\u56db\u90e8\u5206\uff08\u914d\u65b9\u8868 / \u5b9e\u9645\u6210\u5206\u542b\u91cf / \u98ce\u9669\u7269\u8d28\u5371\u5bb3\u8bc6\u522b / \u6210\u5206\u5b89\u5168\u8bc4\u4f30\uff09"""

    def __init__(self, tables: dict, section4: dict, parent=None):
        super().__init__(parent)
        self.tables = tables
        self.section4 = section4
        self.setWindowTitle('\u9884\u89c8\u62a5\u544a\u6838\u5fc3\u5185\u5bb9')
        self.resize(1000, 700)

        layout = QVBoxLayout(self)

        self.tabs = QTabWidget()
        for key in ('table1', 'table2', 'table3'):
            tab = tables[key]
            page = QWidget()
            v = QVBoxLayout(page)

            title = QLabel(tab['title'])
            title.setStyleSheet("font-weight: bold; font-size: 14px; padding: 6px;")
            title.setAlignment(Qt.AlignCenter)
            v.addWidget(title)

            table = QTableWidget()
            headers = tab['headers']
            rows = tab['rows']
            table.setColumnCount(len(headers))
            table.setRowCount(len(rows))
            table.setHorizontalHeaderLabels(headers)
            table.verticalHeader().setVisible(False)
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            table.setSelectionBehavior(QAbstractItemView.SelectRows)
            table.setAlternatingRowColors(True)

            header = table.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.Stretch)
            header.setStyleSheet(
                "QHeaderView::section { background: #dbeafe; font-weight: bold; padding: 6px; }"
            )

            for r, row_data in enumerate(rows):
                for c, val in enumerate(row_data):
                    item = QTableWidgetItem(str(val) if val is not None else '')
                    item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(r, c, item)

            table.resizeRowsToContents()
            v.addWidget(table)

            if key == 'table3':
                src_html = (
                    "<b>\u6570\u636e\u6765\u6e90\u8bf4\u660e\uff1a</b><br>"
                    "\u2022 <b>\u6807\u51c6\u4e2d\u6587\u540d\u79f0</b>\uff1a"
                    "safety_engine.assess_formula() \u8fd4\u56de\u7684 ingredient \u5b57\u6bb5\uff1b"
                    "\u98ce\u9669\u7269\u8d28\u884c\u76f4\u63a5\u4f7f\u7528\u98ce\u9669\u7269\u8d28\u672c\u8eab\u540d\u79f0\u3002<br>"
                    "\u2022 <b>\u53ef\u80fd\u542b\u6709\u7684\u98ce\u9669\u7269\u8d28</b>\uff1a"
                    "database.check_banned() / check_restricted() \u7b49\u7981\u7528/\u9650\u7528\u8868\u8bc6\u522b\u7ed3\u679c\uff1b"
                    "\u65e0\u98ce\u9669\u5219\u6807\u8bb0\u4e3a\u201c\u901a\u8fc7\u5b89\u5168\u8bc4\u4f30\u201d\u3002<br>"
                    "\u2022 <b>\u5907\u6ce8</b>\uff1a"
                    "\u6839\u636e risk_substances \u5217\u8868\u662f\u5426\u4e3a\u7a7a\u5224\u65ad\uff1a"
                    "\u201c\u7b26\u5408\u9650\u91cf\u8981\u6c42\u201d\u6216\u201c\u65e0\u98ce\u9669\u201d\u3002"
                )
                src_label = QLabel(src_html)
                src_label.setWordWrap(True)
                src_label.setTextFormat(Qt.RichText)
                src_label.setStyleSheet(
                    "color: #475569; padding: 8px; font-size: 11px; "
                    "background: #f8fafc; border-left: 3px solid #3b82f6;"
                )
                v.addWidget(src_label)

            self.tabs.addTab(page, tab['title'])

        # 4th tab: Section 4 — per-ingredient safety assessment
        self.tabs.addTab(self._build_section4_page(section4), section4['title'])

        layout.addWidget(self.tabs)

        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _build_section4_page(self, section4: dict) -> QWidget:
        page = QWidget()
        v = QVBoxLayout(page)

        title = QLabel(section4['title'])
        title.setStyleSheet("font-weight: bold; font-size: 14px; padding: 6px;")
        title.setAlignment(Qt.AlignCenter)
        v.addWidget(title)

        text_browser = QTextBrowser()
        text_browser.setOpenExternalLinks(True)
        text_browser.setStyleSheet(
            "QTextBrowser { background: #ffffff; padding: 8px; font-size: 12pt; "
            "font-family: 'Times New Roman', '宋体'; }"
        )

        html_parts = ['<html><body style="line-height: 1.6;">']
        for p in section4['paragraphs']:
            if p['type'] == 'single':
                html_parts.append(
                    f"<p style='margin: 8px 0 4px 0; text-indent: 2em;'>"
                    f"<b>{p['material_idx']}\u53f7\u539f\u6599\uff1a</b>"
                    f"{p['name']}\uff0c{p['text']}</p>"
                )
            elif p['type'] == 'composite_head':
                comps = '\u3001'.join(p['components'][:-1]) + '\u548c' + p['components'][-1]
                html_parts.append(
                    f"<p style='margin: 8px 0 4px 0; text-indent: 2em;'>"
                    f"<b>{p['material_idx']}\u53f7\u539f\u6599\uff1a</b>{comps}\u7684\u6df7\u5408\u7269\u3002</p>"
                )
            elif p['type'] == 'composite_sub':
                html_parts.append(
                    f"<p style='margin: 4px 0; text-indent: 2em;'>"
                    f"{p['name']}\uff0c{p['text']}</p>"
                )

        if section4.get('references'):
            html_parts.append(
                "<hr style='margin: 12px 0; border: none; border-top: 1px solid #e5e7eb;'>"
                "<p style='color: #475569; font-size: 10pt;'>"
                "<b>\u53c2\u8003\u6587\u732e\uff1a</b></p><ol style='font-size: 10pt; color: #475569;'>"
            )
            for ref in section4['references']:
                html_parts.append(f"<li>{ref}</li>")
            html_parts.append('</ol>')

        html_parts.append('</body></html>')
        text_browser.setHtml('\n'.join(html_parts))
        v.addWidget(text_browser)

        return page


# ── Label Preview Dialog ──

class LabelPreviewDialog(QDialog):
    """标签成分表预览对话框"""
    
    def __init__(self, label_text, parent=None):
        super().__init__(parent)
        self.label_text = label_text
        self.setWindowTitle("标签成分表预览")
        self.setMinimumSize(800, 600)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # 标题
        title_label = QLabel("化妆品标签成分表")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #1a73e8;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 预览区域
        self.preview_text = QPlainTextEdit()
        self.preview_text.setPlainText(self.label_text)
        self.preview_text.setReadOnly(True)
        self.preview_text.setStyleSheet("""
            QPlainTextEdit {
                font-family: "宋体", "SimSun", sans-serif;
                font-size: 14px;
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 10px;
            }
        """)
        layout.addWidget(self.preview_text)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        layout.addWidget(line)
        
        # 统计信息
        stats_layout = QHBoxLayout()
        stats_layout.addWidget(QLabel("成分数量："))
        self.stats_label = QLabel()
        self.update_stats()
        stats_layout.addWidget(self.stats_label)
        stats_layout.addStretch()
        layout.addLayout(stats_layout)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        self.btn_copy = QPushButton("复制到剪贴板")
        self.btn_copy.clicked.connect(self._on_copy)
        btn_layout.addWidget(self.btn_copy)
        
        self.btn_export = QPushButton("导出为文本文件")
        self.btn_export.clicked.connect(self._on_export)
        btn_layout.addWidget(self.btn_export)
        
        self.btn_close = QPushButton("关闭")
        self.btn_close.clicked.connect(self.close)
        btn_layout.addWidget(self.btn_close)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
    
    def update_stats(self):
        """更新统计信息"""
        parts = self.label_text.split('\n\n【合规说明】')
        if parts:
            main_text = parts[0]
            # 统计成分数量（逗号分隔）
            if '成分表：' in main_text:
                ingredients = main_text.replace('成分表：', '').strip()
                if ingredients:
                    count = len([x for x in ingredients.split(',') if x.strip()])
                    self.stats_label.setText(f"{count} 种")
                else:
                    self.stats_label.setText("0 种")
    
    def _on_copy(self):
        """复制到剪贴板"""
        clipboard = QApplication.clipboard()
        clipboard.setText(self.label_text)
        QMessageBox.information(self, "成功", "已复制到剪贴板")
    
    def _on_export(self):
        """导出为文本文件"""
        filename = f"标签成分表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        path, _ = QFileDialog.getSaveFileName(self, "保存标签成分表", filename, "文本文件 (*.txt)")
        if path:
            with open(path, 'w', encoding='utf-8') as f:
                f.write(self.label_text)
            QMessageBox.information(self, "成功", f"导出成功: {os.path.basename(path)}")


# ── Material Management Tab ──

class MaterialManagementTab(QWidget):
    def __init__(self, catalog):
        super().__init__()
        self.catalog = catalog
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # Top bar
        top = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索原料...")
        self.search_input.textChanged.connect(self._do_search)
        top.addWidget(self.search_input, 1)

        self.btn_add = QPushButton("添加原料")
        self.btn_add.clicked.connect(self._on_add)
        top.addWidget(self.btn_add)

        self.btn_fill_inci = QPushButton("批量自动匹配INCI")
        self.btn_fill_inci.clicked.connect(self._on_fill_inci)
        top.addWidget(self.btn_fill_inci)

        self.btn_fill_cas = QPushButton("批量匹配CAS号")
        self.btn_fill_cas.clicked.connect(self._on_fill_cas)
        top.addWidget(self.btn_fill_cas)

        self.btn_download_template = QPushButton("下载导入模板")
        self.btn_download_template.clicked.connect(self._on_download_template)
        top.addWidget(self.btn_download_template)

        self.btn_import_excel = QPushButton("从Excel导入")
        self.btn_import_excel.clicked.connect(self._on_import_excel)
        top.addWidget(self.btn_import_excel)

        self.btn_batch_delete = QPushButton("批量删除")
        self.btn_batch_delete.clicked.connect(self._on_batch_delete)
        top.addWidget(self.btn_batch_delete)

        layout.addLayout(top)

        # Table
        self.table = QTableWidget(0, 11)
        self.table.setHorizontalHeaderLabels([
            "选择", "ID", "内部代码", "原料名称/商品名", "INCI名称",
            "成分组成", "主要使用目的", "原料报送码", "备注", "CAS号", "生产商"
        ])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.table.setColumnWidth(0, 40)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(10, QHeaderView.ResizeToContents)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setWordWrap(True)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked)
        self._row_id_map = {}
        self.table.cellChanged.connect(self._on_material_cell_changed)
        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #d0d7de;
                gridline-color: #d0d7de;
            }
            QTableWidget::item {
                border: 1px solid #e5e7eb;
            }
            QTableWidget::item:selected {
                background-color: #dbeafe;
            }
            QHeaderView::section {
                background-color: #f3f4f6;
                border: 1px solid #d0d7de;
                padding: 4px;
            }
        """)
        layout.addWidget(self.table)

    def refresh(self):
        self._do_search()

    def _do_search(self):
        keyword = self.search_input.text().strip().lower()
        materials = db.get_all_raw_materials()
        if keyword:
            materials = [
                m for m in materials
                if keyword in m['name_zh'].lower()
                or keyword in (m.get('internal_code') or '').lower()
                or keyword in (m.get('name_inci') or '').lower()
            ]

        self.table.blockSignals(True)
        self.table.setRowCount(len(materials))
        self._row_id_map.clear()
        for i, m in enumerate(materials):
            comps = json.loads(m['composition']) if m['composition'] else []
            display_name = m['name_zh']
            if m.get('trade_name'):
                display_name += f" ({m['trade_name']})"

            def _ci(text):
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignCenter)
                return item

            cb = QCheckBox()
            cb.setStyleSheet("margin-left:8px;")
            self.table.setCellWidget(i, 0, cb)
            self._row_id_map[i] = m['id']
            self.table.setItem(i, 1, _ci(str(m['id'])))
            self.table.setItem(i, 2, _ci(m.get('internal_code', '')))
            self.table.setItem(i, 3, _ci(display_name))
            self.table.setItem(i, 4, _ci(m.get('name_inci', '')))
            self.table.setItem(i, 5, _ci(_comp_str_to_display(comps)))
            self.table.item(i, 5).setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.table.setRowHeight(i, max(self.table.rowHeight(i), 40))
            self.table.setItem(i, 6, _ci(m.get('default_purpose', '')))
            self.table.setItem(i, 7, _ci(m.get('supplier_code', '')))
            self.table.setItem(i, 8, _ci(m.get('remarks', '')))
            self.table.setItem(i, 9, _ci(m.get('cas_number', '')))
            self.table.setItem(i, 10, _ci(m.get('supplier_name', '')))
        self.table.blockSignals(False)

    def _on_material_cell_changed(self, row, col):
        if row not in self._row_id_map:
            return
        editable = {2, 3, 4, 6, 7, 8, 9, 10}
        if col not in editable:
            return
        rid = self._row_id_map[row]
        # Re-read full row data from table

        def _txt(c):
            it = self.table.item(row, c)
            return it.text().strip() if it else ''
        internal_code = _txt(2)
        name_zh = _txt(3)
        name_inci = _txt(4)
        default_purpose = _txt(6)
        supplier_code = _txt(7)
        remarks = _txt(8)
        cas_number = _txt(9)
        supplier_name = _txt(10)
        # Auto-match CAS when INCI name changes and CAS cell is empty
        if col == 4 and not cas_number:
            matched_cas = self.catalog.match_cas(name_inci)
            if not matched_cas and name_zh:
                matched_cas = self.catalog.match_cas(name_zh)
            if matched_cas:
                self.table.blockSignals(True)
                self.table.setItem(row, 9, QTableWidgetItem(matched_cas))
                self.table.blockSignals(False)
                cas_number = matched_cas
        # Check for duplicate name (excluding self)
        if col == 3 and name_zh:
            all_mats = db.get_all_raw_materials()
            for m in all_mats:
                if m['id'] != rid and m['name_zh'] == name_zh:
                    QMessageBox.warning(
                        self, "重复检测",
                        f"原料名称 '{name_zh}' 已存在（ID: {m['id']}，内部代码: "
                        f"{m.get('internal_code', '无')}）。\n\n"
                        f"同名但内部编码不同可能为不同供应商的原料，请确认是否继续。"
                    )
                    # Revert the cell to original value
                    mat = db.get_raw_material_by_id(rid)
                    if mat:
                        self.table.blockSignals(True)
                        self.table.item(row, col).setText(mat['name_zh'])
                        self.table.blockSignals(False)
                    return
        # Rebuild composition with new name if name changed
        mat = db.get_raw_material_by_id(rid)
        if not mat:
            return
        old_comps = json.loads(mat['composition']) if mat['composition'] else []
        new_comps = []
        for c in old_comps:
            new_comps.append({
                'name': name_zh if c['name'] == mat['name_zh'] else c['name'],
                'percent': c['percent'],
                'inci': c.get('inci', ''),
            })
        db.update_raw_material(
            rid, name_zh, name_inci, json.dumps(new_comps, ensure_ascii=False),
            default_purpose, supplier_code,
            remarks, mat.get('trade_name', ''),
            internal_code, supplier_name, cas_number
        )

    def _on_add(self):
        dlg = MaterialEditDialog(self.catalog, None, self)
        if dlg.exec():
            data = dlg.get_data()
            db.add_raw_material(
                data['name_zh'], data['name_inci'], data['composition'],
                data['default_purpose'], data['supplier_code'],
                data['remarks'], data['trade_name'], data['internal_code'],
                data['supplier_name'], data['cas_number']
            )
            self._do_search()

    def _on_edit(self, rid):
        mat = db.get_raw_material_by_id(rid)
        if not mat:
            return
        dlg = MaterialEditDialog(self.catalog, mat, self)
        if dlg.exec():
            data = dlg.get_data()
            db.update_raw_material(
                rid, data['name_zh'], data['name_inci'], data['composition'],
                data['default_purpose'], data['supplier_code'],
                data['remarks'], data['trade_name'], data['internal_code'],
                data['supplier_name'], data['cas_number']
            )
            self._do_search()

    def _on_delete(self, rid):
        mat = db.get_raw_material_by_id(rid)
        if not mat:
            return
        if QMessageBox.question(self, "确认", f"确定删除原料 '{mat['name_zh']}'？") == QMessageBox.Yes:
            db.delete_raw_material(rid)
            self._do_search()

    def _on_batch_delete(self):
        ids = []
        for row in range(self.table.rowCount()):
            cb = self.table.cellWidget(row, 0)
            if cb and cb.isChecked():
                id_item = self.table.item(row, 1)
                if id_item:
                    ids.append(int(id_item.text()))
        if not ids:
            QMessageBox.information(self, "提示", "请先勾选要删除的原料")
            return
        if QMessageBox.question(self, "确认", f"确定删除选中的 {len(ids)} 条原料？") == QMessageBox.Yes:
            for rid in ids:
                db.delete_raw_material(rid)
            self._do_search()

    def _on_fill_inci(self):
        materials = db.get_all_raw_materials()
        updated = 0
        updated_comps = 0
        updated_cas = 0
        for m in materials:
            composition = json.loads(m['composition']) if m['composition'] else []
            comp_incis = []
            comp_updated = False
            for comp in composition:
                cn = comp.get('name', '')
                if cn:
                    inci = self.catalog.match_inci(cn)
                    if inci and not comp.get('inci'):
                        comp['inci'] = inci
                        comp_updated = True
                        updated_comps += 1
                    if inci and inci not in comp_incis:
                        comp_incis.append(inci)
            if comp_updated:
                db.update_raw_material_composition(m['id'], json.dumps(composition, ensure_ascii=False))
            if not m['name_inci'] and m['name_zh']:
                inci = self.catalog.match_inci(m['name_zh'])
                if not inci and comp_incis:
                    inci = '、'.join(comp_incis)
                if inci:
                    db.update_raw_material_inci(m['id'], inci)
                    updated += 1
            # 自动匹配CAS号
            if not m.get('cas_number') and m['name_inci']:
                cas = self.catalog.match_cas(m['name_inci'])
                if not cas and m['name_zh']:
                    cas = self.catalog.match_cas(m['name_zh'])
                if cas:
                    db.update_raw_material_cas(m['id'], cas)
                    updated_cas += 1
        msg = f"更新了 {updated} 条原料的INCI名, {updated_comps} 个成分的INCI名"
        if updated_cas:
            msg += f", {updated_cas} 条原料的CAS号"
        QMessageBox.information(self, "完成", msg)

    def _on_fill_cas(self):
        materials = db.get_all_raw_materials()
        updated = 0
        for m in materials:
            if m.get('cas_number'):
                continue
            cas = ''
            if m.get('name_inci'):
                cas = self.catalog.match_cas(m['name_inci'])
            if not cas and m.get('name_zh'):
                cas = self.catalog.match_cas(m['name_zh'])
            if cas:
                db.update_raw_material_cas(m['id'], cas)
                updated += 1
        QMessageBox.information(self, "完成", f"批量匹配CAS号完成，更新了 {updated} 条原料")

    def _on_download_template(self):
        path, _ = QFileDialog.getSaveFileName(self, "保存导入模板", "原料导入模板.xlsx", "Excel文件 (*.xlsx)")
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "原料导入模板"
        ws.append(['原料名称/商品名', '成分组成', '内部代码', '主要使用目的', '原料报送码', '备注', 'CAS号', '生产商'])
        for col, w in [('A', 15), ('B', 50), ('C', 12), ('D', 15), ('E', 15), ('F', 20), ('G', 15), ('H', 20)]:
            ws.column_dimensions[col].width = w
        wb.save(path)
        QMessageBox.information(self, "成功", f"模板已保存: {os.path.basename(path)}")

    def _on_import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)")
        if not path:
            return
        try:
            # 局部导入pandas，加快启动速度
            import pandas as pd
            df = pd.read_excel(path)
            success = 0
            errors = []
            for idx, row in df.iterrows():
                try:
                    raw_name = str(row.get('原料名称/商品名', '')).strip()
                    if not raw_name or raw_name == 'nan':
                        errors.append(f"第{idx + 2}行：缺少原料名称")
                        continue
                    comp_str = str(row.get('成分组成', '')).strip()
                    comps, comp_incis = _build_composition_from_natural(
                        raw_name,
                        comp_str if comp_str != 'nan' else '',
                        self.catalog
                    )

                    # 自动匹配INCI
                    name_inci = self.catalog.match_inci(raw_name)
                    if not name_inci and comp_incis:
                        name_inci = '、'.join(comp_incis)

                    db.add_raw_material(
                        name_zh=raw_name,
                        name_inci=name_inci,
                        composition_json=json.dumps(comps, ensure_ascii=False),
                        default_purpose=str(row.get('主要使用目的', '')).strip() if pd.notna(row.get('主要使用目的')) else '',
                        supplier_code=str(row.get('原料报送码', '')).strip() if pd.notna(row.get('原料报送码')) else '',
                        remarks=str(row.get('备注', '')).strip() if pd.notna(row.get('备注')) else '',
                        trade_name='',
                        internal_code=str(row.get('内部代码', '')).strip() if pd.notna(row.get('内部代码')) else '',
                        supplier_name=str(row.get('生产商', '')).strip() if pd.notna(row.get('生产商')) else '',
                        cas_number=str(row.get('CAS号', '')).strip() if pd.notna(row.get('CAS号')) else ''
                    )
                    success += 1
                except Exception as e:
                    errors.append(f"第{idx + 2}行：{str(e)}")
            msg = f"导入完成！成功 {success} 条"
            if errors:
                msg += "\n\n错误详情:\n" + "\n".join(errors[:20])
            QMessageBox.information(self, "导入结果", msg)
            self._do_search()
        except Exception as e:
            QMessageBox.critical(self, "导入失败", str(e))


# ── Crawl Progress Tab ──

class CrawlProgressTab(QWidget):
    def __init__(self, catalog):
        super().__init__()
        self.catalog = catalog
        self.worker = None
        self.cas_worker = None
        self._build_ui()
        self._update_catalog_info()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # Catalog info cards
        info_frame = QFrame()
        info_frame.setObjectName("catalogInfo")
        info_frame.setStyleSheet("""
            QFrame#catalogInfo {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #eef4ff, stop:1 #f0f9ff);
                border: 1px solid #c7d7fe;
                border-radius: 8px;
                padding: 14px 18px;
            }
            QLabel {
                font-size: 14px;
                color: #1d2939;
            }
        """)
        info_layout = QHBoxLayout(info_frame)
        self.lbl_cat_i = QLabel("目录I: - 条")
        self.lbl_cat_i.setStyleSheet("font-weight: 700; color: #344054;")
        info_layout.addWidget(self.lbl_cat_i)
        info_layout.addStretch()
        self.lbl_cat_ii = QLabel("目录II: - 条")
        self.lbl_cat_ii.setStyleSheet("font-weight: 700; color: #344054;")
        info_layout.addWidget(self.lbl_cat_ii)
        info_layout.addStretch()
        self.lbl_cas_count = QLabel("CAS号数据: - 条")
        self.lbl_cas_count.setStyleSheet("font-weight: 700; color: #344054;")
        info_layout.addWidget(self.lbl_cas_count)
        layout.addWidget(info_frame)

        # Buttons
        btn_row = QHBoxLayout()
        self.btn_crawl_i = QPushButton("同步目录I（已使用化妆品原料目录）")
        self.btn_crawl_i.clicked.connect(lambda: self._start_crawl('I'))
        btn_row.addWidget(self.btn_crawl_i)
        self.btn_crawl_ii = QPushButton("同步目录II")
        self.btn_crawl_ii.clicked.connect(lambda: self._start_crawl('II'))
        btn_row.addWidget(self.btn_crawl_ii)

        # CAS 获取区域
        btn_row.addStretch()
        cas_label = QLabel("CAS方式:")
        cas_label.setStyleSheet("font-size: 13px;")
        self.cas_combo = QComboBox()
        self.cas_combo.addItems(["COSING + PubChem", "仅 COSING", "仅 PubChem"])
        self.cas_combo.setMinimumWidth(140)
        self.btn_crawl_cas = QPushButton("获取CAS号")
        self.btn_crawl_cas.clicked.connect(self._start_crawl_cas)
        btn_row.addWidget(cas_label)
        btn_row.addWidget(self.cas_combo)
        btn_row.addWidget(self.btn_crawl_cas)
        layout.addLayout(btn_row)

        # Progress
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)

        # Log
        self.log_text = QPlainTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumBlockCount(500)
        layout.addWidget(self.log_text, 1)

    def _update_catalog_info(self):
        info = self.catalog.get_catalog_info()
        self.lbl_cat_i.setText(f"目录I: {info['catalog_i']} 条")
        self.lbl_cat_ii.setText(f"目录II: {info['catalog_ii']} 条")
        # Update CAS count
        from catalog_mgr import USERDATA_DIR
        cas_path = os.path.join(USERDATA_DIR, "cosing_data.json")
        if os.path.exists(cas_path):
            try:
                with open(cas_path, 'r', encoding='utf-8') as f:
                    cas_count = len(json.load(f))
                self.lbl_cas_count.setText(f"CAS号数据: {cas_count} 条")
            except Exception:
                self.lbl_cas_count.setText("CAS号数据: - 条")
        else:
            self.lbl_cas_count.setText("CAS号数据: 0 条")

    def _start_crawl(self, cat_type):
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(self, "提示", "正在爬取中，请稍候...")
            return
        if self.cas_worker and self.cas_worker.isRunning():
            QMessageBox.warning(self, "提示", "正在获取CAS号中，请稍候...")
            return

        self.btn_crawl_i.setEnabled(False)
        self.btn_crawl_ii.setEnabled(False)
        self.btn_crawl_cas.setEnabled(False)
        self.cas_combo.setEnabled(False)
        self.progress_bar.setValue(0)
        self.log_text.clear()

        self.worker = CrawlWorker()
        self.worker.cat_type = cat_type
        self.worker.log_updated.connect(self._on_log)
        self.worker.progress_updated.connect(self._on_progress)
        self.worker.finished_signal.connect(self._on_crawl_finished)
        self.worker.start()

    def _start_crawl_cas(self):
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(self, "提示", "正在爬取中，请稍候...")
            return
        if self.cas_worker and self.cas_worker.isRunning():
            QMessageBox.warning(self, "提示", "正在获取CAS号中，请稍候...")
            return

        self.btn_crawl_i.setEnabled(False)
        self.btn_crawl_ii.setEnabled(False)
        self.btn_crawl_cas.setEnabled(False)
        self.cas_combo.setEnabled(False)
        self.progress_bar.setValue(0)
        self.log_text.clear()

        mode = self.cas_combo.currentIndex()
        self.cas_worker = CasCrawlWorker(mode=mode)
        self.cas_worker.log_updated.connect(self._on_log)
        self.cas_worker.progress_updated.connect(self._on_progress)
        self.cas_worker.finished_signal.connect(self._on_cas_finished)
        self.cas_worker.start()

    def _on_log(self, msg):
        self.log_text.appendPlainText(msg)

    def _on_progress(self, page, total, count):
        pct = int(page / total * 100) if total > 0 else 0
        self.progress_bar.setValue(pct)

    def _on_crawl_finished(self, success, _count):
        self.btn_crawl_i.setEnabled(True)
        self.btn_crawl_ii.setEnabled(True)
        self.btn_crawl_cas.setEnabled(True)
        self.cas_combo.setEnabled(True)
        if success:
            self.catalog.load_all()
            self.catalog.build_indexes()
            self._update_catalog_info()
            self.log_text.appendPlainText("\n爬取完成！")
            self.progress_bar.setValue(100)
        else:
            self.log_text.appendPlainText("\n爬取失败，请检查网络后重试")

    def _on_cas_finished(self, success, count):
        self.btn_crawl_i.setEnabled(True)
        self.btn_crawl_ii.setEnabled(True)
        self.btn_crawl_cas.setEnabled(True)
        self.cas_combo.setEnabled(True)
        if success:
            self.catalog.load_cosing()
            self._update_catalog_info()
            self.log_text.appendPlainText(f"\nCAS号获取完成！共 {count} 条数据")
            self.progress_bar.setValue(100)
        else:
            self.log_text.appendPlainText("\nCAS号获取失败，请检查网络后重试")


"""SafetyAssessmentTab code fragment - to be inserted into app_qt.py"""

# ── Product Category Presets ──
_PRODUCT_CATEGORIES = [
    ('\u9762\u90e8', [
        '\u5316\u5986\u6c34', '\u4e73\u6db2', '\u9762\u971c/\u65e5\u971c', '\u7cbe\u534e\u6db2',
        '\u6d17\u9762\u5976', '\u5378\u5986\u6db2/\u6cb9', '\u7c89\u5e95\u7c89', '\u6563\u7c89/\u5bc6\u7c89',
        '\u53e3\u7ea2', '\u5507\u818f/\u5507\u5f69', '\u773c\u5f71', '\u773c\u7ebf\u7b14/\u773c\u7ebf\u6db2',
        '\u7709\u7b14/\u7709\u7c89', '\u776b\u6bdb\u818f', '\u776b\u6bdb\u6db2',
        '\u9762\u819c', '\u971c/\u971c', '\u62c9\u62c9/\u51dd\u80f6',
    ]),
    ('\u5168\u8eab\u76ae\u80a4', [
        '\u6d17\u6fa1\u9732', '\u6d17\u6fa1\u4e73', '\u6d17\u6fa1\u6ce1', '\u4f53\u4e73/\u4f53\u971c',
        '\u62a4\u624b\u971c', '\u9632\u6652\u971c', '\u6cdb\u6cfd\u6db2/\u6cdb\u6cfd\u6cb9',
        '\u6e05\u6d01\u971c/\u6d17\u9762\u5976', '\u5a74\u513f\u7528\u54c1',
    ]),
    ('\u5934\u53d1', [
        '\u6d17\u53d1\u6c34', '\u62a4\u53d1\u7d20', '\u53d1\u819c', '\u53d1\u9187/\u51dd\u80f6',
        '\u673a\u6cb9/\u7cbe\u534e\u6cb9', '\u5934\u76ae\u62a4\u7406\u4ea7\u54c1',
        '\u67d3\u53d1\u5242\uff08\u6c27\u5316\u578b\uff09', '\u67d3\u53d1\u5242\uff08\u975e\u6c27\u5316\uff09',
    ]),
    ('\u773c\u90e8', [
        '\u773c\u5f71', '\u776b\u6bdb\u818f', '\u773c\u7ebf\u7b14', '\u773c\u971c/\u773c\u818f',
        '\u7709\u7b14/\u7709\u7c89', '\u773c\u88c5\u5378\u5986',
    ]),
    ('\u5634\u5507', [
        '\u5507\u818f/\u5507\u5f69', '\u5507\u91c9/\u5507\u5f69', '\u5507\u91c9', '\u5507\u819c', '\u53e3\u7ea2',
    ]),
    ('\u53e3\u8154', [
        '\u7259\u818f\uff08\u6210\u4eba\uff09', '\u7259\u818f\uff08\u513f\u7ae5\uff09',
        '\u6a61\u76ae\u818f', '\u6a61\u7c89', '\u732b\u556a', '\u5403\u9165\u5439',
        '\u6d17\u9762\u5976', '\u6d17\u9762\u4e73\u6db2', '\u6d17\u9762\u9ecf\u571f',
        '\u5378\u5986\u4e73', '\u9798\u9178/\u8d77\u6ce1\u7f51/\u6d17\u8138\u86cb\u767d/\u6d17\u9762\u6ce1\u6cab\u6db2',
        '\u6d17\u8138\u4e73/\u6d17\u9762\u4e73/\u817b\u5ea6\u6d17\u9762\u5976',
        '\u9762\u819c\uff08\u7247\u72b6\uff09', '\u9762\u819c\uff08\u8131\u62c9\u5f0f\u6216\u6ce5\u819c\uff09',
        '\u6c89\u6dc0\u5f0f\u9762\u819c',
        '\u9762\u971c/\u4e73\u6db2/\u7cbe\u534e\u7d20/\u7cbe\u534e\u6db2',
        '\u7c89\u5e95\u7c89/\u7c89\u7c89/\u5bc6\u7c89',
        '\u949d\u7eba\u7ebf\u53cc\u773c\u76ae / \u5243\u773c\u5f71',
        '\u773c\u7ebf\u6db2/\u773c\u7ebf\u7b14',
        '\u7709\u7b14/\u7709\u7c89/\u7709\u818f',
        '\u53e3\u7ea2/\u5507\u818f/\u5507\u91c9',
        '\u53e3\u7ea2\u6cb9/\u5507\u5f69/\u5507\u91c9/\u5507\u818f/\u5507\u5f69',
        '\u9762\u90e8\u7cbe\u534e\u6db2/\u7cbe\u534e\u7d20/\u6c34\u7cbe\u534e',
        '\u4f53\u4e73/\u4f53\u971c/\u8eab\u4f53\u4e73\u6db2',
        '\u62a4\u624b\u971c/\u62a4\u624b\u971c\u971c/\u62a4\u624b\u971c/\u62a4\u624b\u971c\u971c',
        '\u6d17\u6fa1\u9732/\u6d17\u6fa1\u4e73/\u6d17\u6fa1\u971c',
        '\u6c89\u6dc0\u6d17\u53d1\u6c34/\u6563\u88c5\u6d17\u53d1\u6c34',
        '\u9f13\u98ce\u673a/\u5439\u98ce\u673a',
    ]),
    ('\u624b\u90e8/\u8db3\u90e8', [
        '\u62a4\u624b\u971c/\u62a4\u624b\u971c', '\u62a4\u624b\u971c/\u62a4\u624b\u971c\u971c/\u62a4\u624b\u971c\u971c/\u62a4\u624b\u971c\u971c',
        '\u62a4\u624b\u971c/\u62a4\u624b\u971c\u971c/\u62a4\u624b\u971c\u971c/\u62a4\u624b\u971c\u971c',
        '\u62a4\u624b\u971c/\u62a4\u624b\u971c\u971c/\u62a4\u624b\u971c\u971c/\u62a4\u624b\u971c\u971c',
    ]),
]

_APPLICATION_SITES = [
    '\u9762\u90e8', '\u5168\u8eab\u76ae\u80a4', '\u5934\u53d1', '\u773c\u90e8',
    '\u5634\u5507', '\u53e3\u8154', '\u624b\u90e8/\u8db3\u90e8', '\u6307\u7532/\u8dbd\u7532',
]

_PURPOSE_OPTIONS = [
    '', '\u6d01\u9762', '\u4fdd\u6e7f', '\u9632\u6652', '\u6297\u8870\u8001',
    '\u63d0\u4eae', '\u7f8e\u767d', '\u6536\u7f29\u6bdb\u5b54', '\u53bb\u89d2\u8d28',
    '\u9632\u8150\u5242', '\u9753\u80f6\u5242', '\u9632\u6652\u5242', '\u7740\u8272\u5242',
    '\u67d3\u53d1\u5242', '\u4fdd\u6e7f\u5242', '\u60ee\u5316\u5242', '\u4e0d\u65ad\u5730',
    '\u7a33\u5b9a\u5242', '\u6297\u6c27\u5316\u5242', '\u8774\u6db2\u8c03\u8282\u5242',
    '\u4e3b\u8981\u7528\u9014: \u4fdd\u6e7f',
    '\u4e3b\u8981\u7528\u9014: \u62a4\u53d1',
    '\u4e3b\u8981\u7528\u9014: \u6d17\u53d1',
    '\u4e3b\u8981\u7528\u9014: \u6d17\u9762',
    '\u4e3b\u8981\u7528\u9014: \u4fdd\u6e7f',
    '\u4e3b\u8981\u7528\u9014: \u62a4\u53d1',
    '\u4e3b\u8981\u7528\u9014: \u6d17\u53d1',
]


class SafetyAssessmentTab(QWidget):
    """\u5b89\u5168\u8bc4\u4f30\u6807\u7b7e\u9875 (PRD 5.3 / 5.5)"""

    def __init__(self, catalog):
        super().__init__()
        self.catalog = catalog
        self._results = None
        self._category_cache = {}  # 缓存产品类别信息
        self._expanded_rows = []  # 展开后的行数据
        self._merge_groups = {}   # 合并组信息
        self._build_ui()

    # ── UI Construction ──

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # ── Shared group style ──
        _GROUP_STYLE = """
            QGroupBox {
                font-weight: bold; font-size: 13px;
                border: 1px solid #d0d5dd; border-radius: 6px;
                margin-top: 12px; padding: 12px 8px 8px 8px;
                background: #ffffff;
            }
            QGroupBox::title {
                subcontrol-origin: margin; subcontrol-position: top left;
                padding: 2px 8px; color: #1d2939;
            }
        """

        # ═══════════════════════════════════════════
        # ── Tab widget ──
        # ═══════════════════════════════════════════
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane { border: 1px solid #d0d5dd; border-radius: 4px; padding: 8px; }
            QTabBar::tab { padding: 6px 16px; font-size: 12px; }
            QTabBar::tab:selected { font-weight: bold; border-bottom: 2px solid #1d4ed8; }
        """)

        # ── Tab 1: 产品信息 ──
        tab1 = QWidget()
        tab1_layout = QVBoxLayout(tab1)
        tab1_layout.setSpacing(8)

        grp_basic = QGroupBox('产品基础信息')
        grp_basic.setStyleSheet(_GROUP_STYLE)
        basic_layout = QVBoxLayout(grp_basic)
        basic_layout.setSpacing(6)

        r1 = QHBoxLayout()
        r1.addWidget(QLabel('产品名称:'))
        self.product_name_input = QLineEdit()
        self.product_name_input.setPlaceholderText('输入产品名称...')
        r1.addWidget(self.product_name_input, 1)
        basic_layout.addLayout(r1)

        tab1_layout.addWidget(grp_basic)

        grp_exposure = QGroupBox('暴露参数')
        grp_exposure.setStyleSheet(_GROUP_STYLE)
        exp_grid = QGridLayout(grp_exposure)
        exp_grid.setSpacing(8)
        exp_grid.setContentsMargins(10, 14, 10, 10)

        # Row 0 — 日均使用量参考报告 | 产品类别 | 使用类型 | 使用部位
        exp_grid.addWidget(QLabel('日均使用量参考报告:'), 0, 0)
        self.data_source_combo = QComboBox()
        self.data_source_combo.setEditable(True)
        self.data_source_combo.addItems([
            '',
            '欧盟SCCS指南',
            '日本JCIA消费者调研数据',
            '荷兰RIVM暴露因子报告',
            '美国PCPC消费者调研',
            '中国T/FDCA 13-2025团标',
            '《化妆品安全评估技术导则》附录',
        ])
        self.data_source_combo.currentTextChanged.connect(self._on_data_source_changed)
        exp_grid.addWidget(self.data_source_combo, 0, 1)

        exp_grid.addWidget(QLabel('产品类别:'), 0, 2)
        self.category_combo = QComboBox()
        self.category_combo.setEditable(True)
        self._load_product_categories()
        exp_grid.addWidget(self.category_combo, 0, 3)

        exp_grid.addWidget(QLabel('使用类型:'), 0, 4)
        self.rinse_type_combo = QComboBox()
        self.rinse_type_combo.addItems(['', '驻留型', '淋洗型'])
        self.rinse_type_combo.currentTextChanged.connect(self._on_rinse_type_changed)
        exp_grid.addWidget(self.rinse_type_combo, 0, 5)

        exp_grid.addWidget(QLabel('使用部位:'), 0, 6)
        self.site_combo = QComboBox()
        self.site_combo.setEditable(True)
        self.site_combo.addItems(_APPLICATION_SITES)
        self.site_combo.currentTextChanged.connect(self._update_exposure_params)
        exp_grid.addWidget(self.site_combo, 0, 7)

        # v2 S4: 推进剂百分比（气雾剂专用，0 表示无推进剂）
        exp_grid.addWidget(QLabel('推进剂占比(%):'), 0, 8)
        self.propellant_spin = QDoubleSpinBox()
        self.propellant_spin.setRange(0, 80)
        self.propellant_spin.setSuffix('%')
        self.propellant_spin.setValue(0)
        self.propellant_spin.setToolTip('气雾剂产品中推进剂占总配方的百分比。\n'
            '其他原料浓度将扣除推进剂后 ×100% 重新计算。\n'
            '推进剂应与其他原料分开评估。')
        exp_grid.addWidget(self.propellant_spin, 0, 9)

        # Row 1 — 使用人群 | 驻留因子 | 日均使用量 | 体重
        exp_grid.addWidget(QLabel('使用人群:'), 1, 0)
        self.population_combo = QComboBox()
        conn = db.get_db()
        age_rows = conn.execute(
            'SELECT age_group, default_weight_kg FROM exposure_body_weight ORDER BY default_weight_kg'
        ).fetchall()
        conn.close()
        self._pop_weight_map = {}
        for ar in age_rows:
            label = ar['age_group']
            self._pop_weight_map[label] = ar['default_weight_kg']
            self.population_combo.addItem(label)
        self.population_combo.currentTextChanged.connect(self._on_population_changed)
        exp_grid.addWidget(self.population_combo, 1, 1)

        exp_grid.addWidget(QLabel('驻留因子:'), 1, 2)
        self.retention_spin = QDoubleSpinBox()
        self.retention_spin.setRange(0, 1)
        self.retention_spin.setDecimals(2)
        self.retention_spin.setSingleStep(0.1)
        self.retention_spin.setValue(1.0)
        self.retention_spin.setFixedWidth(100)
        self.retention_spin.setButtonSymbols(QAbstractSpinBox.NoButtons)
        exp_grid.addWidget(self.retention_spin, 1, 3)

        exp_grid.addWidget(QLabel('日均使用量(g/day):'), 1, 4)
        self.daily_amount_spin = QDoubleSpinBox()
        self.daily_amount_spin.setRange(0, 1000)
        self.daily_amount_spin.setDecimals(4)
        self.daily_amount_spin.setValue(0)
        self.daily_amount_spin.setSuffix(' g/day')
        self.daily_amount_spin.setFixedWidth(130)
        self.daily_amount_spin.setButtonSymbols(QAbstractSpinBox.NoButtons)
        exp_grid.addWidget(self.daily_amount_spin, 1, 5)

        exp_grid.addWidget(QLabel('体重:'), 1, 6)
        self.bw_spin = QDoubleSpinBox()
        self.bw_spin.setRange(1, 200)
        self.bw_spin.setValue(60)
        self.bw_spin.setSuffix(' kg')
        self.bw_spin.setFixedWidth(100)
        self.bw_spin.setReadOnly(True)
        self.bw_spin.setButtonSymbols(QAbstractSpinBox.NoButtons)
        exp_grid.addWidget(self.bw_spin, 1, 7)

        # Column stretches: label cols fixed, widget cols expand evenly
        for col in range(8):
            exp_grid.setColumnStretch(col, 0 if col % 2 == 0 else 1)

        tab1_layout.addWidget(grp_exposure)

        self.class_group = QGroupBox('产品分类判定')
        self.class_group.setStyleSheet(_GROUP_STYLE)
        class_layout = QHBoxLayout(self.class_group)
        class_layout.setSpacing(8)

        # -- 第一类判定 --
        case1_frame = QFrame()
        case1_frame.setStyleSheet(
            "QFrame { background: #fef2f2; border: 1px solid #fecaca; border-radius: 4px; padding: 6px; }"
        )
        case1_inner = QVBoxLayout(case1_frame)
        case1_inner.setContentsMargins(8, 4, 8, 4)
        lbl_case1 = QLabel('第一类判定')
        lbl_case1.setStyleSheet("font-weight: bold; color: #dc2626; font-size: 12px;")
        case1_inner.addWidget(lbl_case1)
        self.is_children_check = QCheckBox('儿童产品')
        self.is_children_check.setToolTip(
            '适用人群含婴幼儿/儿童的化妆品。\n'
            '依据《儿童化妆品监督管理规定》，儿童化妆品是指适用于年龄在12岁以下（含12岁）儿童的化妆品。'
        )
        case1_inner.addWidget(self.is_children_check)
        self.contains_new_ingredient_check = QCheckBox('含监测期内新原料')
        self.contains_new_ingredient_check.setToolTip(
            '配方中含有处于监测期内的新原料。\n'
            '监测期通常为3-4年，期间需收集不良反应报告。\n'
            '新原料完成监测期后方可纳入已使用化妆品原料目录。'
        )
        case1_inner.addWidget(self.contains_new_ingredient_check)
        case1_inner.addStretch()
        class_layout.addWidget(case1_frame, 1)

        # -- 第二类情形一 --
        case2_frame = QFrame()
        case2_frame.setStyleSheet(
            "QFrame { background: #fffbeb; border: 1px solid #fde68a; border-radius: 4px; padding: 6px; }"
        )
        case2_inner = QVBoxLayout(case2_frame)
        case2_inner.setContentsMargins(8, 4, 8, 4)
        lbl_case2 = QLabel('第二类情形一')
        lbl_case2.setStyleSheet("font-weight: bold; color: #b45309; font-size: 12px;")
        case2_inner.addWidget(lbl_case2)
        self.contains_nano_check = QCheckBox('含纳米原料')
        self.contains_nano_check.setToolTip(
            '产品配方中含有纳米原料（粒径1-100nm）。\n'
            '纳米原料需按《化妆品纳米原料管理要求》提供安全性评价资料，\n'
            '包括但不限于：粒径分布、表面性质、生物持久性、毒代动力学数据。'
        )
        case2_inner.addWidget(self.contains_nano_check)
        self.contains_unlisted_sunscreen_check = QCheckBox('非防晒类用未收载防晒剂')
        self.contains_unlisted_sunscreen_check.setToolTip(
            '非防晒类产品中使用未列入《准用防晒剂目录》的防晒剂成分。\n'
            '需提供该防晒剂在相应使用条件下的安全评估资料。'
        )
        case2_inner.addWidget(self.contains_unlisted_sunscreen_check)
        self.uses_device_check = QCheckBox('配合仪器工具使用')
        self.uses_device_check.setToolTip(
            '产品需要配合仪器或工具使用（如美容仪、导入仪等）。\n'
            '需评估仪器与产品的相互作用及安全性。'
        )
        case2_inner.addWidget(self.uses_device_check)
        case2_inner.addStretch()
        class_layout.addWidget(case2_frame, 1)

        # -- 第二类情形二 --
        case3_frame = QFrame()
        case3_frame.setStyleSheet(
            "QFrame { background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 4px; padding: 6px; }"
        )
        case3_inner = QVBoxLayout(case3_frame)
        case3_inner.setContentsMargins(8, 4, 8, 4)
        lbl_case3 = QLabel('第二类情形二（默认）')
        lbl_case3.setStyleSheet("font-weight: bold; color: #16a34a; font-size: 12px;")
        case3_inner.addWidget(lbl_case3)
        lbl_default = QLabel(
            '不符合第一类或第二类情形一条件的普通化妆品，按第二类情形二进行安全评估。'
        )
        lbl_default.setStyleSheet("color: #4b5563; font-size: 11px;")
        case3_inner.addWidget(lbl_default)
        case3_inner.addStretch()
        class_layout.addWidget(case3_frame, 1)

        self.product_class_label = QLabel('')
        self.class_reason_label = QLabel('')
        tab1_layout.addWidget(self.class_group)

        grp_label = QGroupBox('标签与使用说明')
        grp_label.setStyleSheet(_GROUP_STYLE)
        label_layout = QVBoxLayout(grp_label)
        label_layout.setSpacing(6)

        lr1 = QHBoxLayout()
        lr1.addWidget(QLabel('使用方法:'))
        self.usage_method_input = QLineEdit()
        self.usage_method_input.setPlaceholderText('如：洁面后取适量均匀涂抹于面部...')
        lr1.addWidget(self.usage_method_input, 1)
        lr1.addWidget(QLabel('   注意事项:'))
        self.precautions_input = QLineEdit()
        self.precautions_input.setPlaceholderText('如：避免接触眼睛，如不慎入眼请立即用清水冲洗...')
        lr1.addWidget(self.precautions_input, 1)
        label_layout.addLayout(lr1)

        lr2 = QHBoxLayout()
        lr2.addWidget(QLabel('标签警示语:'))
        self.warning_label_input = QLineEdit()
        self.warning_label_input.setPlaceholderText('如：本品含XX成分，敏感肌请先做皮肤测试...')
        lr2.addWidget(self.warning_label_input, 1)
        lr2.addWidget(QLabel('   危害识别表:'))
        self.hazard_format_combo = QComboBox()
        self.hazard_format_combo.addItems(['简化版 (单成分/单一指标)',
                                           '标准版 (多指标交叉引用)',
                                           '完整版 (含文献引用)'])
        lr2.addWidget(self.hazard_format_combo)
        label_layout.addLayout(lr2)

        tab1_layout.addWidget(grp_label)
        tab1_layout.addStretch()
        self.tab_widget.addTab(tab1, '产品信息与分类')

        # ── Tab 3: 评估结果 ──
        tab3 = QWidget()
        tab3_layout = QVBoxLayout(tab3)
        tab3_layout.setSpacing(6)

        summary_frame = QFrame()
        summary_frame.setFrameShape(QFrame.StyledPanel)
        summary_frame.setStyleSheet(
            "QFrame { border: 1px solid #d0d5dd; border-radius: 6px; background: #f9fafb; padding: 6px; }"
        )
        summary_layout = QHBoxLayout(summary_frame)
        summary_layout.setContentsMargins(8, 4, 8, 4)
        self.summary_label = QLabel('请点击"运行安全评估"开始')
        self.summary_label.setStyleSheet("font-size: 13px; font-weight: bold;")
        summary_layout.addWidget(self.summary_label)
        tab3_layout.addWidget(summary_frame)

        self.table = QTableWidget(0, 11)
        self.table.setHorizontalHeaderLabels([
            '原料名称', '标准中文名称', 'INCI名称',
            '添加量(%)', '原料中成分含量(%)', '实际成分含量(%)',
            '禁用', '限用', '使用参考', 'SED', '综合结果'
        ])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(True)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.setStyleSheet("""
            QTableWidget { border: 1px solid #d0d7de; gridline-color: #d0d7de; }
            QTableWidget::item { border: 1px solid #e5e7eb; padding: 4px; }
            QTableWidget::item:selected { background-color: #dbeafe; }
            QHeaderView::section {
                background-color: #f3f4f6; border: 1px solid #d0d7de;
                padding: 4px; font-weight: bold;
            }
        """)
        self.table.cellClicked.connect(self._on_cell_clicked)
        tab3_layout.addWidget(self.table)
        self.tab_widget.addTab(tab3, '评估结果')

        layout.addWidget(self.tab_widget, 1)

        # ═══════════════════════════════════════════
        # ── Action buttons (always visible) ──
        # ═══════════════════════════════════════════
        action_bar = QHBoxLayout()
        self.btn_load = QPushButton('从当前配方加载')
        self.btn_load.clicked.connect(self._on_load_formula)
        action_bar.addWidget(self.btn_load)

        self.btn_import_formula = QPushButton('从Excel导入配方')
        self.btn_import_formula.clicked.connect(self._on_import_formula_excel)
        action_bar.addWidget(self.btn_import_formula)

        self.btn_run = QPushButton('运行安全评估')
        self.btn_run.setStyleSheet(
            "QPushButton { background: #1d4ed8; color: white; font-weight: bold; padding: 6px 16px; }"
            "QPushButton:hover { background: #1e40af; }"
        )
        self.btn_run.clicked.connect(self._on_run_assessment)
        action_bar.addWidget(self.btn_run)

        self.btn_word_report = QPushButton('生成Word报告')
        self.btn_word_report.clicked.connect(self._on_generate_word_report)
        action_bar.addWidget(self.btn_word_report)

        self.btn_preview_3tables = QPushButton('预览 3 大表')
        self.btn_preview_3tables.clicked.connect(self._on_preview_3tables)
        action_bar.addWidget(self.btn_preview_3tables)

        self.btn_save_project = QPushButton('保存项目')
        self.btn_save_project.clicked.connect(self._on_save_project)
        action_bar.addWidget(self.btn_save_project)

        self.btn_load_project = QPushButton('加载项目')
        self.btn_load_project.clicked.connect(self._on_load_project)
        action_bar.addWidget(self.btn_load_project)

        action_bar.addStretch()

        self.ingredient_count_label = QLabel('当前配方: 0 个原料')
        action_bar.addWidget(self.ingredient_count_label)

        layout.addLayout(action_bar)

        # ── Signal connections ──
        self.category_combo.currentTextChanged.connect(self._on_category_changed)
        self.is_children_check.stateChanged.connect(self._on_product_class_changed)
        self.contains_nano_check.stateChanged.connect(self._on_product_class_changed)
        self.contains_new_ingredient_check.stateChanged.connect(self._on_product_class_changed)
        self.contains_unlisted_sunscreen_check.stateChanged.connect(self._on_product_class_changed)
        self.uses_device_check.stateChanged.connect(self._on_product_class_changed)

    # ── Helper Methods ──

    def _determine_product_class(self):
        """根据当前参数判定产品分类"""
        # 获取当前参数
        is_children = self.is_children_check.isChecked()
        contains_nano = self.contains_nano_check.isChecked()
        contains_new_ingredient = self.contains_new_ingredient_check.isChecked()
        contains_unlisted_sunscreen = self.contains_unlisted_sunscreen_check.isChecked()
        uses_device = self.uses_device_check.isChecked()
        category = self.category_combo.currentText().strip()

        # 特殊化妆品功效列表
        special_efficacies = {
            '祛斑美白': '祛斑美白功效',
            '防晒': '防晒功效',
            '染发': '染发功效',
            '烫发': '烫发功效',
            '防脱发': '防脱发功效'
        }

        # 第二类情形一功效列表
        case1_efficacies = {
            '祛痘': '祛痘功效',
            '抗皱': '抗皱功效',
            '除臭': '除臭功效',
            '去屑': '去屑功效',
            '脱毛': '脱毛功效',
            '去角质': '去角质功效'
        }

        # 第二类情形一剂型列表
        case1_formulations = {
            '贴膜': '贴膜剂型',
            '气雾剂': '气雾剂剂型'
        }

        # 第一类判定（按优先级）
        # 1. 特殊化妆品功效（最高优先级）
        for efficacy, desc in special_efficacies.items():
            if efficacy in category:
                return ('第一类', '特殊化妆品', desc)

        # 2. 儿童产品
        if is_children:
            return ('第一类', '儿童化妆品', '适用人群含婴幼儿/儿童')

        # 3. 含监测期内新原料
        if contains_new_ingredient:
            return ('第一类', '含新原料化妆品', '配方中含有监测期内新原料')

        # 第二类情形一判定
        # 1. 含纳米原料
        if contains_nano:
            return ('第二类情形一', '含纳米原料化妆品', '配方中含有纳米原料')

        # 2. 非防晒类使用未收载防晒剂
        if contains_unlisted_sunscreen:
            return ('第二类情形一', '使用未收载防晒剂', '非防晒类产品使用未收载防晒剂')

        # 3. 配合仪器工具使用
        if uses_device:
            return ('第二类情形一', '配合仪器使用', '产品需要配合仪器或工具使用')

        # 4. 特定功效
        for efficacy, desc in case1_efficacies.items():
            if efficacy in category:
                return ('第二类情形一', '特定功效化妆品', desc)

        # 5. 特定剂型
        for formulation, desc in case1_formulations.items():
            if formulation in category:
                return ('第二类情形一', '特定剂型化妆品', desc)

        # 默认：第二类情形二
        return ('第二类情形二', '普通化妆品', '不满足第一类或第二类情形一判定条件')

    def _on_product_class_changed(self):
        """产品分类参数变化时更新显示"""
        product_class, class_name, reason = self._determine_product_class()
        self.product_class_label.setText(f'{product_class} - {class_name}')
        self.class_reason_label.setText(reason)

        # 根据分类设置颜色
        if product_class == '第一类':
            self.product_class_label.setStyleSheet("color: #dc2626; font-weight: bold; font-size: 14px;")
            self.class_group.setStyleSheet("QGroupBox { border: 2px solid #dc2626; border-radius: 6px; }")
        elif product_class == '第二类情形一':
            self.product_class_label.setStyleSheet("color: #ca8a04; font-weight: bold; font-size: 14px;")
            self.class_group.setStyleSheet("QGroupBox { border: 2px solid #ca8a04; border-radius: 6px; }")
        else:
            self.product_class_label.setStyleSheet("color: #16a34a; font-weight: bold; font-size: 14px;")
            self.class_group.setStyleSheet("QGroupBox { border: 2px solid #16a34a; border-radius: 6px; }")

    def _load_product_categories(self):
        """从数据库加载产品类别"""
        # 初始化产品类别数据
        db.init_product_categories()

        # 加载类别
        categories = db.get_all_product_categories()
        self._category_cache = {cat['name_zh']: cat for cat in categories}

        # 填充下拉框
        self.category_combo.clear()
        self.category_combo.addItem('')  # 空选项
        for cat in categories:
            self.category_combo.addItem(cat['name_zh'])

    def _on_category_changed(self, category_name):
        """产品类别选择变更时更新UI"""
        if not category_name:
            self.rinse_type_combo.setCurrentText('')
            return

        cat = self._category_cache.get(category_name)
        if cat:
            # 更新驻留/淋洗下拉框
            if self.rinse_type_combo.currentText().strip() == '':
                self.rinse_type_combo.setCurrentText('淋洗型' if cat['is_rinsed'] else '驻留型')

            # 自动填充使用部位
            if cat['application_site'] and self.site_combo.currentText().strip() == '':
                self.site_combo.setCurrentText(cat['application_site'])

        # 尝试从数据库查询日均使用量和驻留因子
        self._update_exposure_params()

    def _update_exposure_params(self):
        """根据产品类别、使用部位、数据来源查询日均使用量和驻留因子"""
        category = self.category_combo.currentText().strip()
        site = self.site_combo.currentText().strip()
        source = self.data_source_combo.currentText().strip()

        if not category or not site:
            return

        # 策略1: 精确匹配产品类别
        rows = db.query_exposure(
            product_category=category,
            application_site=site
        )

        # 策略2: 如果精确匹配不到，尝试模糊匹配
        if not rows:
            # 从类别名称中提取关键词进行模糊匹配
            keywords = ['霜', '乳', '液', '水', '油', '膏', '粉', '蜜', '胶', '露']
            for kw in keywords:
                if kw in category:
                    rows = db.query_exposure(
                        product_category=kw,
                        application_site=site
                    )
                    if rows:
                        break

        # 策略3: 如果还是没有，只按使用部位查询
        if not rows:
            rows = db.query_exposure(application_site=site)

        if rows:
            # 优先匹配数据来源
            matched = None
            if source:
                matched = next((r for r in rows if source in r.get('source', '')), None)
            if not matched:
                matched = rows[0]

            # 自动填充日均使用量和驻留因子（仅在未手动设置时）
            if self.daily_amount_spin.value() == 0:
                if matched.get('daily_amount_g'):
                    self.daily_amount_spin.setValue(float(matched['daily_amount_g']))
            if self.retention_spin.value() == 1.0:
                if matched.get('retention_factor'):
                    self.retention_spin.setValue(float(matched['retention_factor']))

    def _on_data_source_changed(self, source_name):
        """\u6570\u636e\u6765\u6e90\u9009\u62e9\u53d8\u66f4\u65f6\u81ea\u52a8\u586b\u5145\u76f8\u5173\u53c2\u6570"""
        if not source_name:
            return

        # \u6839\u636e\u6570\u636e\u6765\u6e90\u81ea\u52a8\u586b\u5145\u76f8\u5173\u53c2\u6570
        # \u6b27\u6d32SCCS\u6307\u5357 - \u901a\u5e38\u914d\u5408\u6b27\u6d32\u53d1\u7528\u4ea7\u54c1
        if '欧盟' in source_name or 'SCCS' in source_name:
            # \u5982\u679c\u4ea7\u54c1\u7c7b\u522b\u672a\u9009\u62e9\uff0c\u53ef\u4ee5\u5efa\u8bae\u4e00\u4e9b\u6b27\u6d32\u5e38\u89c1\u4ea7\u54c1\u7c7b\u522b
            if self.category_combo.currentText().strip() == '':
                self.category_combo.setCurrentText('护肤霜')

        # \u65e5\u672cJCIA - \u901a\u5e38\u914d\u5408\u65e5\u672c\u53d1\u7528\u4ea7\u54c1
        elif '日本' in source_name or 'JCIA' in source_name:
            if self.category_combo.currentText().strip() == '':
                self.category_combo.setCurrentText('化妆水')

        # \u83ef\u5170RIVM - \u663e\u8457\u91cf\u6d4b\u8bd5\u6570\u636e\u4e30\u5bcc
        elif '荷兰' in source_name or 'RIVM' in source_name:
            pass

        # \u7f8e\u56fdPCPC - \u591a\u5e94\u7528\u4ea7\u54c1\u6570\u636e
        elif '美国' in source_name or 'PCPC' in source_name:
            if self.category_combo.currentText().strip() == '':
                self.category_combo.setCurrentText('洗发水')

        # \u4e2d\u56fd\u56e2\u6807 - \u9002\u7528\u4e8e\u4e2d\u56fd\u5ba1\u6838
        elif '中国' in source_name or 'T/FDCA' in source_name:
            if self.category_combo.currentText().strip() == '':
                self.category_combo.setCurrentText('护肤乳液')

    # ── Actions ──

    def _on_load_formula(self):
        """\u4ece\u5f53\u524d\u914d\u65b9\u52a0\u8f7d\u539f\u6599\u5217\u8868"""
        formula = db.get_current_formula()
        self.ingredient_count_label.setText(
            f'\u5f53\u524d\u914d\u65b9: {len(formula)} \u4e2a\u539f\u6599'
        )

        # Build expanded rows with multi-component support
        expanded_rows = []
        for idx, item in enumerate(formula):
            # Parse composition
            comps = []
            composition = item.get('composition', '')
            if composition:
                try:
                    comps = json.loads(composition)
                except Exception:
                    pass
            if not comps:
                comps = [{'name': item.get('name_zh', ''), 'percent': 100}]

            name_zh = item.get('name_zh', '')
            name_inci = item.get('name_inci', '') or ''
            conc = item.get('added_percent', 0)

            # Main row (first component)
            expanded_rows.append({
                'row_type': 'main',
                'formula_id': item.get('id'),
                'idx': idx,
                'name_zh': name_zh,
                'std_name': comps[0]['name'],
                'inci_name': comps[0].get('inci', '') or name_inci,
                'concentration': conc,
                'comp_percent': comps[0]['percent'],
                'actual_concentration': round(conc * comps[0]['percent'] / 100, 4),
                'banned': '',
                'restricted': '',
                'usage_reference': '',
                'sed': '',
                'overall': ''
            })

            # Subsequent component rows
            for comp in comps[1:]:
                expanded_rows.append({
                    'row_type': 'comp',
                    'formula_id': item.get('id'),
                    'idx': idx,
                    'name_zh': '',
                    'std_name': comp['name'],
                    'inci_name': comp.get('inci', ''),
                    'concentration': '',
                    'comp_percent': comp['percent'],
                    'actual_concentration': round(conc * comp['percent'] / 100, 4),
                    'banned': '',
                    'restricted': '',
                    'usage_reference': '',
                    'sed': '',
                    'overall': ''
                })

        # Track merge groups
        merge_groups = {}
        group_id = None
        for i, row_data in enumerate(expanded_rows):
            if row_data['row_type'] == 'main':
                group_id = row_data['idx']
                merge_groups[group_id] = {'start': i, 'count': 1}
            else:
                if group_id is not None:
                    merge_groups[group_id]['count'] += 1

        # Fill table
        self.table.setRowCount(len(expanded_rows))
        cols = ['name_zh', 'std_name', 'inci_name', 'concentration',
                'comp_percent', 'actual_concentration', 'banned',
                'restricted', 'usage_reference', 'sed', 'overall']

        for i, row_data in enumerate(expanded_rows):
            for col_idx, col_name in enumerate(cols):
                value = row_data[col_name]
                if col_name in ['concentration', 'comp_percent']:
                    if value != '':
                        item = QTableWidgetItem(f"{value:.2f}")
                    else:
                        item = QTableWidgetItem('')
                elif col_name == 'actual_concentration':
                    item = QTableWidgetItem(f"{value:.4f}")
                else:
                    item = QTableWidgetItem(str(value) if value else '')
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(i, col_idx, item)

        # Apply cell merging for multi-component materials
        # Merge cols: name_zh(0), concentration(3)
        # Note: banned, restricted, usage_reference, sed, overall should NOT be merged
        #       because each component should be evaluated independently
        merge_cols = [0, 3]
        for gid, info in merge_groups.items():
            if info['count'] > 1:
                start = info['start']
                span = info['count']
                for col in merge_cols:
                    self.table.setSpan(start, col, span, 1)

        self._expanded_rows = expanded_rows
        self._merge_groups = merge_groups

    def _on_import_formula_excel(self):
        """从Excel导入配方并填入评估表"""
        path, _ = QFileDialog.getOpenFileName(
            self, '选择配方Excel文件', '', 'Excel文件 (*.xlsx *.xls)'
        )
        if not path:
            return

        try:
            # 局部导入pandas，加快启动速度
            import pandas as pd
            df = pd.read_excel(path)
            if df.empty:
                QMessageBox.warning(self, '导入失败', 'Excel文件为空')
                return

            # Normalize column names (strip whitespace)
            df.columns = [str(c).strip() for c in df.columns]

            col_map = {}
            for col in df.columns:
                if '原料名称' in col or '商品名' in col or '中文名' in col:
                    col_map['name'] = col
                elif 'INCI' in col.upper() or '英文名' in col:
                    col_map['inci'] = col
                elif '添加量' in col or '比例' in col or '含量' in col or '浓度' in col or '%' in col:
                    col_map['percent'] = col
                elif '成分' in col and ('组' in col or '成' in col):
                    col_map['composition'] = col

            if 'name' not in col_map:
                QMessageBox.warning(self, '导入失败',
                                    '未找到原料名称列\n\n'
                                    '请确保Excel包含以下列之一：\n'
                                    '· 原料名称 / 商品名 / 中文名\n'
                                    '· INCI名称 / 英文名（可选）\n'
                                    '· 添加量(%) / 比例 / 含量（可选，默认为0）')
                return
            if 'percent' not in col_map:
                col_map['percent'] = None

            expanded_rows = []
            self.table.setRowCount(0)
            merge_groups = {}

            for idx, row in df.iterrows():
                raw_name = str(row[col_map['name']]).strip()
                if not raw_name or raw_name == 'nan':
                    continue

                # Parse concentration
                conc = 0.0
                if col_map['percent'] is not None:
                    try:
                        val = row[col_map['percent']]
                        if pd.notna(val):
                            conc = float(val)
                    except (ValueError, TypeError):
                        conc = 0.0

                # Parse INCI
                inci_name = ''
                if col_map.get('inci'):
                    val = row[col_map['inci']]
                    if pd.notna(val):
                        inci_name = str(val).strip()
                if not inci_name:
                    inci_name = self.catalog.match_inci(raw_name) or ''

                # Parse composition
                comps = []
                if col_map.get('composition'):
                    comp_str = str(row[col_map['composition']]).strip() if pd.notna(row[col_map['composition']]) else ''
                    if comp_str and comp_str != 'nan':
                        comps, _ = _build_composition_from_natural(raw_name, comp_str, self.catalog)

                if not comps:
                    comps = [{'name': raw_name, 'inci': inci_name, 'percent': 100}]

                # Main row
                expanded_rows.append({
                    'row_type': 'main',
                    'formula_id': None,
                    'idx': idx,
                    'name_zh': raw_name,
                    'std_name': comps[0]['name'],
                    'inci_name': comps[0].get('inci', '') or inci_name,
                    'concentration': conc,
                    'comp_percent': comps[0]['percent'],
                    'actual_concentration': round(conc * comps[0]['percent'] / 100, 4),
                    'banned': '',
                    'restricted': '',
                    'usage_reference': '',
                    'sed': '',
                    'overall': ''
                })
                merge_groups[idx] = {'start': len(expanded_rows) - 1, 'count': 1}

                # Component rows
                for comp in comps[1:]:
                    expanded_rows.append({
                        'row_type': 'comp',
                        'formula_id': None,
                        'idx': idx,
                        'name_zh': '',
                        'std_name': comp['name'],
                        'inci_name': comp.get('inci', ''),
                        'concentration': '',
                        'comp_percent': comp['percent'],
                        'actual_concentration': round(conc * comp['percent'] / 100, 4),
                        'banned': '',
                        'restricted': '',
                        'usage_reference': '',
                        'sed': '',
                        'overall': ''
                    })
                    merge_groups[idx]['count'] += 1

            if not expanded_rows:
                QMessageBox.warning(self, '导入失败', '未能从Excel中解析出有效原料数据')
                return

            self.ingredient_count_label.setText(
                f'从Excel导入: {len([r for r in expanded_rows if r["row_type"] == "main"])} 个原料'
            )

            # Fill table
            self.table.setRowCount(len(expanded_rows))
            cols = ['name_zh', 'std_name', 'inci_name', 'concentration',
                    'comp_percent', 'actual_concentration', 'banned',
                    'restricted', 'usage_reference', 'sed', 'overall']

            for i, row_data in enumerate(expanded_rows):
                for col_idx, col_name in enumerate(cols):
                    value = row_data[col_name]
                    if col_name in ['concentration', 'comp_percent']:
                        if value != '':
                            item = QTableWidgetItem(f"{value:.2f}")
                        else:
                            item = QTableWidgetItem('')
                    elif col_name == 'actual_concentration':
                        item = QTableWidgetItem(f"{value:.4f}")
                    else:
                        item = QTableWidgetItem(str(value) if value else '')
                    item.setTextAlignment(Qt.AlignCenter)
                    self.table.setItem(i, col_idx, item)

            # Cell merging for multi-component materials
            merge_cols = [0, 3]
            for gid, info in merge_groups.items():
                if info['count'] > 1:
                    start = info['start']
                    span = info['count']
                    for col in merge_cols:
                        self.table.setSpan(start, col, span, 1)

            self._expanded_rows = expanded_rows
            self._merge_groups = merge_groups

        except Exception as e:
            QMessageBox.critical(self, '导入失败', f'读取Excel出错: {e}')

    def _on_rinse_type_changed(self, text: str):
        """驻留型/淋洗型切换时自动调整驻留因子"""
        if text == '淋洗型':
            self.retention_spin.setValue(0.01)
        elif text == '驻留型':
            self.retention_spin.setValue(1.0)

    def _on_population_changed(self, label: str):
        """人群选择变化时自动填充体重"""
        weight = self._pop_weight_map.get(label)
        if weight is not None:
            self.bw_spin.setValue(float(weight))

    def _on_run_assessment(self):
        """\u8fd0\u884c\u5b89\u5168\u8bc4\u4f30"""
        formula = db.get_current_formula()
        if not formula:
            QMessageBox.information(self, '\u63d0\u793a', '\u914d\u65b9\u4e3a\u7a7a\uff0c\u8bf7\u5148\u6dfb\u52a0\u539f\u6599')
            return

        product_category = self.category_combo.currentText().strip()
        application_site = self.site_combo.currentText().strip()

        # Gather product-level exposure overrides from input fields
        daily_amount_override = self.daily_amount_spin.value()
        daily_amount_override = daily_amount_override if daily_amount_override > 0 else None
        retention_override = self.retention_spin.value()
        data_source = self.data_source_combo.currentText().strip() or None
        population = self.population_combo.currentText().strip()
        body_weight_override = self.bw_spin.value()
        body_weight_override = body_weight_override if body_weight_override > 0 else None

        self.summary_label.setText('\u6b63\u5728\u8fd0\u884c\u5b89\u5168\u8bc4\u4f30...')
        self.summary_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #2563eb;")
        self.btn_run.setEnabled(False)

        try:
            result = se.assess_formula(formula, product_category, application_site,
                                       daily_amount_override, retention_override, data_source,
                                       population, body_weight_override,
                                       propellant_percent=self.propellant_spin.value())
            self._display_results(result)
        except Exception as e:
            QMessageBox.warning(self, '\u8bc4\u4f30\u5931\u8d25', f'\u5b89\u5168\u8bc4\u4f30\u51fa\u9519: {e}')
            self.summary_label.setText(f'\u8bc4\u4f30\u5931\u8d25: {e}')
        finally:
            self.btn_run.setEnabled(True)

    def _display_results(self, result):
        """\u663e\u793a\u8bc4\u4f30\u7ed3\u679c"""
        self._results = result
        items = result['results']

        # Get original formula data for raw material names
        formula = db.get_current_formula()
        formula_map = {item['id']: item for item in formula}

        # Summary bar
        passed = result['passed']
        total = result['total_ingredients']
        failed = result['failed']
        if result['all_pass']:
            color = '#16a34a'
            status_text = '\u2713 \u5168\u90e8\u901a\u8fc7'
        else:
            color = '#dc2626'
            status_text = f'\u2717 {failed} \u4e2a\u6210\u5206\u6709\u95ee\u9898'

        self.summary_label.setText(
            f'\u5b89\u5168\u8bc4\u4f30\u7ed3\u679c: {passed}/{total} \u901a\u8fc7 {status_text}'
        )
        self.summary_label.setStyleSheet(f"font-size: 13px; font-weight: bold; color: {color};")

        # Build expanded rows with component-level assessment results
        # Now each item is already a component-level assessment from the engine
        expanded_rows = []
        formula_id_to_raw_name = {}

        # First pass: build mapping for raw material names
        for r in items:
            formula_id = r.get('formula_id')
            if formula_id and formula_id not in formula_id_to_raw_name:
                if formula_id in formula_map:
                    formula_id_to_raw_name[formula_id] = formula_map[formula_id].get('name_zh', '')
                else:
                    formula_id_to_raw_name[formula_id] = ''

        # Build rows with component-level assessments
        for idx, r in enumerate(items):
            std_name = r.get('std_name', '') or ''
            inci = r.get('inci_name', '')
            conc = r.get('concentration', 0)
            comp_percent = r.get('component_percent', 100)
            base_conc = r.get('base_concentration', conc)
            raw_material_name = r.get('raw_material_name', '') or formula_id_to_raw_name.get(r.get('formula_id'), '')
            component_idx = r.get('component_index', 0)

            # Extract assessment results
            banned = r.get('banned', {}).get('banned', False)
            restricted = r.get('restricted', {})
            usage = r.get('usage_reference', {})
            exposure = r.get('exposure', {})
            sed = exposure.get('sed')
            overall_pass = r.get('overall_pass', False)

            # Build banned text
            banned_text = '\u2713' if not banned else '\u2717 \u7981\u7528'
            banned_color = Qt.green if not banned else Qt.red

            # Build restricted text
            restricted_flag = restricted.get('restricted', False)
            if restricted_flag:
                conc_ok = restricted.get('concentration_ok', True)
                if not conc_ok:
                    restricted_text = '\u2717 \u8d85\u6807'
                    restricted_color = Qt.red
                else:
                    restricted_text = '\u26a0 \u9650\u7528'
                    restricted_color = QColor('#ca8a04')
            else:
                restricted_text = '\u2713'
                restricted_color = Qt.green

            # Build usage text
            within = usage.get('within_range')
            if within is True:
                usage_text = '\u2713 \u5728\u8303\u56f4\u5185'
                usage_color = Qt.green
            elif within is False:
                usage_text = '\u2717 \u8d85\u51fa\u53c2\u8003'
                usage_color = Qt.red
            else:
                usage_text = '\u2014 \u65e0\u6570\u636e'
                usage_color = QColor('#6b7280')

            # Build overall text
            overall_text = '\u2713 \u901a\u8fc7' if overall_pass else '\u2717 \u4e0d\u901a\u8fc7'
            overall_fg_color = Qt.green if overall_pass else Qt.red
            overall_bg_color = QColor('#dcfce7') if overall_pass else QColor('#fee2e2')

            # Build SED text
            sed_text = f'{sed:.6f}' if sed is not None else '\u2014'

            # Determine row type (main or comp)
            row_type = 'main' if component_idx == 0 else 'comp'

            # For main row, show raw material name; for comp rows, show empty
            display_name = raw_material_name if component_idx == 0 else ''

            expanded_rows.append({
                'row_type': row_type,
                'idx': idx,
                'formula_id': r.get('formula_id'),
                'name': display_name,
                'std_name': std_name,
                'inci': inci,
                'conc': base_conc,
                'comp_percent': comp_percent,
                'actual_conc': conc,
                'banned': {'text': banned_text, 'color': banned_color},
                'restricted': {'text': restricted_text, 'color': restricted_color},
                'usage': {'text': usage_text, 'color': usage_color},
                'sed': sed_text,
                'overall': {'text': overall_text, 'fg_color': overall_fg_color, 'bg_color': overall_bg_color}
            })

        # Track merge groups
        merge_groups = {}
        group_id = None
        for i, row_data in enumerate(expanded_rows):
            if row_data['row_type'] == 'main':
                group_id = row_data['idx']
                merge_groups[group_id] = {'start': i, 'count': 1}
            else:
                if group_id is not None:
                    merge_groups[group_id]['count'] += 1

        # Fill table
        self.table.setRowCount(len(expanded_rows))

        for i, row_data in enumerate(expanded_rows):
            # Column 0: 原料名称
            item = QTableWidgetItem(row_data['name'])
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, 0, item)

            # Column 1: 标准中文名称
            item = QTableWidgetItem(row_data['std_name'])
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, 1, item)

            # Column 2: INCI名称
            item = QTableWidgetItem(row_data['inci'])
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, 2, item)

            # Column 3: 添加量(%)
            if row_data['conc'] != '':
                item = QTableWidgetItem(f"{row_data['conc']:.2f}")
            else:
                item = QTableWidgetItem('')
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, 3, item)

            # Column 4: 原料中成分含量(%)
            item = QTableWidgetItem(f"{row_data['comp_percent']:.2f}")
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, 4, item)

            # Column 5: 实际成分含量(%)
            item = QTableWidgetItem(f"{row_data['actual_conc']:.4f}")
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, 5, item)

            # Column 6: 禁用
            banned = row_data['banned']
            item = QTableWidgetItem(banned['text'])
            item.setTextAlignment(Qt.AlignCenter)
            if banned['color']:
                item.setForeground(banned['color'])
            self.table.setItem(i, 6, item)

            # Column 7: 限用
            restricted = row_data['restricted']
            item = QTableWidgetItem(restricted['text'])
            item.setTextAlignment(Qt.AlignCenter)
            if restricted['color']:
                item.setForeground(restricted['color'])
            self.table.setItem(i, 7, item)

            # Column 8: 使用参考
            usage = row_data['usage']
            item = QTableWidgetItem(usage['text'])
            item.setTextAlignment(Qt.AlignCenter)
            if usage['color']:
                item.setForeground(usage['color'])
            self.table.setItem(i, 8, item)

            # Column 9: SED
            item = QTableWidgetItem(row_data['sed'])
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(i, 9, item)

            # Column 10: 综合结果
            overall = row_data['overall']
            item = QTableWidgetItem(overall['text'])
            item.setTextAlignment(Qt.AlignCenter)
            if overall['fg_color']:
                item.setForeground(overall['fg_color'])
            if overall['bg_color']:
                item.setBackground(overall['bg_color'])
            self.table.setItem(i, 10, item)

        # Apply cell merging for multi-component materials
        # Merge cols: name(0), concentration(3)
        # Note: banned, restricted, usage_reference, sed, overall should NOT be merged
        #       because each component should be evaluated independently
        merge_cols = [0, 3]
        for gid, info in merge_groups.items():
            if info['count'] > 1:
                start = info['start']
                span = info['count']
                for col in merge_cols:
                    self.table.setSpan(start, col, span, 1)

        self._expanded_rows = expanded_rows
        self._merge_groups = merge_groups

    def _on_cell_clicked(self, row, col):
        """\u70b9\u51fb\u67e5\u770b\u8be6\u7ec6\u4fe1\u606f"""
        if not self._results:
            return

        # Find the main row index from expanded rows
        main_idx = None
        if row < len(self._expanded_rows):
            main_idx = self._expanded_rows[row]['idx']

        if main_idx is None or main_idx >= len(self._results['results']):
            return

        r = self._results['results'][main_idx]
        detail_lines = [
            f"\u539f\u6599: {r.get('ingredient', '')} (INCI: {r.get('inci_name', '')})",
            f"\u6dfb\u52a0\u91cf: {r.get('concentration', '')}%",
            f"\u7528\u9014: {r.get('purpose', '')}",
            '',
        ]
        # Banned details
        banned = r.get('banned', {})
        detail_lines.append(f"\u3010\u7981\u7528\u68c0\u67e5\u3011 {banned.get('summary', '')}")
        if banned.get('matches'):
            for m in banned['matches']:
                detail_lines.append(f"  - {m.get('name_zh', '')} ({m.get('name_en', '')})")

        # Restricted details
        restricted = r.get('restricted', {})
        detail_lines.append(f"\u3010\u9650\u7528\u68c0\u67e5\u3011 {restricted.get('summary', '')}")
        for v in restricted.get('violations', []):
            detail_lines.append(f"  ! {v}")

        # Allowed lists
        for label, label_cn in [('preservative', '防腐剂'), ('sunscreen', '防晒剂'),
                                ('colorant', '着色剂'), ('hair_dye', '染发剂')]:
            key = f'allowed_{label}'
            if key in r and r[key].get('in_allowed_list'):
                detail_lines.append(f"【准用{label_cn}】 {r[key].get('summary', '')}")

        # Usage data
        usage = r.get('usage_reference', {})
        detail_lines.append(f"\u3010\u4f7f\u7528\u53c2\u8003\u3011 {usage.get('summary', '')}")
        for rec in usage.get('records', [])[:3]:
            detail_lines.append(
                f"  [{rec.get('source_type', '')}] {rec.get('used_part', '')} "
                f"{rec.get('max_percent', '')}% ({rec.get('method', '')})")

        # SED
        exposure = r.get('exposure', {})
        detail_lines.append(f"\u3010SED\u3011 {exposure.get('summary', '')}")
        if exposure.get('daily_amount_g'):
            detail_lines.append(
                f"  \u65e5\u7528\u91cf: {exposure['daily_amount_g']}g | "
                f"\u9a7b\u7559\u56e0\u5b50: {exposure['retention_factor']} | "
                f"\u4f53\u91cd: {exposure['body_weight_kg']}kg")

        detail_lines.append('')
        detail_lines.append(f"\u3010\u7efc\u5408\u7ed3\u8bba\u3011 {r.get('summary', '')}")

        QMessageBox.information(
            self, f"\u5b89\u5168\u8bc4\u4f30\u8be6\u60c5 - {r.get('ingredient', '')}",
            '\n'.join(detail_lines))

    # ── Word report ──

    def _on_generate_word_report(self):
        """\u751f\u6210Word\u5b89\u5168\u8bc4\u4f30\u62a5\u544a"""
        if not self._results or not isinstance(self._results, dict) or 'results' not in self._results:
            QMessageBox.information(self, '\u63d0\u793a', '\u8bf7\u5148\u8fd0\u884c\u5b89\u5168\u8bc4\u4f30')
            return
        product_name = self.product_name_input.text().strip() or '\u672a\u547d\u540d\u4ea7\u54c1'
        category = self.category_combo.currentText().strip()
        site = self.site_combo.currentText().strip()

        path, _ = QFileDialog.getSaveFileName(
            self, '\u4fdd\u5b58Word\u62a5\u544a', f'{product_name}_\u5b89\u5168\u8bc4\u4f30\u62a5\u544a.docx',
            'Word\u6587\u6863 (*.docx)'
        )
        if not path:
            return

        try:
            product_class, class_name, class_reason = self._determine_product_class()
            doc_bytes = rg.generate_safety_report(
                product_name=product_name,
                product_category=category,
                application_site=site,
                assessment_result=self._results,
                extra_info={
                    'usage_method': self.usage_method_input.text().strip(),
                    'precautions': self.precautions_input.text().strip(),
                    'warning_label': self.warning_label_input.text().strip(),
                    'hazard_format': self.hazard_format_combo.currentText(),
                    'data_source': self.data_source_combo.currentText().strip(),
                    'daily_amount': self.daily_amount_spin.value(),
                    'retention_factor': self.retention_spin.value(),
                    'is_children': self.is_children_check.isChecked(),
                    'contains_nano': self.contains_nano_check.isChecked(),
                    'contains_new_ingredient': self.contains_new_ingredient_check.isChecked(),
                    'contains_unlisted_sunscreen': self.contains_unlisted_sunscreen_check.isChecked(),
                    'uses_device': self.uses_device_check.isChecked(),
                    'is_aerosol': '气雾剂' in category or '喷雾' in category,
                    'is_patch': '贴膜' in category or '面膜' in category or '贴' in category,
                    'is_special_cosmetic': product_class == '第一类',
                    'is_case1_efficacy': '祛痘' in category or '抗皱' in category
                                       or '去屑' in category or '除臭' in category
                                       or '脱毛' in category or '去角质' in category,
                    'product_class': product_class,
                    'product_class_name': class_name,
                },
            )
            with open(path, 'wb') as f:
                f.write(doc_bytes)
            QMessageBox.information(self, '\u6210\u529f',
                                    f'Word\u62a5\u544a\u5df2\u751f\u6210: {path}')
        except Exception as e:
            QMessageBox.warning(self, '\u5931\u8d25', f'\u62a5\u544a\u751f\u6210\u51fa\u9519: {e}')

    def _on_preview_3tables(self):
        """\u9884\u89c8 3 \u5927\u8868 + \u7b2c\u56db\u90e8\u5206"""
        if not self._results or not isinstance(self._results, dict) or 'results' not in self._results:
            QMessageBox.information(self, '\u63d0\u793a', '\u8bf7\u5148\u8fd0\u884c\u5b89\u5168\u8bc4\u4f30')
            return
        formula = db.get_current_formula()
        results = self._results.get('results', [])
        try:
            tables = rg.build_3tables_data(formula, results)
            section4 = rg.build_section4_data(results)
        except Exception as e:
            QMessageBox.warning(self, '\u5931\u8d25', f'\u6784\u5efa\u9884\u89c8\u6570\u636e\u51fa\u9519: {e}')
            return
        dlg = ThreeTablesPreviewDialog(tables, section4, self)
        dlg.exec()

    # ── Project persistence ──

    def _on_save_project(self):
        """\u4fdd\u5b58\u9879\u76ee"""
        product_name = self.product_name_input.text().strip()
        if not product_name:
            product_name = '\u672a\u547d\u540d\u4ea7\u54c1'
        formula = db.get_current_formula()

        extra_info = {
            'usage_method': self.usage_method_input.text().strip(),
            'precautions': self.precautions_input.text().strip(),
            'warning_label': self.warning_label_input.text().strip(),
            'hazard_format': self.hazard_format_combo.currentText(),
            'data_source': self.data_source_combo.currentText().strip(),
            'daily_amount': self.daily_amount_spin.value(),
            'retention_factor': self.retention_spin.value(),
        }
        project = pm.create_project(
            name=product_name,
            product_name=product_name,
            product_category=self.category_combo.currentText().strip(),
            application_site=self.site_combo.currentText().strip(),
            population=self.population_combo.currentText().strip(),
            body_weight_kg=self.bw_spin.value(),
            market='CN',
            formula=formula,
            assessment_result=self._results,
        )
        project['extra_info'] = extra_info
        pm.save_project(project)
        QMessageBox.information(self, '\u6210\u529f', f'\u9879\u76ee\u5df2\u4fdd\u5b58: {product_name}')

    def _on_load_project(self):
        """\u52a0\u8f7d\u9879\u76ee"""
        projects = pm.list_projects()
        if not projects:
            QMessageBox.information(self, '\u63d0\u793a', '\u6ca1\u6709\u4fdd\u5b58\u7684\u9879\u76ee')
            return
        items = [f"{p['name']} ({p.get('updated_at', '')[:10]})" for p in projects]
        item, ok = QInputDialog.getItem(self, '\u9009\u62e9\u9879\u76ee',
                                        '\u8bf7\u9009\u62e9\u8981\u52a0\u8f7d\u7684\u9879\u76ee:', items, False)
        if ok and item:
            idx = items.index(item)
            proj = projects[idx]
            data = pm.load_project(proj['path'])
            if data:
                p_info = data.get('product_info', {})
                self.product_name_input.setText(p_info.get('product_name', ''))
                self.category_combo.setCurrentText(p_info.get('product_category', ''))
                self.site_combo.setCurrentText(p_info.get('application_site', ''))
                pop = p_info.get('population', '')
                if pop:
                    idx = self.population_combo.findText(pop)
                    if idx >= 0:
                        self.population_combo.setCurrentIndex(idx)
                # Weight auto-populates from DB via _on_population_changed signal
                ex = data.get('extra_info', {})
                self.usage_method_input.setText(ex.get('usage_method', ''))
                self.precautions_input.setText(ex.get('precautions', ''))
                self.warning_label_input.setText(ex.get('warning_label', ''))
                hf = ex.get('hazard_format', '')
                idx = self.hazard_format_combo.findText(hf)
                if idx >= 0:
                    self.hazard_format_combo.setCurrentIndex(idx)
                self.data_source_combo.setCurrentText(ex.get('data_source', ''))
                da = ex.get('daily_amount', 0)
                if da:
                    self.daily_amount_spin.setValue(float(da))
                rf = ex.get('retention_factor', 1.0)
                self.retention_spin.setValue(float(rf))
                self._results = data.get('assessment', None)
                if self._results:
                    self._display_results(self._results)
                if hasattr(self, 'tab_formula'):
                    self.tab_formula.refresh()
                QMessageBox.information(self, '\u6210\u529f', f'\u9879\u76ee\u5df2\u52a0\u8f7d: {proj["name"]}')


# ── Physicochem & Microbio Test Tab ──

class PhysicochemTestTab(QWidget):
    """理化微生物检验 (PRD 5.11)"""
    _DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'userdata', 'tests')

    def __init__(self, catalog):
        super().__init__()
        self.catalog = catalog
        self._records = []
        self._data_file = os.path.join(self._DATA_DIR, 'physicochem.json')
        self._build_ui()
        self._load_records()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        form = QHBoxLayout()
        form.addWidget(QLabel('产品:'))
        self.ph_product = QLineEdit()
        self.ph_product.setPlaceholderText('产品名称')
        form.addWidget(self.ph_product, 1)
        form.addWidget(QLabel('类型:'))
        self.ph_type = QComboBox()
        self.ph_type.addItems(['', '理化检验', '微生物检验'])
        form.addWidget(self.ph_type)
        form.addWidget(QLabel('项目:'))
        self.ph_item = QLineEdit()
        self.ph_item.setPlaceholderText('检验项目')
        form.addWidget(self.ph_item, 1)
        form.addWidget(QLabel('结果:'))
        self.ph_result = QComboBox()
        self.ph_result.addItems(['', '合格', '不合格'])
        form.addWidget(self.ph_result)
        self.btn_ph_add = QPushButton('添加')
        self.btn_ph_add.clicked.connect(self._on_add)
        form.addWidget(self.btn_ph_add)
        layout.addLayout(form)

        self.ph_table = QTableWidget(0, 6)
        self.ph_table.setHorizontalHeaderLabels(['产品', '检验类型', '检验项目', '标准值', '实测值', '结果'])
        self.ph_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.ph_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.ph_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.ph_table.verticalHeader().setVisible(False)
        layout.addWidget(self.ph_table)

    def _load_records(self):
        self._records = []
        if os.path.exists(self._data_file):
            try:
                with open(self._data_file, 'r', encoding='utf-8') as f:
                    self._records = json.load(f)
            except Exception:
                pass
        self._refresh()

    def _save_records(self):
        os.makedirs(self._DATA_DIR, exist_ok=True)
        with open(self._data_file, 'w', encoding='utf-8') as f:
            json.dump(self._records, f, ensure_ascii=False, indent=2)
        self._refresh()

    def _refresh(self):
        self.ph_table.setRowCount(len(self._records))
        for i, r in enumerate(self._records):
            self.ph_table.setItem(i, 0, QTableWidgetItem(r.get('product', '')))
            self.ph_table.setItem(i, 1, QTableWidgetItem(r.get('test_type', '')))
            self.ph_table.setItem(i, 2, QTableWidgetItem(r.get('item', '')))
            self.ph_table.setItem(i, 3, QTableWidgetItem(r.get('standard', '')))
            self.ph_table.setItem(i, 4, QTableWidgetItem(r.get('actual', '')))
            self.ph_table.setItem(i, 5, QTableWidgetItem(r.get('result', '')))

    def _on_add(self):
        self._records.append({
            'product': self.ph_product.text().strip(),
            'test_type': self.ph_type.currentText(),
            'item': self.ph_item.text().strip(),
            'standard': '',
            'actual': '',
            'result': self.ph_result.currentText(),
        })
        self._save_records()
        self.ph_product.clear()
        self.ph_item.clear()
        self.ph_type.setCurrentIndex(0)
        self.ph_result.setCurrentIndex(0)


# ── Review Tab ──

class ReviewTab(QWidget):
    """质量安全负责人审核 (PRD 5.12)"""

    def __init__(self, catalog):
        super().__init__()
        self.catalog = catalog
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        form = QHBoxLayout()
        form.addWidget(QLabel('产品名称:'))
        self.rv_product = QLineEdit()
        self.rv_product.setPlaceholderText('输入产品名称')
        form.addWidget(self.rv_product, 1)
        form.addWidget(QLabel('报告编号:'))
        self.rv_report = QLineEdit()
        self.rv_report.setPlaceholderText('报告编号')
        form.addWidget(self.rv_report)
        form.addWidget(QLabel('提交人:'))
        self.rv_submitter = QLineEdit()
        self.rv_submitter.setPlaceholderText('提交人')
        form.addWidget(self.rv_submitter)
        self.btn_rv_submit = QPushButton('提交审核')
        self.btn_rv_submit.clicked.connect(self._on_submit)
        form.addWidget(self.btn_rv_submit)
        layout.addLayout(form)

        # Action buttons
        action_bar = QHBoxLayout()
        self.btn_rv_approve = QPushButton('✓ 通过')
        self.btn_rv_approve.setStyleSheet("background: #16a34a; color: white; font-weight: bold;")
        self.btn_rv_approve.clicked.connect(self._on_approve)
        action_bar.addWidget(self.btn_rv_approve)

        self.btn_rv_reject = QPushButton('✗ 驳回')
        self.btn_rv_reject.setStyleSheet("background: #dc2626; color: white; font-weight: bold;")
        self.btn_rv_reject.clicked.connect(self._on_reject)
        action_bar.addWidget(self.btn_rv_reject)

        self.btn_rv_refresh = QPushButton('刷新')
        self.btn_rv_refresh.clicked.connect(self._refresh)
        action_bar.addWidget(self.btn_rv_refresh)
        action_bar.addStretch()
        layout.addLayout(action_bar)

        # Review comment input
        comment_layout = QHBoxLayout()
        comment_layout.addWidget(QLabel('审核意见:'))
        self.rv_comment = QLineEdit()
        comment_layout.addWidget(self.rv_comment, 1)
        layout.addLayout(comment_layout)

        # Table
        self.rv_table = QTableWidget(0, 6)
        self.rv_table.setHorizontalHeaderLabels(['产品名称', '报告编号', '状态', '提交人', '审核意见', '更新时间'])
        self.rv_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.rv_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        self.rv_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.rv_table.verticalHeader().setVisible(False)
        layout.addWidget(self.rv_table)

        self._refresh()

    def _refresh(self):
        reviews = rw.list_reviews()
        self.rv_table.setRowCount(len(reviews))
        for i, r in enumerate(reviews):
            self.rv_table.setItem(i, 0, QTableWidgetItem(r.get('product_name', '')))
            self.rv_table.setItem(i, 1, QTableWidgetItem(r.get('report_number', '')))
            status = r.get('status', '')
            status_map = {'pending': '待审核', 'in_review': '审核中', 'approved': '已通过', 'rejected': '已驳回', 'revision_needed': '需修改'}
            self.rv_table.setItem(i, 2, QTableWidgetItem(status_map.get(status, status)))
            self.rv_table.setItem(i, 3, QTableWidgetItem(r.get('submitted_by', '')))
            history = r.get('history', [])
            last_comment = history[-1].get('comment', '') if history else ''
            self.rv_table.setItem(i, 4, QTableWidgetItem(last_comment))
            self.rv_table.setItem(i, 5, QTableWidgetItem(r.get('updated_at', '')[:16] if r.get('updated_at') else ''))

    def _on_submit(self):
        product = self.rv_product.text().strip()
        if not product:
            QMessageBox.warning(self, '提示', '请输入产品名称')
            return
        report = self.rv_report.text().strip() or f'REP-{datetime.now().strftime("%Y%m%d")}-001'
        submitter = self.rv_submitter.text().strip() or '系统用户'
        record = rw.create_review(product, report, submitter)
        rw.submit_review(record)
        rw.save_review(record)
        QMessageBox.information(self, '成功', f'产品 "{product}" 已提交审核')
        self.rv_product.clear()
        self.rv_report.clear()
        self.rv_submitter.clear()
        self._refresh()

    def _on_approve(self):
        sel = self.rv_table.currentRow()
        if sel < 0:
            QMessageBox.warning(self, '提示', '请先在表格中选择一个审核记录')
            return
        reviews = rw.list_reviews()
        if sel < len(reviews):
            review_id = reviews[sel].get('review_id', '')
            record = rw.load_review(review_id)
            if not record:
                QMessageBox.warning(self, '错误', '未找到审核记录')
                return
            comment = self.rv_comment.text().strip() or '审核通过'
            rw.approve_review(record, '质量安全负责人', comment)
            rw.save_review(record)
            QMessageBox.information(self, '成功', '审核已通过')
            self.rv_comment.clear()
            self._refresh()

    def _on_reject(self):
        sel = self.rv_table.currentRow()
        if sel < 0:
            QMessageBox.warning(self, '提示', '请先在表格中选择一个审核记录')
            return
        reviews = rw.list_reviews()
        if sel < len(reviews):
            review_id = reviews[sel].get('review_id', '')
            record = rw.load_review(review_id)
            if not record:
                QMessageBox.warning(self, '错误', '未找到审核记录')
                return
            comment = self.rv_comment.text().strip() or '审核不通过，请修改后重新提交'
            rw.reject_review(record, '质量安全负责人', comment)
            rw.save_review(record)
            QMessageBox.information(self, '成功', '审核已驳回')
            self.rv_comment.clear()
            self._refresh()


# ── Main Window ──

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("备案配方表助手")
        self.setMinimumSize(1100, 700)
        self.resize(1200, 750)
        self._apply_style()

        # Init catalog
        self.catalog = CatalogManager()

        # Central tab widget
        self.tabs = QTabWidget()

        self.tab_formula = FormulaDesignTab(self.catalog)
        self.tab_materials = MaterialManagementTab(self.catalog)
        self.tab_crawl = CrawlProgressTab(self.catalog)
        self.tab_safety = SafetyAssessmentTab(self.catalog)
        self.tab_physicochem = PhysicochemTestTab(self.catalog)
        self.tab_review = ReviewTab(self.catalog)

        self.tabs.addTab(self.tab_formula, "配方设计")
        self.tabs.addTab(self.tab_materials, "原料管理")
        self.tabs.addTab(self.tab_crawl, "同步目录")
        self.tabs.addTab(self.tab_safety, "安全评估")
        self.tabs.addTab(self.tab_physicochem, "理化微生物")
        self.tabs.addTab(self.tab_review, "审核流程")

        self.setCentralWidget(self.tabs)

        # Status bar
        info = self.catalog.get_catalog_info()
        from catalog_mgr import USERDATA_DIR as _ud
        _cas_path = os.path.join(_ud, "cosing_data.json")
        _cas_count = 0
        if os.path.exists(_cas_path):
            try:
                _cas_count = len(json.load(open(_cas_path, 'r', encoding='utf-8')))
            except Exception:
                pass
        self.statusBar().showMessage(
            f"目录I: {info['catalog_i']}条 | "
            f"目录II: {info['catalog_ii']}条 | "
            f"CAS号: {_cas_count}条 | "
            f"数据库原料: {len(db.get_all_raw_materials())}条"
        )

        self.tabs.currentChanged.connect(self._on_tab_changed)

    def _apply_style(self):
        self.setStyleSheet("""
            /* ── Global ── */
            QMainWindow, QWidget {
                font-family: "Microsoft YaHei", "Segoe UI", sans-serif;
                font-size: 12px;
            }
            QTabWidget::pane {
                border: 1px solid #d0d5dd;
                border-top: none;
                background: #f9fafb;
            }
            QTabBar::tab {
                background: #e4e7eb;
                color: #475467;
                padding: 8px 20px;
                margin-right: 2px;
                border: 1px solid #d0d5dd;
                border-bottom: none;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-size: 13px;
            }
            QTabBar::tab:selected {
                background: #f9fafb;
                color: #1d2939;
                font-weight: bold;
                border-bottom: 1px solid #f9fafb;
            }
            QTabBar::tab:hover:!selected {
                background: #eaecf0;
            }

            /* ── Buttons ── */
            QPushButton {
                background: #ffffff;
                border: 1px solid #d0d5dd;
                border-radius: 6px;
                padding: 6px 14px;
                color: #344054;
                font-size: 12px;
            }
            QPushButton:hover {
                background: #f2f4f7;
                border-color: #98a2b3;
            }
            QPushButton:pressed {
                background: #e4e7eb;
            }
            QPushButton:disabled {
                background: #f2f4f7;
                color: #98a2b3;
            }

            /* ── Table ── */
            QTableWidget {
                background: #ffffff;
                border: 1px solid #eaecf0;
                border-radius: 6px;
                gridline-color: #f2f4f7;
                selection-background-color: #eef4ff;
                selection-color: #1d2939;
                alternate-background-color: #fcfcfd;
            }
            QHeaderView::section {
                background: #f9fafb;
                color: #475467;
                font-weight: 600;
                padding: 8px 6px;
                border: none;
                border-bottom: 2px solid #eaecf0;
                border-right: 1px solid #eaecf0;
                font-size: 12px;
            }
            QHeaderView::section:hover {
                background: #f2f4f7;
            }

            /* ── Inputs ── */
            QLineEdit, QDoubleSpinBox, QSpinBox {
                background: #ffffff;
                border: 1px solid #d0d5dd;
                border-radius: 6px;
                padding: 6px 10px;
                color: #1d2939;
                font-size: 12px;
            }
            QLineEdit:focus, QDoubleSpinBox:focus, QSpinBox:focus {
                border-color: #667eea;
                outline: none;
            }
            QLineEdit:hover, QDoubleSpinBox:hover, QSpinBox:hover {
                border-color: #98a2b3;
            }

            /* ── ComboBox (minimal — avoid full style to keep native dropdown arrow) ── */
            QComboBox {
                min-height: 18px;
            }
            QComboBox QAbstractItemView {
                background: #ffffff;
                border: 1px solid #d0d5dd;
                selection-background-color: #eef4ff;
                selection-color: #1d2939;
            }

            /* ── Status Bar ── */
            QStatusBar {
                background: #f9fafb;
                border-top: 1px solid #eaecf0;
                color: #6b7280;
                font-size: 11px;
                padding: 2px 12px;
            }

            /* ── Progress Bar ── */
            QProgressBar {
                background: #eaecf0;
                border: none;
                border-radius: 4px;
                height: 8px;
                text-align: center;
                font-size: 10px;
                color: #344054;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #667eea, stop:1 #764ba2);
                border-radius: 4px;
            }

            /* ── PlainTextEdit (log) ── */
            QPlainTextEdit {
                background: #ffffff;
                border: 1px solid #eaecf0;
                border-radius: 6px;
                padding: 6px;
                color: #344054;
                font-family: "Consolas", "Courier New", monospace;
                font-size: 11px;
            }

            /* ── ScrollBar ── */
            QScrollBar:vertical {
                background: #f2f4f7;
                width: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: #c4c8ce;
                border-radius: 4px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background: #98a2b3;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0;
            }
            QScrollBar:horizontal {
                background: #f2f4f7;
                height: 8px;
                border-radius: 4px;
            }
            QScrollBar::handle:horizontal {
                background: #c4c8ce;
                border-radius: 4px;
                min-width: 30px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #98a2b3;
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0;
            }
        """)

    def _on_tab_changed(self, idx):
        if idx == 0:
            self.tab_formula.refresh()
        elif idx == 1:
            self.tab_materials.refresh()
        elif idx == 2:
            self.tab_crawl._update_catalog_info()
        elif idx == 3:
            self.tab_safety._on_load_formula()
        elif idx == 4:
            self.tab_physicochem._load_records()
        elif idx == 5:
            self.tab_review._refresh()
        info = self.catalog.get_catalog_info()
        from catalog_mgr import USERDATA_DIR as _ud
        _cas_path = os.path.join(_ud, "cosing_data.json")
        _cas_count = 0
        if os.path.exists(_cas_path):
            try:
                _cas_count = len(json.load(open(_cas_path, 'r', encoding='utf-8')))
            except Exception:
                pass
        self.statusBar().showMessage(
            f"目录I: {info['catalog_i']}条 | "
            f"目录II: {info['catalog_ii']}条 | "
            f"CAS号: {_cas_count}条 | "
            f"数据库原料: {len(db.get_all_raw_materials())}条"
        )
