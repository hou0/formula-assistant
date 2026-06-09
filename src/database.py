import os
import re
import sqlite3
import sys


try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def get_userdata_dir():
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
        return os.path.join(base, 'userdata')
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, 'userdata')


def get_data_dir():
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
        return os.path.join(base, 'data')
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(os.path.normpath(os.path.join(base, '..')), 'data')


DB_PATH = os.path.join(get_userdata_dir(), 'safety.db')
EXCEL_PATH = os.path.join(get_data_dir(), '安全评估数据库_new.xlsx')

_excel_wb = None


def _load_excel():
    global _excel_wb
    if _excel_wb is None and HAS_OPENPYXL and os.path.exists(EXCEL_PATH):
        _excel_wb = openpyxl.load_workbook(EXCEL_PATH,
                                            data_only=True, read_only=True)
    return _excel_wb


def _read_excel_sheet(sheet_name):
    wb = _load_excel()
    if wb is None or sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS raw_materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name_zh TEXT NOT NULL,
            name_inci TEXT,
            composition TEXT NOT NULL,
            default_purpose TEXT,
            supplier_code TEXT,
            remarks TEXT,
            trade_name TEXT,
            internal_code TEXT,
            supplier_name TEXT,
            contact_person TEXT,
            contact_phone TEXT,
            contact_email TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS formula (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            raw_material_id INTEGER NOT NULL,
            added_percent REAL NOT NULL,
            sort_order INTEGER NOT NULL,
            is_new_material INTEGER DEFAULT 0,
            registration_number TEXT DEFAULT '',
            purpose_override TEXT DEFAULT '',
            remarks TEXT DEFAULT ''
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS custom_columns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            width INTEGER DEFAULT 120,
            sort_order INTEGER DEFAULT 0
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS regulation_banned_chemical (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            seq TEXT,
            name_zh TEXT NOT NULL,
            name_en TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS regulation_banned_botanical (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            seq TEXT,
            name_zh TEXT NOT NULL,
            latin_name TEXT,
            notes TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS regulation_restricted (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            seq TEXT,
            name_zh TEXT NOT NULL,
            name_en TEXT,
            inci_name TEXT,
            scope_of_use TEXT,
            max_concentration TEXT,
            restrictions TEXT,
            label_requirements TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS regulation_allowed_preservative (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            seq TEXT,
            name_zh TEXT NOT NULL,
            name_en TEXT,
            inci_name TEXT,
            max_concentration TEXT,
            scope_and_conditions TEXT,
            label_requirements TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS regulation_allowed_sunscreen (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            seq TEXT,
            name_zh TEXT NOT NULL,
            name_en TEXT,
            inci_name TEXT,
            max_concentration TEXT,
            restrictions TEXT,
            label_requirements TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS regulation_allowed_colorant (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            seq TEXT,
            name_zh TEXT,
            color_index TEXT,
            ci_generic_name TEXT,
            color TEXT,
            scope_1 INTEGER DEFAULT 0,
            scope_2 INTEGER DEFAULT 0,
            scope_3 INTEGER DEFAULT 0,
            scope_4 INTEGER DEFAULT 0,
            restrictions TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS regulation_allowed_hair_dye (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            seq TEXT,
            name_zh TEXT NOT NULL,
            inci_name TEXT,
            max_conc_oxidative TEXT,
            max_conc_non_oxidative TEXT,
            restrictions TEXT,
            label_requirements TEXT
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS product_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE,
            name_zh TEXT NOT NULL,
            category_group TEXT NOT NULL,
            product_type TEXT NOT NULL DEFAULT '普通',
            is_rinsed BOOLEAN NOT NULL DEFAULT 0,
            is_professional BOOLEAN NOT NULL DEFAULT 0,
            application_site TEXT NOT NULL DEFAULT '',
            description TEXT,
            sort_order INTEGER NOT NULL DEFAULT 0,
            is_active BOOLEAN NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()


def get_all_raw_materials():
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM raw_materials ORDER BY name_zh'
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_raw_material_by_id(rid):
    conn = get_db()
    row = conn.execute(
        'SELECT * FROM raw_materials WHERE id = ?', (rid,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def add_raw_material(
    name_zh, name_inci='', composition_json='', default_purpose='',
    supplier_code='', remarks='', trade_name='', internal_code='',
    supplier_name='', cas_number=''
):
    conn = get_db()
    conn.execute(
        'INSERT INTO raw_materials '
        '(name_zh, name_inci, composition, default_purpose, supplier_code, '
        'remarks, trade_name, internal_code, supplier_name, cas_number) '
        'VALUES (?,?,?,?,?,?,?,?,?,?)',
        (name_zh, name_inci, composition_json, default_purpose,
         supplier_code, remarks, trade_name, internal_code,
         supplier_name, cas_number)
    )
    conn.commit()
    conn.close()


def update_raw_material(
    rid, name_zh, name_inci='', composition_json='', default_purpose='',
    supplier_code='', remarks='', trade_name='', internal_code='',
    supplier_name='', cas_number=''
):
    conn = get_db()
    conn.execute(
        'UPDATE raw_materials SET '
        'name_zh=?, name_inci=?, composition=?, default_purpose=?, '
        'supplier_code=?, remarks=?, trade_name=?, internal_code=?, '
        'supplier_name=?, cas_number=? WHERE id=?',
        (name_zh, name_inci, composition_json, default_purpose,
         supplier_code, remarks, trade_name, internal_code,
         supplier_name, cas_number, rid)
    )
    conn.commit()
    conn.close()


def delete_raw_material(rid):
    conn = get_db()
    conn.execute('DELETE FROM raw_materials WHERE id = ?', (rid,))
    conn.commit()
    conn.close()


def update_raw_material_inci(rid, name_inci):
    conn = get_db()
    conn.execute(
        'UPDATE raw_materials SET name_inci = ? WHERE id = ?',
        (name_inci, rid)
    )
    conn.commit()
    conn.close()


def update_raw_material_composition(rid, composition_json):
    conn = get_db()
    conn.execute(
        'UPDATE raw_materials SET composition = ? WHERE id = ?',
        (composition_json, rid)
    )
    conn.commit()
    conn.close()


def update_raw_material_cas(rid, cas_number):
    conn = get_db()
    conn.execute(
        'UPDATE raw_materials SET cas_number = ? WHERE id = ?',
        (cas_number, rid)
    )
    conn.commit()
    conn.close()


def get_current_formula():
    conn = get_db()
    rows = conn.execute('''
        SELECT f.id, f.raw_material_id, f.added_percent, f.sort_order,
               r.name_zh, r.name_inci, r.composition, r.default_purpose,
               r.supplier_code, r.remarks, r.trade_name, r.internal_code,
               f.is_new_material, f.registration_number, f.purpose_override
        FROM formula f
        JOIN raw_materials r ON f.raw_material_id = r.id
        ORDER BY f.sort_order
    ''').fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_to_formula(raw_material_id, added_percent, is_new_material=0):
    conn = get_db()
    cur = conn.execute('SELECT COALESCE(MAX(sort_order), 0) FROM formula')
    max_order = cur.fetchone()[0]
    new_order = max_order + 1
    conn.execute(
        'INSERT INTO formula '
        '(raw_material_id, added_percent, sort_order, '
        'is_new_material, registration_number, purpose_override) '
        'VALUES (?,?,?,?,?,?)',
        (raw_material_id, added_percent, new_order,
         1 if is_new_material else 0, '', '')
    )
    conn.commit()
    conn.close()


def remove_from_formula(formula_id):
    conn = get_db()
    conn.execute('DELETE FROM formula WHERE id = ?', (formula_id,))
    conn.commit()
    conn.close()


def clear_formula():
    conn = get_db()
    conn.execute('DELETE FROM formula')
    conn.commit()
    conn.close()


def update_formula_sort_order(formula_ids_in_order):
    conn = get_db()
    for idx, fid in enumerate(formula_ids_in_order):
        conn.execute(
            'UPDATE formula SET sort_order = ? WHERE id = ?',
            (idx + 1, fid)
        )
    conn.commit()
    conn.close()


def update_formula_percent(formula_id, new_percent):
    conn = get_db()
    conn.execute(
        'UPDATE formula SET added_percent = ? WHERE id = ?',
        (new_percent, formula_id)
    )
    conn.commit()
    conn.close()


def update_formula_material_info(
        formula_id, is_new_material, registration_number, purpose_override):
    conn = get_db()
    conn.execute(
        'UPDATE formula SET is_new_material = ?, '
        'registration_number = ?, purpose_override = ? WHERE id = ?',
        (1 if is_new_material else 0,
         registration_number, purpose_override, formula_id)
    )
    conn.commit()
    conn.close()


def update_formula_remarks(formula_id, remarks):
    conn = get_db()
    conn.execute(
        'UPDATE formula SET remarks = ? WHERE id = ?',
        (remarks, formula_id)
    )
    conn.commit()
    conn.close()


def _match_name(table, name_field, value):
    if not value or not value.strip():
        return []
    conn = get_db()
    v = value.strip()
    rows = conn.execute(
        f'SELECT * FROM {table} WHERE {name_field} LIKE ?',
        (f'%{v}%',)
    ).fetchall()
    if not rows and len(v) >= 2:
        rows = conn.execute(
            f'SELECT * FROM {table} WHERE {name_field} LIKE ?',
            (f'{v[:4]}%',)
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _extract_cas(text):
    if not text:
        return []
    pattern = r'CAS\s*(?:No\.)?\s*(\d{2,7}-\d{2}-\d)'
    matches = re.findall(pattern, text, re.IGNORECASE)
    return matches


def check_banned_chemical(name_zh, name_inci='', cas_no=''):
    conn = get_db()
    results = []
    search_name = name_zh.strip() if name_zh else ''
    search_inci = name_inci.strip() if name_inci else ''
    search_cas = cas_no.strip() if cas_no else ''

    if search_cas:
        rows = conn.execute(
            'SELECT * FROM regulation_banned_chemical WHERE name_en LIKE ?',
            (f'%{search_cas}%',)
        ).fetchall()
        for r in rows:
            cas_numbers = _extract_cas(r['name_en'])
            if search_cas in cas_numbers:
                results.append(dict(r))

    if not results and search_inci:
        rows = conn.execute(
            'SELECT * FROM regulation_banned_chemical WHERE name_en = ?',
            (search_inci,)
        ).fetchall()
        results = [dict(r) for r in rows]

    if not results and search_inci:
        rows = conn.execute(
            'SELECT * FROM regulation_banned_chemical WHERE name_en LIKE ?',
            (f'%{search_inci}%',)
        ).fetchall()

        filtered_results = []
        for r in rows:
            name_en_full = r['name_en']
            if name_en_full == search_inci:
                filtered_results.append(r)
                continue

            pattern = r'(?:^|[^a-zA-Z])(%s)(?:$|[^a-zA-Z])'
            pattern = pattern % re.escape(search_inci)
            if re.search(pattern, name_en_full, re.IGNORECASE):
                compound_patterns = [
                    r'esters?\s+of\s+' + re.escape(search_inci),
                    r'salts?\s+of\s+' + re.escape(search_inci),
                    r'derivatives?\s+of\s+' + re.escape(search_inci),
                    re.escape(search_inci) + r'\s+esters?',
                    re.escape(search_inci) + r'\s+salts?',
                ]
                is_compound = any(
                    re.search(p, name_en_full, re.IGNORECASE)
                    for p in compound_patterns
                )
                if not is_compound:
                    filtered_results.append(r)

        results = filtered_results

    if not results and search_name:
        rows = conn.execute(
            'SELECT * FROM regulation_banned_chemical WHERE name_zh = ?',
            (search_name,)
        ).fetchall()
        results = [dict(r) for r in rows]

    if not results and search_name:
        rows = conn.execute(
            'SELECT * FROM regulation_banned_chemical WHERE name_zh LIKE ?',
            (f'%{search_name}%',)
        ).fetchall()

        filtered_results = []
        for r in rows:
            name_zh_full = r['name_zh']
            if name_zh_full == search_name:
                filtered_results.append(r)
            else:
                idx = name_zh_full.find(search_name)
                if idx != -1:
                    end_idx = idx + len(search_name)
                    skip = False

                    if end_idx < len(name_zh_full):
                        next_char = name_zh_full[end_idx]
                        compound_chars = '酯盐酰胺酸基化合聚共醚醇酮醛'

                        if next_char in compound_chars:
                            skip = True
                        elif next_char == '等':
                            skip = True
                        elif next_char == '类':
                            skip = True
                        elif next_char == '盐':
                            skip = True

                    if not skip:
                        filtered_results.append(r)

        results = filtered_results

    conn.close()
    results = [dict(r) for r in results]
    for r in results:
        r['table_type'] = 'chemical'
    return results


def check_banned_botanical(name_zh, latin_name='', cas_no=''):
    conn = get_db()
    results = []
    search_name = name_zh.strip() if name_zh else ''
    search_latin = latin_name.strip() if latin_name else ''
    search_cas = cas_no.strip() if cas_no else ''

    if search_cas:
        rows = conn.execute(
            'SELECT * FROM regulation_banned_botanical '
            'WHERE latin_name LIKE ?',
            (f'%{search_cas}%',)
        ).fetchall()
        for r in rows:
            cas_numbers = _extract_cas(r['latin_name'])
            if search_cas in cas_numbers:
                results.append(dict(r))

    if not results and search_latin:
        rows = conn.execute(
            'SELECT * FROM regulation_banned_botanical WHERE latin_name = ?',
            (search_latin,)
        ).fetchall()
        results = [dict(r) for r in rows]

    if not results and search_latin:
        sql = 'SELECT * FROM regulation_banned_botanical '
        sql += 'WHERE latin_name LIKE ?'
        rows = conn.execute(
            sql,
            (f'%{search_latin}%',)
        ).fetchall()
        results = [dict(r) for r in rows]

    if not results and search_name:
        rows = conn.execute(
            'SELECT * FROM regulation_banned_botanical WHERE name_zh LIKE ?',
            (f'%{search_name}%',)
        ).fetchall()
        results = [dict(r) for r in rows]

    conn.close()
    results = [dict(r) for r in results]
    for r in results:
        r['table_type'] = 'botanical'
    return results


def check_banned(name_zh, name_inci='', cas_no=''):
    results = check_banned_chemical(name_zh, name_inci, cas_no)
    results += check_banned_botanical(name_zh, name_inci, cas_no)
    return results


def check_restricted(name_zh, name_inci=''):
    results = _match_name('regulation_restricted', 'name_zh', name_zh)
    if name_inci:
        conn = get_db()
        rows = conn.execute(
            'SELECT * FROM regulation_restricted WHERE inci_name LIKE ?',
            (f'%{name_inci.strip()}%',)
        ).fetchall()
        conn.close()
        for r in rows:
            d = dict(r)
            if d not in results:
                results.append(d)
    return results


def check_allowed_preservative(name_zh, name_inci=''):
    results = _match_name(
        'regulation_allowed_preservative', 'name_zh', name_zh
    )
    if name_inci:
        conn = get_db()
        rows = conn.execute(
            'SELECT * FROM regulation_allowed_preservative '
            'WHERE inci_name LIKE ?',
            (f'%{name_inci.strip()}%',)
        ).fetchall()
        conn.close()
        for r in rows:
            d = dict(r)
            if d not in results:
                results.append(d)
    return results


def check_allowed_sunscreen(name_zh, name_inci=''):
    results = _match_name(
        'regulation_allowed_sunscreen', 'name_zh', name_zh
    )
    if name_inci:
        conn = get_db()
        rows = conn.execute(
            'SELECT * FROM regulation_allowed_sunscreen '
            'WHERE inci_name LIKE ?',
            (f'%{name_inci.strip()}%',)
        ).fetchall()
        conn.close()
        for r in rows:
            d = dict(r)
            if d not in results:
                results.append(d)
    return results


def check_allowed_colorant(name_zh=''):
    return _match_name('regulation_allowed_colorant', 'name_zh', name_zh)


def check_allowed_hair_dye(name_zh, name_inci=''):
    results = []

    if name_inci:
        search_inci = name_inci.strip()
        conn = get_db()

        rows = conn.execute(
            'SELECT * FROM regulation_allowed_hair_dye WHERE inci_name = ?',
            (search_inci,)
        ).fetchall()
        results = [dict(r) for r in rows]

        if not results:
            sql = 'SELECT * FROM regulation_allowed_hair_dye '
            sql += 'WHERE inci_name LIKE ?'
            rows = conn.execute(
                sql,
                (f'%{search_inci}%',)
            ).fetchall()
            results = [dict(r) for r in rows]

        conn.close()

    if not results and name_zh:
        results = _match_name(
            'regulation_allowed_hair_dye', 'name_zh', name_zh
        )

    return results


def query_usage_data(name_zh='', inci_name=''):
    """Query usage data from unified Excel (fallback to DB)."""
    # Try Excel first
    if HAS_OPENPYXL:
        try:
            results = []
            for sheet_name, source_label in [
                ('已上市使用量', '已上市'),
                ('国际索引', '国际')
            ]:
                rows = _read_excel_sheet(sheet_name)
                for row in rows:
                    # row: 序号, 目录序号, 原料名称, INCI名称, 作用部位,
                    #      使用方法, 最大使用量(%), 备注
                    name_zh_val = row[2] or ''
                    inci_name_val = row[3] or ''
                    match = True
                    if name_zh:
                        nz = name_zh.strip()
                        if nz not in name_zh_val:
                            match = False
                    if inci_name:
                        iz = inci_name.strip()
                        if iz not in inci_name_val:
                            match = False
                    if match:
                        results.append({
                            'catalog_seq': row[1],
                            'name_zh': row[2],
                            'inci_name': row[3],
                            'used_part': row[4],
                            'method': row[5],
                            'max_percent': row[6],
                            'remarks': row[7],
                            'source_type': source_label
                        })
            return results
        except Exception:
            pass
    # Fallback to DB
    results = []
    conn = get_db()
    for table, source_label in [
        ('safety_usage_market', '已上市'),
        ('safety_usage_international', '国际')
    ]:
        clauses = []
        params = []
        if name_zh:
            clauses.append('name_zh LIKE ?')
            params.append(f'%{name_zh.strip()}%')
        if inci_name:
            clauses.append('inci_name LIKE ?')
            params.append(f'%{inci_name.strip()}%')
        if not clauses:
            continue
        sql = 'SELECT *, ? AS source_type FROM {table}'
        sql = sql.format(table=table)
        sql += ' WHERE ' + ' OR '.join(clauses)
        rows = conn.execute(sql, [source_label] + params).fetchall()
        results.extend([dict(r) for r in rows])
    conn.close()
    return results


def query_exposure(product_category='', application_site=''):
    """Query exposure data from unified Excel (fallback to DB)."""
    if HAS_OPENPYXL:
        try:
            rows = _read_excel_sheet('暴露量参数')
            results = []
            for row in rows:
                # row: 产品类别, 使用部位, 数据来源, 日均使用量(g), 驻留因子, 备注
                match = True
                if product_category:
                    if not row[0] or product_category.strip() not in row[0]:
                        match = False
                if application_site:
                    if not row[1] or application_site.strip() not in str(row[1]):
                        match = False
                if match:
                    results.append({
                        'product_category': row[0],
                        'application_site': row[1],
                        'source': row[2],
                        'daily_amount_g': row[3],
                        'retention_factor': row[4],
                        'notes': row[5]
                    })
            return results
        except Exception:
            pass
    # Fallback to DB
    conn = get_db()
    clauses = []
    params = []
    if product_category:
        clauses.append('product_category LIKE ?')
        params.append(f'%{product_category.strip()}%')
    if application_site:
        clauses.append('application_site LIKE ?')
        params.append(f'%{application_site.strip()}%')
    if not clauses:
        rows = conn.execute('SELECT * FROM exposure_daily_usage').fetchall()
    else:
        sql = 'SELECT * FROM exposure_daily_usage WHERE '
        sql += ' AND '.join(clauses)
        rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_body_weight(age_group='成人'):
    """Query body weight from unified Excel (fallback to DB)."""
    if HAS_OPENPYXL:
        try:
            rows = _read_excel_sheet('体重参数')
            for row in rows:
                if row[0] and age_group.strip() in row[0]:
                    return {
                        'age_group': row[0],
                        'default_weight_kg': row[1],
                        'source': row[2],
                        'notes': row[3]
                    }
        except Exception:
            pass
    # Fallback to DB
    conn = get_db()
    rows = conn.execute(
        'SELECT * FROM exposure_body_weight WHERE age_group LIKE ?',
        (f'%{age_group.strip()}%',)
    ).fetchall()
    conn.close()
    if rows:
        return dict(rows[0])
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM exposure_body_weight WHERE age_group LIKE '%成人%'"
    ).fetchall()
    conn.close()
    return dict(rows[0]) if rows else {
        'age_group': '成人',
        'default_weight_kg': 60.0,
        'source': 'default'
    }


def query_risk_substances(substance_name='', trigger_keyword=''):
    rows = _read_excel_sheet('风险物质识别')
    if not rows:
        return []
    results = []
    for row in rows:
        name = row[1] or ''
        keywords = (row[2] or '').strip()
        limit_val = row[4] or ''
        ref = row[5] or ''
        desc = row[6] or ''

        if substance_name and substance_name.strip():
            if substance_name.strip() not in name:
                continue

        if trigger_keyword and trigger_keyword.strip():
            if trigger_keyword.strip() not in keywords:
                continue

        results.append({
            'risk_type': row[0] or '',
            'substance_name': name,
            'trigger_keywords': keywords,
            'trigger_all': row[3] or '',
            'limit': limit_val,
            'source': ref,
            'description': desc,
        })
    return results


def query_authority_opinions(name_zh='', inci_name=''):
    rows = _read_excel_sheet('权威评估意见')
    if not rows:
        return []
    results = []
    for row in rows:
        name = row[0] or ''
        inci = row[1] or ''
        if name_zh and name_zh.strip() not in name:
            continue
        if inci_name and inci_name.strip() not in inci:
            continue
        results.append({
            'name_zh': name,
            'name_inci': inci,
            'cas_no': row[2] or '',
            'authority': row[3] or '',
            'opinion': row[4] or '',
            'reference': row[5] or '',
        })
    return results


def init_product_categories():
    conn = get_db()
    count = conn.execute(
        'SELECT COUNT(*) FROM product_categories'
    ).fetchone()[0]
    if count > 0:
        conn.close()
        return

    default_categories = [
        ('toner', '化妆水', 'skincare', '普通', 0, 0, '面部', '', 1),
        ('lotion', '乳液', 'skincare', '普通', 0, 0, '面部', '', 2),
        ('cream', '面霜/日霜', 'skincare', '普通', 0, 0, '面部', '', 3),
        ('essence', '精华液', 'skincare', '普通', 0, 0, '面部', '', 4),
        ('cleanser', '洗面奶', 'skincare', '普通', 1, 0, '面部', '', 5),
        ('makeup_remover', '卸妆液/油', 'skincare', '普通', 1, 0, '面部', '', 6),
        ('mask', '面膜', 'skincare', '普通', 0, 0, '面部', '', 7),
        ('sunscreen', '防晒霜', 'skincare', '特殊', 0, 0, '全身皮肤', '', 8),
        ('eye_cream', '眼霜/眼胶', 'skincare', '普通', 0, 0, '眼部', '', 9),
        ('hand_cream', '护手霜', 'skincare', '普通', 0, 0, '手部/足部', '', 10),
        ('body_lotion', '体乳/体霜', 'skincare', '普通', 0, 0, '全身皮肤', '', 11),
        ('body_wash', '沐浴露', 'skincare', '普通', 1, 0, '全身皮肤', '', 12),
        ('shampoo', '洗发水', 'haircare', '普通', 1, 0, '头发', '', 20),
        ('conditioner', '护发素', 'haircare', '普通', 1, 0, '头发', '', 21),
        ('hair_mask', '发膜', 'haircare', '普通', 0, 0, '头发', '', 22),
        ('hair_oil', '护发精油', 'haircare', '普通', 0, 0, '头发', '', 23),
        ('hair_dye_oxidative', '染发剂（氧化型）',
         'haircare', '特殊', 0, 0, '头发', '', 24),
        ('hair_dye_nonoxidative', '染发剂（非氧化型）',
         'haircare', '特殊', 0, 0, '头发', '', 25),
        ('foundation', '粉底粉', 'color', '普通', 0, 0, '面部', '', 30),
        ('loose_powder', '散粉/蜜粉', 'color', '普通', 0, 0, '面部', '', 31),
        ('lipstick', '口红', 'color', '普通', 0, 0, '嘴部', '', 32),
        ('lip_gloss', '唇彩/唇釉', 'color', '普通', 0, 0, '嘴部', '', 33),
        ('eye_shadow', '眼影', 'color', '普通', 0, 0, '眼部', '', 34),
        ('eyeliner', '眼线笔/液', 'color', '普通', 0, 0, '眼部', '', 35),
        ('mascara', '睫毛膏', 'color', '普通', 0, 0, '眼部', '', 36),
        ('eyebrow', '眉笔/眉粉', 'color', '普通', 0, 0, '眼部', '', 37),
        ('toothpaste_adult', '牙膏（成人）', 'mouth', '普通', 1, 0, '口腔', '', 40),
        ('toothpaste_child', '牙膏（儿童）', 'mouth', '普通', 1, 0, '口腔', '', 41),
        ('soap', '香皂', 'body', '普通', 1, 0, '全身皮肤', '', 50),
        ('deodorant', '止汗露/香体露', 'body', '普通', 0, 0, '全身皮肤', '', 51),
    ]

    for cat in default_categories:
        conn.execute('''
            INSERT INTO product_categories
            (code, name_zh, category_group, product_type, is_rinsed,
             is_professional, application_site, description, sort_order)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', cat)
    conn.commit()
    conn.close()


def get_all_product_categories(only_active=True):
    conn = get_db()
    if only_active:
        rows = conn.execute('''
            SELECT * FROM product_categories
            WHERE is_active = 1
            ORDER BY sort_order, name_zh
        ''').fetchall()
    else:
        rows = conn.execute('''
            SELECT * FROM product_categories
            ORDER BY sort_order, name_zh
        ''').fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_product_category_by_id(cat_id):
    conn = get_db()
    row = conn.execute(
        'SELECT * FROM product_categories WHERE id = ?', (cat_id,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def get_product_category_by_code(code):
    conn = get_db()
    row = conn.execute(
        'SELECT * FROM product_categories WHERE code = ?', (code,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def get_product_category_by_name(name_zh):
    conn = get_db()
    row = conn.execute(
        'SELECT * FROM product_categories WHERE name_zh = ?', (name_zh,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def get_product_categories_by_group(group):
    conn = get_db()
    rows = conn.execute('''
        SELECT * FROM product_categories
        WHERE category_group = ? AND is_active = 1
        ORDER BY sort_order, name_zh
    ''', (group,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]
