from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import os
import sys
import time
import requests
import json
import traceback
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def get_output_directory():
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        output_dir = filedialog.askdirectory(
            title="选择保存目录",
            initialdir=os.path.dirname(os.path.abspath(__file__))
        )
        root.destroy()
        if output_dir:
            return output_dir
    except Exception as e:
        print(f"打开文件对话框失败: {e}")
    default_dir = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "化妆品原料目录"
    )
    os.makedirs(default_dir, exist_ok=True)
    return default_dir


def check_network():
    try:
        response = requests.get("https://hzpsys.nifdc.org.cn", timeout=10)
        return response.status_code == 200
    except Exception:
        return False


def create_local_template(output_dir):
    columns = ['序号', '中文名称', 'INCI名称/英文名称', '备注']
    wb = Workbook()
    ws = wb.active

    ws.append(columns)

    for col in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fgColor="4472C4")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border

    column_widths = [10, 20, 40, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    excel_path = os.path.join(output_dir, "已使用化妆品原料目录I.xlsx")
    wb.save(excel_path)
    print(f"已创建空模板，保存至: {excel_path}")
    return excel_path


def crawl_with_edge(output_dir=None, progress_callback=None):
    if output_dir is None:
        output_dir = os.path.join(get_base_path(), "化妆品原料目录")

    print("开始爬取《已使用化妆品原料目录》I...")

    edge_options = Options()
    edge_options.add_argument('--headless')
    edge_options.add_argument('--disable-gpu')
    edge_options.add_argument('--no-sandbox')
    edge_options.add_argument(
        '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/91.0.4472.124 Safari/537.36'
    )
    edge_options.add_argument('--ignore-certificate-errors')
    edge_options.add_argument('--disable-extensions')
    edge_options.add_argument('--disable-popup-blocking')
    edge_options.add_argument('--window-size=1920,1080')

    driver = None
    try:
        print("尝试初始化Edge浏览器..")
        driver = webdriver.Edge(options=edge_options)
        print("Edge浏览器初始化成功")

        url = "https://hzpsys.nifdc.org.cn/hzpGS/ysyhzpylmla"
        print(f"访问页面: {url}")

        driver.set_page_load_timeout(60)
        driver.set_script_timeout(60)
        driver.get(url)

        wait = WebDriverWait(driver, 30)
        wait.until(EC.presence_of_element_located((By.ID, "searchInfo")))
        print("页面加载完成")

        try:
            search_button = wait.until(
                EC.element_to_be_clickable((By.ID, "searchInfo"))
            )
            search_button.click()
            print("点击查询按钮")
            time.sleep(3)
        except Exception as e:
            print(f"点击查询按钮失败: {e}")

        try:
            print("尝试设置每页显示100条记录..")
            page_list = driver.find_element(By.CLASS_NAME, "page-list")
            dropdown = page_list.find_element(By.CLASS_NAME, "dropdown")
            dropdown_button = dropdown.find_element(By.TAG_NAME, "button")
            dropdown_button.click()
            time.sleep(1)

            menu_items = dropdown.find_elements(By.TAG_NAME, "li")
            for item in menu_items:
                link = item.find_element(By.TAG_NAME, "a")
                if link.text == "100":
                    link.click()
                    print("已设置每页显示100条记录")
                    time.sleep(3)
                    break
        except Exception as e:
            print(f"设置每页显示记录数失败: {e}")

        all_data = []
        max_pages = 90

        try:
            wait.until(EC.presence_of_element_located((By.CLASS_NAME, "pagination")))
            time.sleep(2)

            pagination_info = driver.find_element(By.CLASS_NAME, "pagination-info")
            pagination_text = pagination_info.text
            print(f"分页信息: {pagination_text}")

            if "总共" in pagination_text:
                total_match = re.search(r'总共 (\d+) 条记录', pagination_text)
                if total_match:
                    total_records = int(total_match.group(1))
                    max_pages = (total_records + 99) // 100
                    print(f"总记录数: {total_records}, 总页数: {max_pages}")
        except Exception as e:
            print(f"获取总页数失败: {e}")

        if progress_callback:
            progress_callback(0, max_pages, 0)

        for page in range(1, max_pages + 1):
            print(f"\n处理第 {page}/{max_pages} 页..")

            success = False
            retries = 3

            while retries > 0 and not success:
                try:
                    wait.until(EC.presence_of_element_located((By.ID, "dg")))
                    time.sleep(2)

                    print("滚动加载数据...")
                    for i in range(3):
                        driver.execute_script(
                            "window.scrollTo(0, document.body.scrollHeight);"
                        )
                        time.sleep(1)

                    page_source = driver.page_source
                    soup = BeautifulSoup(page_source, 'html.parser')
                    table = soup.find('table', id='dg')

                    if table:
                        rows = table.find_all('tr')
                        page_data = []

                        for row in rows:
                            cols = row.find_all('td')
                            if len(cols) >= 4:
                                item = {
                                    '序号': cols[0].text.strip(),
                                    '中文名称': cols[1].text.strip(),
                                    'INCI名称/英文名称': cols[2].text.strip(),
                                    '备注': cols[3].text.strip()
                                }
                                page_data.append(item)

                        if page_data:
                            all_data.extend(page_data)
                            print(
                                f"第 {page} 页获取到 {len(page_data)} 条记录，"
                                f"累计 {len(all_data)} 条"
                            )
                        else:
                            print(f"第 {page} 页未找到数据")
                    else:
                        print(f"第 {page} 页未找到表格")

                    success = True

                except Exception as e:
                    retries -= 1
                    print(f"第 {page} 页处理失败，剩余重试次数: {retries}")
                    print(f"错误详情: {str(e)}")
                    if retries > 0:
                        print("等待5秒后重试...")
                        time.sleep(5)
                    else:
                        print(f"第 {page} 页重试失败，跳过此页继续")
                        break

            if progress_callback:
                progress_callback(page, max_pages, len(all_data))

            if not success:
                continue

            if page < max_pages:
                next_success = False
                next_retries = 3

                while next_retries > 0 and not next_success:
                    try:
                        pagination = wait.until(
                            EC.presence_of_element_located(
                                (By.CLASS_NAME, "pagination")
                            )
                        )

                        if pagination:
                            next_button = pagination.find_element(
                                By.CLASS_NAME, "page-next"
                            )
                            if next_button:
                                next_link = next_button.find_element(
                                    By.TAG_NAME, "a"
                                )
                                driver.execute_script(
                                    "arguments[0].click();", next_link
                                )
                                print("点击下一页按钮")
                                time.sleep(3)
                                next_success = True
                            else:
                                print("未找到下一页链接")
                                next_retries = 0
                        else:
                            print("未找到分页元素")
                            next_retries = 0

                    except Exception as e:
                        next_retries -= 1
                        print(f"分页操作失败，剩余重试次数: {next_retries}")
                        print(f"错误详情: {str(e)}")
                        if next_retries > 0:
                            print("等待5秒后重试...")
                            time.sleep(5)
                        else:
                            print("分页重试失败，终止爬取")

                if not next_success:
                    print("无法继续分页，终止爬取")
                    break

        if all_data:
            wb = Workbook()
            ws = wb.active

            columns = ['序号', '中文名称', 'INCI名称/英文名称', '备注']
            ws.append(columns)

            for col in range(1, len(columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fgColor="4472C4")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border

            for i, item in enumerate(all_data, start=2):
                row = [
                    item.get('序号', ''),
                    item.get('中文名称', ''),
                    item.get('INCI名称/英文名称', ''),
                    item.get('备注', '')
                ]
                ws.append(row)

                for col in range(1, len(columns) + 1):
                    cell = ws.cell(row=i, column=col)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border

            column_widths = [10, 20, 40, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            excel_path = os.path.join(output_dir, "已使用化妆品原料目录I.xlsx")
            wb.save(excel_path)
            print(f"\n目录I数据已保存为Excel，共 {len(all_data)} 条记录")

            catalog = []
            for item in all_data:
                catalog.append({
                    'seq': item.get('序号', ''),
                    'cn_name': item.get('中文名称', ''),
                    'inci_name': item.get('INCI名称/英文名称', ''),
                    'remark': item.get('备注', '')
                })

            userdata_dir = os.path.join(get_base_path(), "userdata")
            os.makedirs(userdata_dir, exist_ok=True)
            json_path = os.path.join(userdata_dir, "ingredient_catalog.json")
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(catalog, f, ensure_ascii=False, indent=2)
            print(f"数据已保存到: {json_path}")
            print(f"JSON共 {len(catalog)} 条记录")

            backup_json_path = os.path.join(output_dir, "ingredient_catalog.json")
            with open(backup_json_path, 'w', encoding='utf-8') as f:
                json.dump(catalog, f, ensure_ascii=False, indent=2)
            print(f"备份已保存到: {backup_json_path}")
        else:
            print("未找到表格数据，创建本地模板...")
            create_local_template(output_dir)

    except Exception as e:
        print(f"爬取失败: {e}")
        traceback.print_exc()
        print("创建本地模板...")
        create_local_template(output_dir)
    finally:
        if 'driver' in locals():
            try:
                driver.quit()
            except Exception:
                pass


if __name__ == "__main__":
    print("开始使用Edge浏览器爬取《已使用化妆品原料目录》I...")

    output_dir = None
    if len(sys.argv) > 1:
        output_dir = sys.argv[1]

    crawl_with_edge(output_dir)
    print("爬取完成")
