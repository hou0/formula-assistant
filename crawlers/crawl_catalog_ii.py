"""
《已使用化妆品原料目录》II 数据抓取脚本
使用Edge浏览器自动化抓取目录II的数据并导出为Excel
"""

from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import os
import time
import re
import sys
import requests
import base64
import json
import traceback
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


def get_base_path():
    """获取应用基础路径，兼容开发环境和打包环境"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


base_path = get_base_path()


def get_output_directory():
    """获取输出目录，支持用户选择"""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        output_dir = filedialog.askdirectory(
            title="选择保存目录",
            initialdir=base_path
        )
        root.destroy()
        if output_dir:
            return output_dir
    except Exception as e:
        print(f"打开文件对话框失败: {e}")
    default_dir = os.path.join(base_path, "化妆品原料目录")
    os.makedirs(default_dir, exist_ok=True)
    return default_dir


def create_local_template(output_dir):
    """创建本地空模板"""
    excel_path = os.path.join(output_dir, "已使用化妆品原料目录II.xlsx")

    wb = Workbook()
    ws = wb.active

    headers = [
        '序号', '中文名称', 'INCI名称', '英文名称', 'CAS号', '原料分子式',
        '化学结构式', '相对分子量', '使用目的', '安全使用量', '安全使用量', '备注'
    ]
    ws.append(headers)

    sub_headers = ['', '', '', '', '', '', '', '', '', '使用范围', '安全使用量', '']
    ws.append(sub_headers)

    for col in range(1, 9):
        ws.merge_cells(
            start_row=1,
            start_column=col,
            end_row=2,
            end_column=col
        )

    ws.merge_cells(start_row=1, start_column=9, end_row=2, end_column=9)
    ws.merge_cells(start_row=1, start_column=12, end_row=2, end_column=12)
    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=11)

    ws.cell(row=1, column=10).alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 15

    wb.save(excel_path)
    print(f"已创建空模板，保存至: {excel_path}")
    return excel_path


def crawl_catalog_ii(output_dir=None, progress_callback=None):
    """抓取目录II的数据"""
    if output_dir is None:
        output_dir = os.path.join(base_path, "化妆品原料目录")

    print("=" * 60)
    print("开始抓取《已使用化妆品原料目录》II")
    print("=" * 60)

    edge_options = Options()
    edge_options.add_argument('--headless')
    edge_options.add_argument('--disable-gpu')
    edge_options.add_argument('--no-sandbox')
    edge_options.add_argument(
        '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/91.0.4472.124 Safari/537.36'
    )
    edge_options.add_argument('--ignore-certificate-errors')
    edge_options.add_argument('--disable-extensions')
    edge_options.add_argument('--disable-popup-blocking')
    edge_options.add_argument('--window-size=1920,1080')

    driver = None
    try:
        print("正在初始化Edge浏览器...")
        driver = webdriver.Edge(options=edge_options)
        print("Edge浏览器初始化成功")

        url = "https://hzpsys.nifdc.org.cn/hzpGS/ysyhzpylmlb"
        print(f"访问页面: {url}")

        driver.set_page_load_timeout(60)
        driver.set_script_timeout(60)

        driver.get(url)

        wait = WebDriverWait(driver, 30)
        wait.until(EC.presence_of_element_located((By.ID, "searchInfo")))
        print("页面加载完成")

        try:
            search_button = wait.until(EC.element_to_be_clickable(
                (By.ID, "searchInfo")
            ))
            search_button.click()
            print("点击查询按钮")
            time.sleep(3)
        except Exception as e:
            print(f"点击查询按钮失败: {e}")

        try:
            print("尝试设置每页显示100条记录..")
            page_list = driver.find_element(By.CLASS_NAME, "page-list")
            dropdown = page_list.find_element(By.CLASS_NAME, "dropdown")
            dropdown_button = dropdown.find_element(By.TAG_NAME, "button")
            dropdown_button.click()
            time.sleep(1)

            menu_items = dropdown.find_elements(By.TAG_NAME, "li")
            for item in menu_items:
                link = item.find_element(By.TAG_NAME, "a")
                if link.text == "100":
                    link.click()
                    print("已设置每页显示100条记录")
                    time.sleep(3)
                    break
        except Exception as e:
            print(f"设置每页显示记录数失败: {e}")

        all_data = []
        max_pages = 1

        try:
            wait.until(EC.presence_of_element_located(
                (By.CLASS_NAME, "pagination")
            ))
            time.sleep(2)

            pagination_info = driver.find_element(
                By.CLASS_NAME, "pagination-info"
            )
            pagination_text = pagination_info.text
            print(f"分页信息: {pagination_text}")

            if "总共" in pagination_text:
                total_match = re.search(r'总共 (\d+) 条记录', pagination_text)
                if total_match:
                    total_records = int(total_match.group(1))
                    max_pages = (total_records + 99) // 100
                    print(f"总记录数: {total_records}, 总页数: {max_pages}")
        except Exception as e:
            print(f"获取总页数失败: {e}")
            max_pages = 10

        if progress_callback:
            progress_callback(0, max_pages, 0)

        for page in range(1, max_pages + 1):
            print(f"\n处理第 {page}/{max_pages} 页..")

            wait.until(EC.presence_of_element_located(
                (By.ID, "dg")
            ))
            time.sleep(2)

            print("滚动加载数据...")
            for i in range(3):
                driver.execute_script(
                    "window.scrollTo(0, document.body.scrollHeight);"
                )
                time.sleep(1)

            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            table = soup.find('table', id='dg')

            if table:
                rows = table.find_all('tr')
                page_data = []

                for row in rows:
                    cols = row.find_all('td')
                    if len(cols) >= 12:
                        chemical_structure = cols[6].text.strip()
                        chemical_structure_img = None
                        img_tag = cols[6].find('img')

                        if img_tag:
                            if 'src' in img_tag.attrs:
                                chemical_structure_img = img_tag['src']
                                print(
                                    f"找到图片src: {chemical_structure_img[:100]}..."
                                )
                            elif 'data-src' in img_tag.attrs:
                                chemical_structure_img = img_tag['data-src']
                                print(
                                    f"找到图片data-src: {chemical_structure_img[:100]}..."
                                )
                            else:
                                print(f"img标签属性: {img_tag.attrs}")

                            if chemical_structure_img:
                                if chemical_structure_img.startswith('data:image'):
                                    print("  -> 这是base64编码的内联图片")
                                elif not chemical_structure_img.startswith('http'):
                                    chemical_structure_img = (
                                        'https://hzpsys.nifdc.org.cn' +
                                        chemical_structure_img
                                    )
                        else:
                            if chemical_structure:
                                print(
                                    f"化学结构式文本内容: {chemical_structure[:50]}..."
                                )

                        item = {
                            '序号': cols[0].text.strip(),
                            '中文名称': cols[1].text.strip(),
                            'INCI名称': cols[2].text.strip(),
                            '英文名称': cols[3].text.strip(),
                            'CAS号': cols[4].text.strip(),
                            '原料分子式': cols[5].text.strip(),
                            '化学结构式': chemical_structure,
                            '化学结构式图': chemical_structure_img,
                            '相对分子量': cols[7].text.strip(),
                            '使用目的': cols[8].text.strip(),
                            '使用范围': cols[9].text.strip(),
                            '安全使用量': cols[10].text.strip(),
                            '备注': cols[11].text.strip()
                        }
                        page_data.append(item)

                if page_data:
                    all_data.extend(page_data)
                    print(
                        "第 {} 页获取到 {} 条记录，累计 {} 条"
                        .format(page, len(page_data), len(all_data))
                    )
                else:
                    print(f"第 {page} 页未找到数据")

                if progress_callback:
                    progress_callback(page, max_pages, len(all_data))
            else:
                print(f"第 {page} 页未找到表格")

            if page < max_pages:
                try:
                    pagination = driver.find_element(
                        By.CLASS_NAME, "pagination"
                    )
                    if pagination:
                        next_button = pagination.find_element(
                            By.CLASS_NAME, "page-next"
                        )
                        if next_button:
                            if 'disabled' in next_button.get_attribute('class'):
                                print("已到达最后一页")
                                break
                            next_link = next_button.find_element(
                                By.TAG_NAME, "a"
                            )
                            next_link.click()
                            print("点击下一页按钮")
                            time.sleep(3)
                        else:
                            print("未找到下一页链接")
                            break
                    else:
                        print("未找到分页元素")
                        break
                except Exception as e:
                    print(f"分页操作失败: {e}")
                    break

        if all_data:
            img_dir = os.path.join(output_dir, "化学结构式图片")
            os.makedirs(img_dir, exist_ok=True)

            img_count = len(
                [i for i in all_data if i.get('化学结构式图')]
            )
            print(
                f"开始下载化学结构式图片，共 {img_count} 张图片..."
            )
            downloaded_count = 0
            failed_count = 0

            for item in all_data:
                if item.get('化学结构式图'):
                    try:
                        img_url = item['化学结构式图']
                        safe_name = "".join(
                            c for c in item['中文名称']
                            if c.isalnum() or c in (' ', '-', '_')
                        ).strip()
                        img_name = f"{item['序号']}_{safe_name}.png"
                        img_path = os.path.join(img_dir, img_name)

                        if os.path.exists(img_path):
                            print(f"图片已存在，跳过: {img_name}")
                            item['化学结构式图片路径'] = img_path
                            downloaded_count += 1
                            continue

                        if img_url.startswith('data:image'):
                            print(f"处理base64图片: {img_name}")
                            header, base64_data = img_url.split(',', 1)
                            img_data = base64.b64decode(base64_data)
                            with open(img_path, 'wb') as f:
                                f.write(img_data)
                            item['化学结构式图片路径'] = img_path
                            print(f"[OK] 已保存base64图片: {img_name}")
                            downloaded_count += 1
                        else:
                            print(f"正在下载: {img_url}")
                            response = requests.get(
                                img_url,
                                timeout=30,
                                headers={
                                    'User-Agent': (
                                        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                                        'AppleWebKit/537.36'
                                    )
                                }
                            )
                            if response.status_code == 200:
                                with open(img_path, 'wb') as f:
                                    f.write(response.content)
                                item['化学结构式图片路径'] = img_path
                                print(f"[OK] 已下载: {img_name}")
                                downloaded_count += 1
                            else:
                                print(
                                    f"[ERR] 下载失败 (状态码 {response.status_code}): "
                                    f"{img_url}"
                                )
                                failed_count += 1
                    except Exception as e:
                        print(f"[ERR] 处理图片时出错: {e}")
                        failed_count += 1

            print(
                f"\n图片下载完成: 成功 {downloaded_count} 张, "
                f"失败 {failed_count} 张"
            )

            wb = Workbook()
            ws = wb.active

            headers = [
                '序号', '中文名称', 'INCI名称', '英文名称', 'CAS号', '原料分子式',
                '化学结构式', '相对分子量', '使用目的', '安全使用量', '安全使用量', '备注'
            ]
            ws.append(headers)

            sub_headers = [
                '', '', '', '', '', '', '', '', '',
                '使用范围', '安全使用量', ''
            ]
            ws.append(sub_headers)

            for col in range(1, 9):
                ws.merge_cells(
                    start_row=1,
                    start_column=col,
                    end_row=2,
                    end_column=col
                )

            ws.merge_cells(
                start_row=1,
                start_column=9,
                end_row=2,
                end_column=9
            )

            ws.merge_cells(
                start_row=1,
                start_column=12,
                end_row=2,
                end_column=12
            )

            ws.merge_cells(
                start_row=1,
                start_column=10,
                end_row=1,
                end_column=11
            )

            for col in range(1, 13):
                if col != 10 and col != 11:
                    ws.cell(row=1, column=col).alignment = Alignment(
                        horizontal='center',
                        vertical='center'
                    )
            ws.cell(row=1, column=10).alignment = Alignment(
                horizontal='center',
                vertical='center'
            )

            for r_idx, item in enumerate(all_data, start=3):
                row = [
                    item.get('序号', ''),
                    item.get('中文名称', ''),
                    item.get('INCI名称', ''),
                    item.get('英文名称', ''),
                    item.get('CAS号', ''),
                    item.get('原料分子式', ''),
                    item.get('化学结构式', ''),
                    item.get('相对分子量', ''),
                    item.get('使用目的', ''),
                    item.get('使用范围', ''),
                    item.get('安全使用量', ''),
                    item.get('备注', '')
                ]
                ws.append(row)

                for col in range(1, 13):
                    cell = ws.cell(row=r_idx, column=col)
                    cell.alignment = Alignment(
                        vertical='center',
                        wrap_text=True
                    )

                if (
                    item.get('化学结构式图片路径')
                    and os.path.exists(item['化学结构式图片路径'])
                ):
                    try:
                        img = Image(item['化学结构式图片路径'])
                        img.width = 150
                        img.height = 100
                        cell = ws.cell(row=r_idx, column=7)
                        ws.add_image(img, get_column_letter(7) + str(r_idx))
                        ws.row_dimensions[r_idx].height = 120
                    except Exception as e:
                        print(f"插入图片时出错: {e}")

            for col in range(1, 13):
                if col == 7:
                    ws.column_dimensions[get_column_letter(col)].width = 25
                else:
                    ws.column_dimensions[get_column_letter(col)].width = 15

            catalog = []
            for item in all_data:
                catalog.append({
                    'seq': item.get('序号', ''),
                    'cn_name': item.get('中文名称', ''),
                    'inci_name': item.get('INCI名称', ''),
                    'en_name': item.get('英文名称', ''),
                    'cas': item.get('CAS号', ''),
                    'formula': item.get('原料分子式', ''),
                    'structure': item.get('化学结构式', ''),
                    'mol_weight': item.get('相对分子量', ''),
                    'use_purpose': item.get('使用目的', ''),
                    'use_range': item.get('使用范围', ''),
                    'safe_amount': item.get('安全使用量', ''),
                    'remark': item.get('备注', '')
                })

            userdata_dir = os.path.join(base_path, "userdata")
            os.makedirs(userdata_dir, exist_ok=True)
            json_path_ii = os.path.join(
                userdata_dir, "ingredient_catalog_ii.json"
            )
            with open(json_path_ii, 'w', encoding='utf-8') as f:
                json.dump(catalog, f, ensure_ascii=False, indent=2)
            print(f"数据已保存到: {json_path_ii}")
            print(f"JSON共 {len(catalog)} 条记录")

            backup_json_path = os.path.join(
                output_dir, "ingredient_catalog_ii.json"
            )
            with open(backup_json_path, 'w', encoding='utf-8') as f:
                json.dump(catalog, f, ensure_ascii=False, indent=2)
            print(f"备份已保存到: {backup_json_path}")

            try:
                excel_path = os.path.join(
                    output_dir, "已使用化妆品原料目录II.xlsx"
                )
                wb.save(excel_path)
                print(f"\n目录II数据已保存为Excel，共 {len(all_data)} 条记录")
                print(f"保存路径: {os.path.abspath(excel_path)}")
                print(f"化学结构式图片保存至: {os.path.abspath(img_dir)}")
            except Exception as e:
                print(f"Excel保存失败（JSON已保存）: {e}")

            return True
        else:
            print("未找到表格数据，创建本地模板...")
            create_local_template(output_dir)
            return False
    except Exception as e:
        print(f"爬取失败: {e}")
        traceback.print_exc()
        if 'all_data' in dir() and all_data:
            try:
                catalog = []
                for item in all_data:
                    catalog.append({
                        'seq': item.get('序号', ''),
                        'cn_name': item.get('中文名称', '')
                    })
                userdata_dir = os.path.join(base_path, "userdata")
                os.makedirs(userdata_dir, exist_ok=True)
                with open(
                    os.path.join(userdata_dir, "ingredient_catalog_ii.json"),
                    'w',
                    encoding='utf-8'
                ) as f:
                    json.dump(catalog, f, ensure_ascii=False, indent=2)
                print(f"异常恢复：已保存 {len(catalog)} 条数据")
            except Exception as e2:
                print(f"异常恢复保存失败: {e2}")
        print("创建本地模板...")
        create_local_template(output_dir)
        return False
    finally:
        if 'driver' in locals():
            try:
                driver.quit()
                print("浏览器已关闭")
            except Exception as e:
                print(f"关闭浏览器失败: {e}")


def main():
    """主函数"""
    print("\n" + "=" * 60)
    print("《已使用化妆品原料目录》II 数据抓取工具")
    print("=" * 60 + "\n")

    try:
        __import__('selenium')
        __import__('bs4')
        __import__('openpyxl')
        __import__('requests')
        print("[OK] 所有依赖库检查通过")
    except ImportError as e:
        print(f"[ERR] 缺少必要的库: {e}")
        print(
            "请运行: pip install selenium beautifulsoup4 openpyxl requests"
        )
        input("\n按回车键退出..")
        sys.exit(1)

    output_dir = None
    if len(sys.argv) > 1:
        output_dir = sys.argv[1]
        print(f"使用命令行指定的输出目录: {output_dir}\n")
    else:
        output_dir = get_output_directory()
        print(f"保存目录: {output_dir}\n")

    success = crawl_catalog_ii(output_dir)

    print("\n" + "=" * 60)
    if success:
        print("[OK] 抓取完成！")
    else:
        print("[ERR] 抓取失败，已创建空模板")
    print("=" * 60)


if __name__ == "__main__":
    main()
